#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import math
import openpyxl
import os
from build123d import *

EXCEL_PATH = os.path.expanduser("~/Desktop/3x5+3_low_clipped(1).xlsx")

# ── Read Excel data ──────────────────────────────────────────────────────────
# Check if file exists, otherwise skip Excel loading and use defaults
wb = None 
if os.path.exists(EXCEL_PATH):
    wb = openpyxl.load_workbook(EXCEL_PATH)
else:
    print(f"Warning: Excel file not found at {EXCEL_PATH}, using minimal setup")

def _dedupe(pts, tol=0.05):
    """Remove consecutive duplicate points (within tolerance) and ensure the
    last point is not a near-duplicate of the first (avoids zero-length edges)."""
    out = [pts[0]]
    for p in pts[1:]:
        if math.hypot(p[0] - out[-1][0], p[1] - out[-1][1]) >= tol:
            out.append(p)
    # Remove last point if it is essentially the same as the first (polygon already closed)
    if math.hypot(out[-1][0] - out[0][0], out[-1][1] - out[0][1]) < tol:
        out.pop()
    return out

def _read_periphery_xz(sheet_name):
    """Return deduplicated list of (X, Z) tuples from a periphery sheet."""
    if wb is None:
        return []
    pts = []
    for row in wb[sheet_name].iter_rows(values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        pts.append((float(row[1]), float(row[3])))   # (X_mm, Z_mm)
    return _dedupe(pts)

def _read_circles():
    """Return list of (center_x, center_z, radius) from the Circles sheet."""
    if wb is None:
        return []
    ws = wb['Circles']
    result = []
    in_definitions = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Circle_ID' and row[1] == 'Center_X_mm':
            in_definitions = True
            continue
        if not in_definitions:
            continue
        if row[0] == 'Circle Vertex Data' or not isinstance(row[0], (int, float)):
            break
        result.append((float(row[1]), float(row[3]), float(row[4])))   # (cx, cz, r)
    return result

outer_pts = _read_periphery_xz('Outer Periphery')
inner_pts = _read_periphery_xz('Inner Periphery')
circles   = _read_circles()

footprint_solid = None
# ── Create a clipping footprint based on the outer periphery ────────────────
if len(outer_pts) > 1:
    with BuildPart() as footprint_builder:
        with BuildSketch(Plane.XZ) as _fs:
            with BuildLine():
                Polyline(*outer_pts, close=True)
            make_face()
        extrude(amount=5000, both=True)
    footprint_solid = footprint_builder.part
    print("Clipping footprint solid created.")
else:
    print("Skipping footprint: no outer periphery data")

print(f"Outer periphery : {len(outer_pts)} points")
print(f"Inner periphery : {len(inner_pts)} points")
print(f"Circles         : {len(circles)}")
for i, (cx, cz, r) in enumerate(circles, 1):
    print(f"  Circle {i}: center=({cx}, {cz}), r={r}")

# ── Build profile and extrude ────────────────────────────────────────────────
# Work plane at Y=-80.  2D (u, v) maps to world (X, Z).
# Plane normal = -Y → extrude amount=-40 travels in +Y → Y=-80 to Y=-40.
_wp = Plane(origin=(0, -80.0, 0), x_dir=(1, 0, 0), z_dir=(0, -1, 0))

with BuildPart() as snap_fit:
    with BuildSketch(_wp) as _sk:
        # ── Outer periphery face ─────────────────────────────────────────
        with BuildLine():
            Polyline(*outer_pts, close=True)
        # Robustness: offset the footprint slightly outwards
        
        make_face()

        # ── Subtract inner periphery ─────────────────────────────────────
        with BuildSketch(_wp, mode=Mode.PRIVATE) as _inner_priv:
            with BuildLine():
                Polyline(*inner_pts, close=True)
            make_face()
        add(_inner_priv.sketch, mode=Mode.SUBTRACT)

        # ── Subtract circles ─────────────────────────────────────────────
        _circle_locs = [(cx, cz) for cx, cz, _ in circles]
        with Locations(*_circle_locs):
            Circle(circles[0][2], mode=Mode.SUBTRACT)   # all circles same r=29 mm

    # Extrude from Y=-80 to Y=-40 (40 mm in +Y)
    extrude(amount=-40.0)
    print("Extrusion complete: Y=-80 → Y=-40")

part = snap_fit.part
if part:
    print(f"Snap-fit Is Manifold : {part.is_manifold}")
    print(f"Snap-fit Volume      : {part.volume:.2f} mm³")
else:
    print("Error: No snap-fit geometry created.")

# ── Lofted sweep profile ──────────────────────────────────────────────────────
# Start profile: 13 points joined in order (given coordinates)
_sw_start = [
    ( 78.75373363,  70.94733715,  144.49463844),
    ( 75.678339,    25.43191195,  -26.41081572),
    ( 75.51553249,  21.94497108,  -52.23324776),
    ( 76.46360397,  20.38958788, -242.22452164),
    ( 76.88437462,  23.45331907, -268.09766769),
    ( 82.46365547,  73.26032162, -468.19126129),
    ( 82.65399933,  74.10850525, -488.26728821),
    ( 78.63717556,  30.67662716, -462.12020874),
    ( 73.4980917,  -15.20066261, -277.81377792),
    ( 72.9142952,  -19.451437,   -241.91606522),
    ( 71.96622372, -17.89605379,  -51.9247961 ),
    ( 72.19211102, -13.05809498,  -16.09744906),
    ( 75.84316731,  40.97708702,  186.7987442 ),
]

# End diagonals define how each start point maps to the far (−X) face:
#   Line 1: _sw_start[0] → (-110.5930233, 86.33929253, 137.93326378)
#   Line 2: _sw_start[5] → (-106.51596069, 92.50883102, -478.90472412)
# All other end points are found by linearly interpolating the delta vector
# between these two Z anchors.
_z1 = _sw_start[0][2];  _z6 = _sw_start[5][2]
_d1 = (-110.5930233   - _sw_start[0][0],
        86.33929253   - _sw_start[0][1],
       137.93326378   - _sw_start[0][2])
_d6 = (-106.51596069  - _sw_start[5][0],
        92.50883102   - _sw_start[5][1],
       -478.90472412  - _sw_start[5][2])

def _end_pt(x, y, z):
    t = (z - _z1) / (_z6 - _z1)
    return (
        x + _d1[0] + t * (_d6[0] - _d1[0]),
        y + _d1[1] + t * (_d6[1] - _d1[1]),
        z + _d1[2] + t * (_d6[2] - _d1[2]),
    )

_sw_end = [_end_pt(*p) for p in _sw_start]

# Print end profile for verification
print("\nEnd profile (computed from diagonal interpolation):")
for i, (p_s, p_e) in enumerate(zip(_sw_start, _sw_end), 1):
    print(f"  P{i:2d}  start={tuple(f'{v:.4f}' for v in p_s)}  "
          f"end={tuple(f'{v:.4f}' for v in p_e)}")

# Build the closed loft solid:
# The two profiles are non-planar in 3D.  We:
#   1. Loft the two 3D wires → lateral shell (open ends).
#   2. Project each profile onto its Newell best-fit plane → two planar cap faces.
#   3. Sew lateral shell + caps together with BRepBuilderAPI_Sewing → closed solid.
from OCP.BRepBuilderAPI import BRepBuilderAPI_Sewing, BRepBuilderAPI_MakeFace as OCC_MF
from OCP.BRep import BRep_Builder
from OCP.TopoDS import TopoDS_Solid as OCP_Solid
from OCP.BRepPrimAPI import BRepPrimAPI_MakePrism
from OCP.gp import gp_Vec, gp_Pln, gp_Ax3, gp_Pnt, gp_Dir

def _make_planar_cap(pts3d):
    """Fit a Newell best-fit plane to pts3d and return a planar OCC Face."""
    n = len(pts3d)
    cx = sum(p[0] for p in pts3d) / n
    cy = sum(p[1] for p in pts3d) / n
    cz = sum(p[2] for p in pts3d) / n
    nx = ny = nz = 0.0
    for i in range(n):
        p1, p2 = pts3d[i], pts3d[(i + 1) % n]
        nx += (p1[1] - p2[1]) * (p1[2] + p2[2])
        ny += (p1[2] - p2[2]) * (p1[0] + p2[0])
        nz += (p1[0] - p2[0]) * (p1[1] + p2[1])
    length = math.sqrt(nx*nx + ny*ny + nz*nz)
    nx /= length;  ny /= length;  nz /= length
    wire = Wire.make_polygon([Vector(*p) for p in pts3d], close=True)
    fb   = OCC_MF(gp_Pln(gp_Ax3(gp_Pnt(cx, cy, cz), gp_Dir(nx, ny, nz))),
                  wire.wrapped, True)
    return Face(fb.Face())

_wire_start = Wire.make_polygon([Vector(*p) for p in _sw_start], close=True)
_wire_end   = Wire.make_polygon([Vector(*p) for p in _sw_end],   close=True)
_cap_start  = _make_planar_cap(_sw_start)
_cap_end    = _make_planar_cap(_sw_end)

# Lateral shell from wire loft
_lateral = Solid.make_loft([_wire_start, _wire_end])

# Sew lateral + two caps into a closed shell, then wrap as Solid
_sew = BRepBuilderAPI_Sewing(1e-2)
_sew.Add(_lateral.wrapped)
_sew.Add(_cap_start.wrapped)
_sew.Add(_cap_end.wrapped)
_sew.Perform()
_ocp_sol = OCP_Solid()
_bld_sol  = BRep_Builder()
_bld_sol.MakeSolid(_ocp_sol)
_bld_sol.Add(_ocp_sol, _sew.SewedShape())
_closed_solid = Solid(_ocp_sol)

with BuildPart() as loft_part:
    add(_closed_solid)

    # ── Three rectangular slot cuts (42 mm deep, both directions) ────────────
    _cut_profiles = [
        [(-88.2062912,  36.3257575,  -78.66913319),
         (-87.51270294, 35.18787146, -217.66275406),
         ( 50.93726635, 22.85993576, -216.87093735),
         ( 50.24367809, 23.99782181,  -77.87733555)],
        [(-86.27497673, 44.20006752, -293.77058029),
         (-82.51532555, 77.76292801, -428.60530853),
         ( 55.93464375, 65.43498993, -427.81349182),
         ( 52.17499733, 31.87213182, -292.97878265)],
        [( 50.72258472, 34.25492764,   -1.91867799),
         (-87.72738457, 46.58286095,   -2.71047771),
         (-85.3107357,  82.34900475,  131.58748627),
         ( 53.13923836, 70.02107143,  132.37928391)],
    ]

    for _ci, _cpts in enumerate(_cut_profiles, 1):
        _n = len(_cpts)
        _cx = sum(p[0] for p in _cpts) / _n
        _cy = sum(p[1] for p in _cpts) / _n
        _cz = sum(p[2] for p in _cpts) / _n
        _cc = Vector(_cx, _cy, _cz)
        _cnx = _cny = _cnz = 0.0
        for _i in range(_n):
            _p1, _p2 = _cpts[_i], _cpts[(_i + 1) % _n]
            _cnx += (_p1[1] - _p2[1]) * (_p1[2] + _p2[2])
            _cny += (_p1[2] - _p2[2]) * (_p1[0] + _p2[0])
            _cnz += (_p1[0] - _p2[0]) * (_p1[1] + _p2[1])
        _cnorm  = Vector(_cnx, _cny, _cnz).normalized()
        _cref   = Vector(0, 0, 1) if abs(_cnorm.Z) < 0.9 else Vector(1, 0, 0)
        _cx_loc = _cnorm.cross(_cref).normalized()
        _cy_loc = _cnorm.cross(_cx_loc).normalized()
        _cproj  = [((Vector(*p) - _cc).dot(_cx_loc),
                    (Vector(*p) - _cc).dot(_cy_loc)) for p in _cpts]
        _cplane = Plane(origin=_cc, x_dir=_cx_loc, z_dir=_cnorm)
        with BuildSketch(_cplane, mode=Mode.PRIVATE) as _csk:
            with BuildLine():
                Polyline(*_cproj, close=True)
            make_face()
        extrude(_csk.sketch, amount=41.0, mode=Mode.SUBTRACT, both=True)
        print(f"Cut {_ci} applied")

lp = loft_part.part
if lp:
    print(f"Loft Is Manifold : {lp.is_manifold}")
    print(f"Loft Volume      : {lp.volume:.2f} mm³")
else:
    print("Error: No loft geometry created.")

# ── Swept solid: 12-point profile lofted along path translation ───────────────
# Line 1: (-142.115…) → (-329.173…)  |  Line 2: (-328.502…) → (-141.444…)
# Both lines share the same translation vector → uniform shift of all profile pts.
_sp_start = [
    (-134.31758881,  36.70519829,   60.55816174),
    (-142.11549759, -11.93355918, -143.58742714),
    (-142.54227638, -14.94356513, -169.46678162),
    (-141.44435883, -13.27223659, -359.45625305),
    (-140.71885109,  -9.80747104, -385.27198792),
    (-130.97485542,  43.88611317, -588.16360474),
    (-138.04813385,   0.60776316, -579.13894653),
    (-147.44586945, -47.84443855, -395.66123962),
    (-148.45247269, -52.65162945, -359.8431778 ),
    (-149.55039024, -54.32295799, -169.85370636),
    (-148.9582634,  -50.14672279, -133.94737244),
    (-140.92096329,  -0.17112618,   69.86096859),
]

# Translation = far-end of line 1 minus near-end of line 1
_sp_dx = -329.1734314   - (-142.11549759)   # -187.05793381
_sp_dy =   21.36377573  - (-11.93355918)    #  +33.29733491
_sp_dz = -144.37548637  - (-143.58742714)   #   -0.78805923

_sp_face   = _make_planar_cap(_sp_start)
_sp_prism  = BRepPrimAPI_MakePrism(_sp_face.wrapped, gp_Vec(_sp_dx, _sp_dy, _sp_dz))
_sp_prism.Build()
_sp_closed = Solid(_sp_prism.Shape())
print(f"Swept prism complete: is_manifold={_sp_closed.is_manifold}  volume={_sp_closed.volume:.2f}")

with BuildPart() as swept_part:
    add(_sp_closed)

# ── Swept solid 2: 12-point profile extruded along translation line ───────────
_sp2_start = [
    (-348.6554718,   35.57249069, -622.57381439),
    (-363.53610992, -16.40210629, -421.5953064 ),
    (-364.5608139,  -19.7904551,  -395.77957153),
    (-365.80024719, -21.35973454, -205.7901001 ),
    (-365.11276245, -18.3983779,  -179.91075516),
    (-350.77335358,  38.64121199,   63.07707787),
    (-360.54286957,   2.56469488,   72.358284  ),
    (-375.25997162, -55.86974144, -170.27069092),
    (-376.21383667, -59.97848034, -206.17702484),
    (-374.97440338, -58.40919971, -396.16649628),
    (-373.55266571, -53.70803356, -431.98455811),  # swapped: was pt 12
    (-373.55266571,  -6.25927269, -615.46226501),  # swapped: was pt 11
]

# Translation vector from line: (-364.561…, -19.790…, -395.780…) → (-548.005…, 29.683…, -396.568…)
_sp2_dx = -548.00487518 - (-364.5608139)   # -183.44406128
_sp2_dy =   29.68334675 - (-19.7904551)    #  +49.47380185
_sp2_dz = -396.56764984 - (-395.77957153)  #   -0.78807831

_sp2_face  = _make_planar_cap(_sp2_start)
_sp2_prism = BRepPrimAPI_MakePrism(_sp2_face.wrapped, gp_Vec(_sp2_dx, _sp2_dy, _sp2_dz))
_sp2_prism.Build()
_sp2_closed = Solid(_sp2_prism.Shape())
print(f"Swept2 prism complete: is_manifold={_sp2_closed.is_manifold}  volume={_sp2_closed.volume:.2f}")

with BuildPart() as swept_part2:
    add(_sp2_closed)

# ── Swept solid 3: 12-point profile extruded along translation line ───────────
_sp3_start = [
    (-534.72625732,  137.26408005,   65.06082058),
    (-552.4338913,    83.46186638, -147.52271652),
    (-553.41617584,   80.12220383, -173.34644318),
    (-552.33371735,   78.27410221, -363.33438873),
    (-551.05690002,   81.11089706, -389.20532227),
    (-531.88514709,  127.97060966, -609.05384064),
    (-546.91318512,   87.53202438, -583.21975708),
    (-564.30206299,   44.64496613, -398.94187927),
    (-566.0736084,    40.70904732, -363.04725647),
    (-567.15606689,   42.5571537,  -173.05931091),
    (-565.793190,     47.19076633, -137.23016739),
    (-548.08555603,  100.99297523,   75.35336971),
]

# Translation vector from line 1: (-552.434…,83.462…,−147.523…) → (−730.870…,148.714…,−149.174…)
_sp3_dx = -730.86982727 - (-552.4338913)   # -178.43593597
_sp3_dy =  148.71439934 -   83.46186638    #  +65.25253296
_sp3_dz = -149.17409897 - (-147.52271652)  #   -1.65138245

_sp3_face  = _make_planar_cap(_sp3_start)
_sp3_prism = BRepPrimAPI_MakePrism(_sp3_face.wrapped, gp_Vec(_sp3_dx, _sp3_dy, _sp3_dz))
_sp3_prism.Build()
_sp3_closed = Solid(_sp3_prism.Shape())
print(f"Swept3 prism complete: is_manifold={_sp3_closed.is_manifold}  volume={_sp3_closed.volume:.2f}")

with BuildPart() as swept_part3:
    add(_sp3_closed)

# ── Swept solid 4: 12-point profile extruded along translation line ───────────
_sp4_start = [
    (-725.88874817,  196.7896843,  -610.77384949),
    (-746.41586304,  154.32138443, -395.83099365),
    (-747.97439575,  151.15730286, -370.01350403),
    (-749.18060303,  149.78930473, -180.02225876),
    (-747.95036316,  152.58116722, -154.14418221),
    (-725.94497681,  200.93118668,   59.34596062),
    (-743.48274231,  166.15316391,   68.52691174),
    (-764.43519592,  117.43898392, -144.4862175 ),
    (-766.14219666,  113.56541634, -180.39077759),
    (-764.93591309,  114.93341446, -370.38200378),
    (-762.77351379,  119.32341576, -406.20250702),
    (-741.43325806,  163.72909546, -589.7038269 ),
]

# Translation from line 1: (-746.416…,154.321…,−395.831…) → (−918.484…,234.896…,−396.343…)
_sp4_dx = -918.48396301 - (-746.41586304)   # -172.06809997
_sp4_dy =  234.89614487 -  154.32138443     #  +80.57476044
_sp4_dz = -396.3433075  - (-395.83099365)   #   -0.51231385

_sp4_face  = _make_planar_cap(_sp4_start)
_sp4_prism = BRepPrimAPI_MakePrism(_sp4_face.wrapped, gp_Vec(_sp4_dx, _sp4_dy, _sp4_dz))
_sp4_prism.Build()
_sp4_closed = Solid(_sp4_prism.Shape())
print(f"Swept4 prism complete: is_manifold={_sp4_closed.is_manifold}  volume={_sp4_closed.volume:.2f}")

with BuildPart() as swept_part4:
    add(_sp4_closed)

# ── 12 quadrilateral cuts through swept prisms ───────────────────────────────
_quad_cuts = [
    # Cut 1
    [(-754.58969116,  199.54776764,    5.41946292),
     (-880.47111511,  258.49456787,    5.04466772),
     (-894.37446594,  227.94614792, -129.84228134),
     (-768.49304199,  168.99934769, -129.46748734)],
    # Cut 2
    [(-772.11212158,  160.78689575, -205.58984756),
     (-897.99354553,  219.73367691, -205.96464157),
     (-897.11105347,  220.73448181, -344.95822906),
     (-771.22962952,  161.78768158, -344.58343506)],
    # Cut 3
    [(-766.64512634,  171.09506607, -420.52757263),
     (-892.52655029,  230.0418663,  -420.90236664),
     (-876.91444397,  262.52811432, -555.1480484 ),
     (-751.03302002,  203.58133316, -554.77325439)],
    # Cut 4
    [(-563.1414032,   132.44882584,   11.21015906),
     (-693.68133545,  180.18619537,   10.00204802),
     (-704.86938477,  146.19298935, -124.31209564),
     (-574.32937622,   98.45561028, -123.10399055)],
    # Cut 5
    [(-577.21885681,   88.63175392, -199.06646729),
     (-707.75886536,  136.369133,   -200.27458191),
     (-706.9669342,   135.01708984, -339.26574707),
     (-576.42696381,   87.27971077, -338.05763245)],
    # Cut 6
    [(-572.67112732,   95.62437057, -414.15897369),
     (-703.21105957,  143.36174965, -415.36708832),
     (-690.48973083,  174.73701477, -550.18089294),
     (-559.94976044,  126.99964523, -548.97277832)],
    # Cut 7
    [(-380.04901886,   26.6913867,   -20.36664486),
     (-514.25281525,   62.88537979,  -20.94317198),
     (-522.43553162,   30.39582968, -155.84488869),
     (-388.23173523,   -5.79815984, -155.26836395)],
    # Cut 8
    [(-390.25402069,  -14.50921535, -231.39446259),
     (-524.45781708,   21.68477535, -231.97097778),
     (-523.55110168,   22.83282757, -370.96328735),
     (-389.3473053,   -13.36116314, -370.38677216)],
    # Cut 9
    [(-386.33300781,   -3.39407831, -446.32572174),
     (-520.5368042,    32.7999115,  -446.90223694),
     (-510.59841156,   67.51242638, -581.13067627),
     (-376.39461517,   31.31843567, -580.55416107)],
    # Cut 10
    [(-160.92475891,   31.68284416,   15.9566772 ),
     (-297.77238846,   56.0424757,    15.38015008),
     (-303.09232712,   22.96338797, -119.52156067),
     (-166.2446785,    -1.39624089, -118.94503593)],
    # Cut 11
    [(-167.50005722,  -10.2504015,  -195.0711441 ),
     (-304.34770584,   14.10922766, -195.6476593 ),
     (-303.54448318,   15.33193588, -334.63996887),
     (-166.69685364,   -9.02769327, -334.06345367)],
    # Cut 12
    [(-164.56272125,    1.16417706, -410.00240326),
     (-301.4103508,    25.52380562, -410.57891846),
     (-294.53517914,   60.97041607, -544.80735779),
     (-157.68754005,   36.6107893,  -544.23084259)],
]

with BuildPart() as swept_combined:
    add(swept_part.part)
    add(swept_part2.part)
    add(swept_part3.part)
    add(swept_part4.part)

    for _qi, _qpts in enumerate(_quad_cuts, 1):
        _qn  = len(_qpts)
        _qcx = sum(p[0] for p in _qpts) / _qn
        _qcy = sum(p[1] for p in _qpts) / _qn
        _qcz = sum(p[2] for p in _qpts) / _qn
        _qcc = Vector(_qcx, _qcy, _qcz)
        _qnx = _qny = _qnz = 0.0
        for _qi2 in range(_qn):
            _qp1, _qp2 = _qpts[_qi2], _qpts[(_qi2 + 1) % _qn]
            _qnx += (_qp1[1] - _qp2[1]) * (_qp1[2] + _qp2[2])
            _qny += (_qp1[2] - _qp2[2]) * (_qp1[0] + _qp2[0])
            _qnz += (_qp1[0] - _qp2[0]) * (_qp1[1] + _qp2[1])
        _qnorm  = Vector(_qnx, _qny, _qnz).normalized()
        _qref   = Vector(0, 0, 1) if abs(_qnorm.Z) < 0.9 else Vector(1, 0, 0)
        _qxloc  = _qnorm.cross(_qref).normalized()
        _qyloc  = _qnorm.cross(_qxloc).normalized()
        _qproj  = [((Vector(*p) - _qcc).dot(_qxloc),
                    (Vector(*p) - _qcc).dot(_qyloc)) for p in _qpts]
        _qplane = Plane(origin=_qcc, x_dir=_qxloc, z_dir=_qnorm)
        with BuildSketch(_qplane, mode=Mode.PRIVATE) as _qsk:
            with BuildLine():
                Polyline(*_qproj, close=True)
            make_face()
        extrude(_qsk.sketch, amount=250.0, mode=Mode.SUBTRACT, both=True)
        print(f"Quad cut {_qi} applied")

# ── Profile from Excel at Y=-10, extruded to Y=-37.5 ─────────────────────────
EXCEL_PATH2 = os.path.expanduser("~/Desktop/3x5+3_low_clipped(2).xlsx")
_wb2 = openpyxl.load_workbook(EXCEL_PATH2) if os.path.exists(EXCEL_PATH2) else None

def _read_periphery_xz2(sheet_name):
    pts = []
    if _wb2 is None:
        return pts
    for row in _wb2[sheet_name].iter_rows(values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        pts.append((float(row[1]), float(row[3])))
    return _dedupe(pts)

def _read_circles2():
    if _wb2 is None:
        return []
    ws = _wb2['Circles']
    result = []
    in_definitions = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Circle_ID' and row[1] == 'Center_X_mm':
            in_definitions = True
            continue
        if not in_definitions:
            continue
        if row[0] == 'Circle Vertex Data' or not isinstance(row[0], (int, float)):
            break
        result.append((float(row[1]), float(row[3]), float(row[4])))
    return result

outer_pts2 = _read_periphery_xz2('Outer Periphery')
inner_pts2 = _read_periphery_xz2('Inner Periphery')
circles2   = _read_circles2()
print(f"[Profile2] Outer: {len(outer_pts2)} pts  Inner: {len(inner_pts2)} pts  Circles: {len(circles2)}")

# Plane at Y=-10, normal in -Y → extrude(27.5) travels to Y=-37.5
_wp2 = Plane(origin=(0, -10.0, 0), x_dir=(1, 0, 0), z_dir=(0, -1, 0))

with BuildPart() as profile_part:
    with BuildSketch(_wp2):
        with BuildLine():
            Polyline(*outer_pts2, close=True)
        make_face()

        with BuildSketch(_wp2, mode=Mode.PRIVATE) as _inner_priv2:
            with BuildLine():
                Polyline(*inner_pts2, close=True)
            make_face()
        add(_inner_priv2.sketch, mode=Mode.SUBTRACT)

        _circle_locs2 = [(cx, cz) for cx, cz, _ in circles2]
        with Locations(*_circle_locs2):
            Circle(circles2[0][2], mode=Mode.SUBTRACT)

    extrude(amount=27.5)
    print("Profile extrusion complete: Y=-10 → Y=-37.5")

pp = profile_part.part
if pp:
    print(f"Profile Is Manifold : {pp.is_manifold}")
    print(f"Profile Volume      : {pp.volume:.2f} mm³")
else:
    print("Error: No profile geometry created.")

# ── Profile from 3skeletyl_shape.xlsx at Y=-10, extruded to Y=-40 ─────────────
EXCEL_PATH3 = os.path.expanduser("~/Desktop/3x5+3_low_clipped.xlsx")
_wb3 = openpyxl.load_workbook(EXCEL_PATH3) if os.path.exists(EXCEL_PATH3) else None

def _read_periphery_xz3(sheet_name):
    pts = []
    if _wb3 is None:
        return pts
    for row in _wb3[sheet_name].iter_rows(values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        pts.append((float(row[1]), float(row[3])))
    return _dedupe(pts)

outer_pts3 = _read_periphery_xz3('Outer Periphery')
inner_pts3 = _read_periphery_xz3('Inner Periphery')
print(f"[Profile3] Outer: {len(outer_pts3)} pts  Inner: {len(inner_pts3)} pts  (no circles sheet)")

# Plane at Y=-37.5, normal in -Y → extrude(2.5) travels to Y=-40
_wp3 = Plane(origin=(0, -37.5, 0), x_dir=(1, 0, 0), z_dir=(0, -1, 0))

with BuildPart() as profile_part2:
    with BuildSketch(_wp3):
        with BuildLine():
            Polyline(*outer_pts3, close=True)
        make_face()

        with BuildSketch(_wp3, mode=Mode.PRIVATE) as _inner_priv3:
            with BuildLine():
                Polyline(*inner_pts3, close=True)
            make_face()
        add(_inner_priv3.sketch, mode=Mode.SUBTRACT)

    extrude(amount=2.5)
    print("Profile2 extrusion complete: Y=-37.5 → Y=-40")

pp2 = profile_part2.part
if pp2:
    print(f"Profile2 Is Manifold : {pp2.is_manifold}")
    print(f"Profile2 Volume      : {pp2.volume:.2f} mm³")
else:
    print("Error: No profile2 geometry created.")

# ── Second lofted sweep: 13-point profile with fillets ───────────────────────
# Start profile points (in order)
_sw2_start = [
    (-1005.54504395,  47.50864983,  468.95496368),
    (-1117.46160719,  53.14289374,  319.34056804),  # fillet r=15.916661
    ( -997.0740509,  146.09633446,  232.78711319),
    ( -963.22525024, 170.67960739,  213.27558517),
    ( -943.34739685, 188.7881279,   186.5873909 ),
    ( -821.24771118, 254.27799225,  140.79333305),
    ( -754.96589661, 281.96849823,  114.73377228),  # anchor 1 (index 6)
    ( -588.84155954, 345.27420105,   86.53644157),  # fillet r=30.000011945
    ( -541.30931854, 298.78492355,  262.19957352),
    ( -707.06489563, 235.11837006,  291.7603302 ),
    ( -724.21043396, 229.46388245,  302.24931717),
    ( -866.18804932, 145.86549759,  374.73155975),  # anchor 2 (index 11)
    ( -885.30578613, 140.46954155,  382.20317841),
]

# Anchor-based end-profile interpolation using the two sweep lines:
#   Line 1 (anchor index 6): start=(-754.96589661,281.96849823,114.73377228)
#                              end=(-743.35983276,245.8565712,102.03627586)
#   Line 2 (anchor index 11): start=(-866.18804932,145.86549759,374.73155975)
#                               end=(-852.50518799,110.82853317,361.12316132)
_sw2_z_anc1 = _sw2_start[6][2]   # 114.73377228
_sw2_z_anc2 = _sw2_start[11][2]  # 374.73155975
_sw2_d_anc1 = (
    -743.35983276 - _sw2_start[6][0],
    245.8565712   - _sw2_start[6][1],
    102.03627586  - _sw2_start[6][2],
)
_sw2_d_anc2 = (
    -852.50518799 - _sw2_start[11][0],
    110.82853317  - _sw2_start[11][1],
    361.12316132  - _sw2_start[11][2],
)

def _end_pt2(x, y, z):
    t = (z - _sw2_z_anc1) / (_sw2_z_anc2 - _sw2_z_anc1)
    return (
        x + _sw2_d_anc1[0] + t * (_sw2_d_anc2[0] - _sw2_d_anc1[0]),
        y + _sw2_d_anc1[1] + t * (_sw2_d_anc2[1] - _sw2_d_anc1[1]),
        z + _sw2_d_anc1[2] + t * (_sw2_d_anc2[2] - _sw2_d_anc1[2]),
    )

_sw2_end = [_end_pt2(*p) for p in _sw2_start]

_wire2_start = Wire.make_polygon([Vector(*p) for p in _sw2_start], close=True)
_wire2_end   = Wire.make_polygon([Vector(*p) for p in _sw2_end],   close=True)
_cap2_start  = _make_planar_cap(_sw2_start)
_cap2_end    = _make_planar_cap(_sw2_end)

_lateral2 = Solid.make_loft([_wire2_start, _wire2_end])

_sew2 = BRepBuilderAPI_Sewing(1e-2)
_sew2.Add(_lateral2.wrapped)
_sew2.Add(_cap2_start.wrapped)
_sew2.Add(_cap2_end.wrapped)
_sew2.Perform()
_ocp_sol2 = OCP_Solid()
_bld_sol2  = BRep_Builder()
_bld_sol2.MakeSolid(_ocp_sol2)

from OCP.TopAbs import TopAbs_SHELL
from OCP.TopExp import TopExp_Explorer
_sewn2 = _sew2.SewedShape()
if _sewn2.ShapeType() == TopAbs_SHELL:
    _bld_sol2.Add(_ocp_sol2, _sewn2)
else:
    _exp2 = TopExp_Explorer(_sewn2, TopAbs_SHELL)
    _exp2.Init(_sewn2, TopAbs_SHELL)
    if _exp2.More():
        _bld_sol2.Add(_ocp_sol2, _exp2.Current())
    else:
        _bld_sol2.Add(_ocp_sol2, _sewn2)

_closed_solid2 = Solid(_ocp_sol2)

_loft2_cuts = [
    [(-1001.43005371,  61.72926903, 438.39286804),
     (-1084.65202332,  65.91894627, 327.13848114),
     ( -989.02992249, 139.7505188,  258.39029312),
     ( -905.80795288, 135.56083679, 369.6446991 )],
    [( -859.09797668, 161.03891373, 342.79415131),
     ( -930.08834839, 179.19240952, 224.67636108),
     ( -820.45013428, 243.74893188, 168.7039566 ),
     ( -749.45976257, 225.5954361,  286.82174683)],
    [( -689.80209351, 250.22073746, 264.58787918),
     ( -724.97840881, 284.62520599, 134.58804131),
     ( -596.6934967,  333.51127625, 112.8134346 ),
     ( -561.51714325, 299.10678864, 242.81328201)],
]

with BuildPart() as loft_part2:
    add(_closed_solid2)
    for _lci, _lcpts in enumerate(_loft2_cuts, 1):
        _ln = len(_lcpts)
        _lcx = sum(p[0] for p in _lcpts) / _ln
        _lcy = sum(p[1] for p in _lcpts) / _ln
        _lcz = sum(p[2] for p in _lcpts) / _ln
        _lcc = Vector(_lcx, _lcy, _lcz)
        _lcnx = _lcny = _lcnz = 0.0
        for _li in range(_ln):
            _lp1, _lp2 = _lcpts[_li], _lcpts[(_li + 1) % _ln]
            _lcnx += (_lp1[1] - _lp2[1]) * (_lp1[2] + _lp2[2])
            _lcny += (_lp1[2] - _lp2[2]) * (_lp1[0] + _lp2[0])
            _lcnz += (_lp1[0] - _lp2[0]) * (_lp1[1] + _lp2[1])
        _lcnorm = Vector(_lcnx, _lcny, _lcnz).normalized()
        _lcref  = Vector(0, 0, 1) if abs(_lcnorm.Z) < 0.9 else Vector(1, 0, 0)
        _lcxl   = _lcnorm.cross(_lcref).normalized()
        _lcyl   = _lcnorm.cross(_lcxl).normalized()
        _lcproj = [((Vector(*p) - _lcc).dot(_lcxl),
                    (Vector(*p) - _lcc).dot(_lcyl)) for p in _lcpts]
        _lcplane = Plane(origin=_lcc, x_dir=_lcxl, z_dir=_lcnorm)
        with BuildSketch(_lcplane, mode=Mode.PRIVATE) as _lcsk:
            with BuildLine():
                Polyline(*_lcproj, close=True)
            make_face()
        extrude(_lcsk.sketch, amount=55.0, mode=Mode.SUBTRACT, both=True)
        print(f"Loft2 cut {_lci} applied")

lp2 = loft_part2.part
if lp2:
    print(f"Loft2 Is Manifold : {lp2.is_manifold}")
    print(f"Loft2 Volume      : {lp2.volume:.2f} mm³")
else:
    print("Error: No loft2 geometry created.")

# ── Loft feature 1: 4-pt start profile → periphery-traced end profile at y=-10 ─
_lf1_start = [
    (-706.78936276, 234.36965464, 291.47849947),
    (-543.4452744,  304.71192416, 264.39917878),
    (-531.37051487, 271.18400637, 251.92634639),
    (-693.76927463, 198.98969867, 278.16081083),
]
_lf1_end_pts = [
    (-686.225, -10.0, 319.889),   # P1
    (-530.129, -10.0, 291.524),   # P2  (P1→P2 longer → periphery arc)
    (-539.207, -10.0, 253.947),   # P3  (P2→P3 shorter → straight)
    (-687.887, -10.0, 284.138),   # P4  (P3→P4 longer → periphery arc, P4→P1 shorter → straight)
]

def _lf_closest(pts_xz, tx, tz):
    return min(range(len(pts_xz)), key=lambda i: (pts_xz[i][0]-tx)**2 + (pts_xz[i][1]-tz)**2)

def _lf_arc_mid(pts_xz, i1, i2, y_val=-10.0):
    """Intermediate periphery points (exclusive) between i1 and i2, shorter arc."""
    n = len(pts_xz)
    fwd, bwd = [], []
    i = (i1 + 1) % n
    while i != i2:
        fwd.append(i); i = (i + 1) % n
    i = (i1 - 1) % n
    while i != i2:
        bwd.append(i); i = (i - 1) % n
    arc = fwd if len(fwd) <= len(bwd) else bwd
    return [(pts_xz[k][0], y_val, pts_xz[k][1]) for k in arc]

# Find which periphery (outer/inner) each arc endpoint belongs to
def _lf_best_perim(xz):
    oi = _lf_closest(outer_pts2, *xz)
    ii = _lf_closest(inner_pts2, *xz)
    od = math.hypot(outer_pts2[oi][0]-xz[0], outer_pts2[oi][1]-xz[1])
    id_ = math.hypot(inner_pts2[ii][0]-xz[0], inner_pts2[ii][1]-xz[1])
    return ('outer', oi, od) if od <= id_ else ('inner', ii, id_)

_p1_xz = (_lf1_end_pts[0][0], _lf1_end_pts[0][2])
_p2_xz = (_lf1_end_pts[1][0], _lf1_end_pts[1][2])
_p3_xz = (_lf1_end_pts[2][0], _lf1_end_pts[2][2])
_p4_xz = (_lf1_end_pts[3][0], _lf1_end_pts[3][2])

_lf1_p1_which, _lf1_p1_i, _lf1_p1_d = _lf_best_perim(_p1_xz)
_lf1_p3_which, _lf1_p3_i, _lf1_p3_d = _lf_best_perim(_p3_xz)
print(f"Loft1 P1 on {_lf1_p1_which} periphery (snap dist={_lf1_p1_d:.2f} mm)")
print(f"Loft1 P3 on {_lf1_p3_which} periphery (snap dist={_lf1_p3_d:.2f} mm)")

_perim12 = outer_pts2 if _lf1_p1_which == 'outer' else inner_pts2
_lf1_p2_i = _lf_closest(_perim12, *_p2_xz)
_lf1_arc12 = _lf_arc_mid(_perim12, _lf1_p1_i, _lf1_p2_i)

_perim34 = outer_pts2 if _lf1_p3_which == 'outer' else inner_pts2
_lf1_p4_i = _lf_closest(_perim34, *_p4_xz)
_lf1_arc34 = _lf_arc_mid(_perim34, _lf1_p3_i, _lf1_p4_i)

print(f"Loft1 arc P1→P2: {len(_lf1_arc12)} intermediate pts")
print(f"Loft1 arc P3→P4: {len(_lf1_arc34)} intermediate pts")

# Assemble end profile: P1 → arc → P2 → straight → P3 → arc → P4 → (closes to P1)
_lf1_end_3d = (
    [_lf1_end_pts[0]] + _lf1_arc12 +
    [_lf1_end_pts[1], _lf1_end_pts[2]] + _lf1_arc34 +
    [_lf1_end_pts[3]]
)

_lf1_wire_start = Wire.make_polygon([Vector(*p) for p in _lf1_start],  close=True)
_lf1_wire_end   = Wire.make_polygon([Vector(*p) for p in _lf1_end_3d], close=True)
_lf1_cap_start  = _make_planar_cap(_lf1_start)
_lf1_cap_end    = _make_planar_cap(_lf1_end_3d)

_lf1_lateral = Solid.make_loft([_lf1_wire_start, _lf1_wire_end])

_lf1_sew = BRepBuilderAPI_Sewing(1e-2)
_lf1_sew.Add(_lf1_lateral.wrapped)
_lf1_sew.Add(_lf1_cap_start.wrapped)
_lf1_sew.Add(_lf1_cap_end.wrapped)
_lf1_sew.Perform()
_lf1_ocp = OCP_Solid()
_lf1_bld = BRep_Builder()
_lf1_bld.MakeSolid(_lf1_ocp)
_lf1_sewn = _lf1_sew.SewedShape()
if _lf1_sewn.ShapeType() == TopAbs_SHELL:
    _lf1_bld.Add(_lf1_ocp, _lf1_sewn)
else:
    _lf1_exp = TopExp_Explorer(_lf1_sewn, TopAbs_SHELL)
    _lf1_exp.Init(_lf1_sewn, TopAbs_SHELL)
    _lf1_bld.Add(_lf1_ocp, _lf1_exp.Current() if _lf1_exp.More() else _lf1_sewn)
_lf1_solid = Solid(_lf1_ocp)

# Edge P1→P2 midpoint on start profile
_lf1_p1p2_mid = (
    (_lf1_start[0][0] + _lf1_start[1][0]) / 2,
    (_lf1_start[0][1] + _lf1_start[1][1]) / 2,
    (_lf1_start[0][2] + _lf1_start[1][2]) / 2,
)

def _find_edge_near(solid, target_xyz, tol=50.0):
    tx, ty, tz = target_xyz
    best_edge, best_dist = None, float("inf")
    for edge in solid.edges():
        verts = edge.vertices()
        if len(verts) < 2:
            continue
        mx = sum(v.X for v in verts) / len(verts)
        my = sum(v.Y for v in verts) / len(verts)
        mz = sum(v.Z for v in verts) / len(verts)
        d = math.sqrt((mx-tx)**2 + (my-ty)**2 + (mz-tz)**2)
        if d < best_dist:
            best_dist, best_edge = d, edge
    if best_dist > tol:
        print(f"WARNING: nearest edge is {best_dist:.2f} mm away from target")
    else:
        print(f"Fillet edge found: dist={best_dist:.2f} mm from P1-P2 midpoint")
    return best_edge

_lf1_fillet_edge = _find_edge_near(_lf1_solid, _lf1_p1p2_mid)

with BuildPart() as loft_feature1:
    add(_lf1_solid)
    if _lf1_fillet_edge is not None:
        try:
            fillet(_lf1_fillet_edge, 68.0428283 / 2)
            print("Loft Feature1 fillet (r=34.021mm) applied on P1-P2 edge")
        except Exception as _lfe:
            print(f"WARNING: fillet skipped — {_lfe}")

lf1 = loft_feature1.part
if lf1:
    print(f"Loft Feature1 Is Manifold : {lf1.is_manifold}")
    print(f"Loft Feature1 Volume      : {lf1.volume:.2f} mm³")
else:
    print("Error: No loft feature1 geometry created.")

# ── Four extruded profiles (BRepPrimAPI_MakePrism) ───────────────────────────
_ep_defs = [
    # Profile 1 — quad
    dict(pts=[
            (-990.09796143,  -17.43142128, 389.92115021),
            (-1036.2436676,  -21.33286953, 324.6880722 ),
            (-1086.70791626, -35.89508533, 361.25728607),
            (-1040.56221008, -31.99363708, 426.49032593),
         ],
         line=((-1036.2436676,  -21.33286953, 324.6880722 ),
               (-1044.77973938,  17.56905198, 328.39981079))),
    # Profile 2 — L-shape
    dict(pts=[
            (-693.96408081,  94.91221428,  25.94517708),
            (-687.7583313,  113.04777145,  20.23631573),
            (-684.37638408, 124.05895688,  58.89227295),
            (-677.27403558, 144.81488715,  52.35855388),
            (-675.47134246, 150.68413678,  72.96321861),
            (-688.77939486, 111.79263385,  85.20579944),
         ],
         line=((-688.77939486, 111.79263385,  85.20579944),
               (-613.05244446,  85.99729538,  85.57812691))),
    # Profile 3 — L-shape
    dict(pts=[
            (-890.57975769, 150.13086319,  26.42907619),
            (-875.39482341, 185.27305621,  14.83410311),
            (-870.69432887, 195.87305385,  53.11654332),
            (-859.98766708, 220.65126945,  44.94110752),
            (-857.40239502, 226.4812683,   65.99645017),
            (-883.29399109, 166.56085968,  85.76685905),
         ],
         line=((-890.57975769, 150.13086319,  26.42907619),
               (-817.16796875, 118.33985329,  26.21783733))),
    # Profile 4 — L-shape
    dict(pts=[
            ( -64.95278358,  -17.80839443, 152.52279282),
            ( -61.46655564,   20.68161295, 142.20942668),
            ( -61.46655564,   31.03437358, 180.84645632),
            ( -59.41302836,   53.70652278, 174.7714724 ),
            ( -59.41302836,   61.60592985, 204.25246168),
            ( -64.95278358,    0.44377327, 220.64081173),
         ],
         line=(( -64.95278358,  -17.80839443, 152.52279282),
               (  14.74279046,  -24.54327345, 154.32739258))),
    # Profile 5 — L-shape
    dict(pts=[
            ( -62.24393368,  -17.24526405, -522.23152161),
            ( -55.39964962,   59.58188946, -522.82632176),
            ( -55.39964962,   59.89156245, -482.82751944),
            ( -58.6946249,    22.90543359, -482.54117717),
            ( -58.6946249,    23.21510652, -442.54238128),
            ( -62.24393368,  -16.62591815, -442.23392487),
         ],
         line=(( -62.24393368,  -16.62591815, -442.23392487),
               (  17.44050503,  -23.7243247,  -442.17895508))),
    # Profile 6 — quad
    dict(pts=[
            (-878.86604309, 196.05964661, -562.13302612),
            (-901.95327802, 143.5626969,  -579.76958738),
            (-897.68409518, 158.07872596, -628.56656442),
            (-874.59686025, 210.57567566, -610.93000316),
         ],
         line=((-878.86604309, 196.05964661, -562.13302612),
               (-805.3288269,  164.69261169, -565.03036499))),
    # Profile 7 — quad
    dict(pts=[
            (-679.42756647, 108.17852993, -561.87336038),
            (-691.77368164,  71.71259403, -572.72727966),
            (-687.81324352,  83.84540852, -617.99484132),
            (-675.46712835, 120.31134441, -607.14092204),
         ],
         line=((-679.42756647, 108.17852993, -561.87336038),
               (-603.63235474,  82.58477211, -562.10189819))),
]

_ep_solids = []
for _ei, _ep in enumerate(_ep_defs, 1):
    _ep_dx = _ep['line'][1][0] - _ep['line'][0][0]
    _ep_dy = _ep['line'][1][1] - _ep['line'][0][1]
    _ep_dz = _ep['line'][1][2] - _ep['line'][0][2]
    _ep_face  = _make_planar_cap(_ep['pts'])
    _ep_prism = BRepPrimAPI_MakePrism(_ep_face.wrapped, gp_Vec(_ep_dx, _ep_dy, _ep_dz))
    _ep_prism.Build()
    _ep_sol = Solid(_ep_prism.Shape())
    _ep_solids.append(_ep_sol)
    print(f"Extruded profile {_ei}: is_manifold={_ep_sol.is_manifold}  volume={_ep_sol.volume:.2f} mm³")

# ── 3-profile cross-section loft (x=-109.4192 → x=40.0 → x=73.8818) ────────
# Profile A: x = 73.8818 (69 pts)
_loft3_A = [
    (73.8818, -10.0, 239.6033), (73.8818, -9.7, 239.6033),
    (73.8818, 0.0, 239.6033), (73.8818, 0.0782, 239.5989),
    (73.8818, 0.3215, 239.5844), (73.8818, 7.4295, 239.1247),
    (73.8818, 7.7542, 239.0667), (73.8818, 14.6729, 237.6901),
    (73.8818, 15.0161, 237.5864), (73.8818, 21.7548, 235.3025),
    (73.8818, 22.1314, 235.1414), (73.8818, 28.7007, 231.9672),
    (73.8818, 29.2056, 231.6807), (73.8818, 35.6273, 227.634),
    (73.8818, 36.2088, 227.2154), (73.8818, 42.4982, 222.3221),
    (73.8818, 43.1762, 221.7309), (73.8818, 49.3476, 216.023),
    (73.8818, 49.8665, 215.5087), (73.8818, 56.1784, 208.7629),
    (73.8818, 56.5511, 208.3228), (73.8818, 56.6695, 208.1742),
    (73.8818, 56.9111, 207.8727), (73.8818, 57.6658, 206.9352),
    (73.8818, 60.711, 203.0092), (73.8818, 62.1809, 200.8719),
    (73.8818, 64.4797, 197.0217), (73.8818, 65.8728, 194.5934),
    (73.8818, 67.1761, 191.8134), (73.8818, 68.781, 188.1878),
    (73.8818, 69.7426, 185.6454), (73.8818, 70.7847, 182.193),
    (73.8818, 71.593, 179.5005), (73.8818, 72.4148, 175.6351),
    (73.8818, 72.9515, 172.4938), (73.8818, 73.437, 168.422),
    (73.8818, 73.5741, 165.493), (73.8818, 73.696, 161.2796),
    (73.8818, 73.4246, 156.851), (73.8818, 73.2273, 154.2377),
    (73.8818, 72.6833, 150.5146), (73.8818, 72.177, 148.0064),
    (73.8818, 71.3524, 144.3596), (73.8818, 70.7712, 142.1903),
    (73.8818, 67.459, 129.8291), (73.8818, 28.5297, 139.6816),
    (73.8818, 40.2697, 183.4957), (73.8818, 41.1517, 186.7875),
    (73.8818, 39.0797, 189.4739), (73.8818, 38.7904, 189.7988),
    (73.8818, 36.3163, 192.2863), (73.8818, 36.1606, 192.4078),
    (73.8818, 33.1669, 194.5164), (73.8818, 32.9954, 194.6134),
    (73.8818, 29.2268, 196.3412), (73.8818, 29.0271, 196.4161),
    (73.8818, 24.2277, 197.7615), (73.8818, 23.988, 197.8161),
    (73.8818, 17.9022, 198.778), (73.8818, 17.7655, 198.7978),
    (73.8818, 17.6122, 198.8135), (73.8818, 14.0075, 199.1502),
    (73.8818, 13.8381, 199.1613), (73.8818, 9.8162, 199.4019),
    (73.8818, 9.6291, 199.4086), (73.8818, 5.1581, 199.553),
    (73.8818, 5.0345, 199.5544), (73.8818, 0.0, 199.6033),
    (73.8818, -10.0, 199.6033),
]

# Profile B: x = 40.0 (78 pts)
_loft3_B = [
    (40.0, -10.0, 239.603), (40.0, -2.0084, 239.603),
    (40.0, 0.0, 239.603), (40.0, 2.0837, 239.4865),
    (40.0, 6.8632, 239.202), (40.0, 7.4158, 239.1098),
    (40.0, 13.6383, 237.9978), (40.0, 14.2029, 237.8412),
    (40.0, 20.3383, 235.9887), (40.0, 20.9189, 235.7637),
    (40.0, 26.9766, 233.1733), (40.0, 27.6913, 232.8098),
    (40.0, 34.8019, 228.7782), (40.0, 35.5521, 228.2974),
    (40.0, 42.5833, 223.2397), (40.0, 43.3769, 222.6249),
    (40.0, 50.3445, 216.5545), (40.0, 51.0145, 215.9518),
    (40.0, 51.4852, 215.4902), (40.0, 54.7402, 212.1733),
    (40.0, 56.8179, 210.1137), (40.0, 57.7604, 209.1064),
    (40.0, 60.25, 206.166), (40.0, 61.5886, 204.6112),
    (40.0, 63.6745, 201.6911), (40.0, 65.0405, 199.8017),
    (40.0, 66.7412, 196.9418), (40.0, 68.0792, 194.718),
    (40.0, 69.4268, 191.9456), (40.0, 70.6807, 189.3962),
    (40.0, 71.5649, 187.146), (40.0, 72.5139, 184.7596),
    (40.0, 73.1892, 182.593), (40.0, 74.0068, 180.0023),
    (40.0, 74.6092, 177.2985), (40.0, 75.4353, 173.6432),
    (40.0, 75.7603, 171.0899), (40.0, 76.2683, 167.1757),
    (40.0, 76.352, 164.7672), (40.0, 76.5026, 160.6566),
    (40.0, 76.3723, 158.3805), (40.0, 76.1353, 154.1431),
    (40.0, 75.8735, 152.3196), (40.0, 75.3578, 148.7072),
    (40.0, 75.2234, 148.1056), (40.0, 75.0712, 147.4232),
    (40.0, 74.9581, 146.915), (40.0, 74.8715, 146.5246),
    (40.0, 74.1696, 143.4204), (40.0, 71.7801, 134.5028),
    (40.0, 71.191, 132.3041), (40.0, 67.5602, 133.2711),
    (40.0, 38.3889, 141.04), (40.0, 32.3906, 142.6374),
    (40.0, 32.6719, 143.6873), (40.0, 36.5447, 158.1409),
    (40.0, 44.1687, 186.5938), (40.0, 41.5861, 189.0873),
    (40.0, 40.9692, 189.7802), (40.0, 40.6404, 190.1108),
    (40.0, 37.5235, 192.5411), (40.0, 37.1124, 192.8307),
    (40.0, 33.5457, 194.8481), (40.0, 33.0144, 195.0917),
    (40.0, 28.7232, 196.7007), (40.0, 28.0327, 196.8943),
    (40.0, 22.744, 198.0985), (40.0, 21.8553, 198.239),
    (40.0, 18.781, 198.6864), (40.0, 15.3302, 199.0377),
    (40.0, 14.7986, 199.0873), (40.0, 10.9618, 199.3383),
    (40.0, 10.3669, 199.3739), (40.0, 6.1095, 199.5245),
    (40.0, 5.447, 199.5459), (40.0, 2.6246, 199.5778),
    (40.0, 0.0, 199.6033), (40.0, -10.0, 199.6033),
]

# Profile C: x = -109.4192 (74 pts, stray outlier removed)
_loft3_C = [
    (-109.4192, -10.0, 239.1376), (-109.4192, 0.0, 239.1328),
    (-109.4192, 6.3587, 238.7545), (-109.4192, 7.0577, 238.7121),
    (-109.4192, 13.4784, 237.5971), (-109.4192, 14.3387, 237.4456),
    (-109.4192, 20.7114, 235.6498), (-109.4192, 21.8312, 235.3302),
    (-109.4192, 28.0471, 232.9339), (-109.4192, 29.5242, 232.357),
    (-109.4192, 36.5264, 228.8875), (-109.4192, 38.5085, 227.9116),
    (-109.4192, 40.765, 226.5717), (-109.4192, 47.4783, 222.465),
    (-109.4192, 50.4489, 220.4012), (-109.4192, 56.3518, 216.1079),
    (-109.4192, 60.1195, 213.1067), (-109.4192, 60.1322, 213.0963),
    (-109.4192, 64.1947, 209.7099), (-109.4192, 66.2214, 207.7374),
    (-109.4192, 68.9122, 205.2376), (-109.4192, 70.1499, 203.911),
    (-109.4192, 71.7614, 202.0289), (-109.4192, 73.213, 200.3357),
    (-109.4192, 75.6362, 197.0554), (-109.4192, 76.8438, 195.3501),
    (-109.4192, 77.0592, 195.0457), (-109.4192, 77.6798, 193.9685),
    (-109.4192, 80.4117, 189.4093), (-109.4192, 81.2671, 187.7232),
    (-109.4192, 81.796, 186.5932), (-109.4192, 82.8208, 184.4282),
    (-109.4192, 82.8796, 184.2707), (-109.4192, 83.6319, 182.3346),
    (-109.4192, 84.882, 179.084), (-109.4192, 86.0857, 174.8625),
    (-109.4192, 87.1819, 170.441), (-109.4192, 87.4795, 168.9296),
    (-109.4192, 87.6292, 168.0327), (-109.4192, 87.7761, 167.0983),
    (-109.4192, 88.4135, 160.7733), (-109.4192, 88.6027, 158.1178),
    (-109.4192, 88.5417, 156.2843), (-109.4192, 88.2227, 148.6405),
    (-109.4192, 88.0408, 147.2079), (-109.4192, 86.7202, 140.2228),
    (-109.4192, 47.9092, 150.0454), (-109.4192, 55.5656, 178.6196),
    (-109.4192, 57.4733, 185.7392), (-109.4192, 54.5396, 187.2885),
    (-109.4192, 51.9611, 188.4259), (-109.4192, 47.8215, 190.2437),
    (-109.4192, 46.5641, 190.8573), (-109.4192, 43.1594, 192.157),
    (-109.4192, 41.4343, 192.8868), (-109.4192, 38.3817, 193.85),
    (-109.4192, 36.0861, 194.6509), (-109.4192, 33.2953, 195.3368),
    (-109.4192, 30.3637, 196.1313), (-109.4192, 27.7619, 196.5954),
    (-109.4192, 24.1107, 197.3112), (-109.4192, 22.8374, 197.4788),
    (-109.4192, 20.622, 197.7978), (-109.4192, 19.3707, 197.9256),
    (-109.4192, 16.9286, 198.1981), (-109.4192, 15.7064, 198.2901),
    (-109.4192, 13.0755, 198.5061), (-109.4192, 11.8457, 198.5687),
    (-109.4192, 8.9833, 198.7276), (-109.4192, 7.7269, 198.7639),
    (-109.4192, 4.6315, 198.8612), (-109.4192, 0.9421, 198.9154),
    (-109.4192, 0.0, 198.9217), (-109.4192, -10.0, 198.9275),
]

def _resample_profile(pts, target_n):
    """Resample a 3D polyline to exactly target_n points via linear interpolation."""
    dists = [0.0]
    for i in range(1, len(pts)):
        d = math.sqrt(sum((pts[i][k] - pts[i-1][k])**2 for k in range(3)))
        dists.append(dists[-1] + d)
    total = dists[-1]
    if total < 1e-9:
        return pts[:target_n]
    result = []
    for j in range(target_n):
        s = j / (target_n - 1) * total
        for i in range(1, len(dists)):
            if dists[i] >= s - 1e-12:
                t = (s - dists[i-1]) / max(dists[i] - dists[i-1], 1e-15)
                t = max(0.0, min(1.0, t))
                p = tuple(pts[i-1][k] + t * (pts[i][k] - pts[i-1][k]) for k in range(3))
                result.append(p)
                break
    return result

_LOFT3_N = 60
_l3a = _resample_profile(_loft3_A, _LOFT3_N)
_l3b = _resample_profile(_loft3_B, _LOFT3_N)
_l3c = _resample_profile(_loft3_C, _LOFT3_N)
print(f"3-profile loft: resampled to {_LOFT3_N} pts each")

_l3_wire_A = Wire.make_polygon([Vector(*p) for p in _l3a], close=True)
_l3_wire_B = Wire.make_polygon([Vector(*p) for p in _l3b], close=True)
_l3_wire_C = Wire.make_polygon([Vector(*p) for p in _l3c], close=True)

_l3_lateral = Solid.make_loft([_l3_wire_C, _l3_wire_B, _l3_wire_A])
_l3_cap_C = _make_planar_cap(_l3c)
_l3_cap_A = _make_planar_cap(_l3a)

_l3_sew = BRepBuilderAPI_Sewing(1e-2)
_l3_sew.Add(_l3_lateral.wrapped)
_l3_sew.Add(_l3_cap_C.wrapped)
_l3_sew.Add(_l3_cap_A.wrapped)
_l3_sew.Perform()
_l3_ocp = OCP_Solid()
_l3_bld = BRep_Builder()
_l3_bld.MakeSolid(_l3_ocp)
_l3_sewn = _l3_sew.SewedShape()
if _l3_sewn.ShapeType() == TopAbs_SHELL:
    _l3_bld.Add(_l3_ocp, _l3_sewn)
else:
    _l3_exp = TopExp_Explorer(_l3_sewn, TopAbs_SHELL)
    _l3_exp.Init(_l3_sewn, TopAbs_SHELL)
    _l3_bld.Add(_l3_ocp, _l3_exp.Current() if _l3_exp.More() else _l3_sewn)
_l3_solid = Solid(_l3_ocp)

with BuildPart() as loft_feature_3prof:
    add(_l3_solid)

_l3p = loft_feature_3prof.part
if _l3p:
    print(f"3-profile loft Is Manifold : {_l3p.is_manifold}")
    print(f"3-profile loft Volume      : {_l3p.volume:.2f} mm³")
else:
    print("Error: No 3-profile loft geometry created.")

# ── Second loft: 4 strip profiles (x=-116.11 → -118.11 → -127.91 → -132.97) ─
# Connects to first loft at Profile C (x=-109.4192)
_loft4_a = [
    (-116.1106, -10.0, 238.0929),
    (-116.1106, -8.9154, 238.0939),
    (-116.1106, 0.0, 238.089),
    (-116.1106, 0.0, 196.3277),
    (-116.1106, -10.0, 196.3347),
]
_loft4_b = [
    (-118.1106, -10.0, 237.5802),
    (-118.1106, -10.0, 237.5802),
    (-118.1106, 0.0, 237.5748),
    (-118.1106, 0.0, 195.0438),
    (-118.1106, -10.0, 195.0508),
]
_loft4_c = [
    (-127.9106, -10.0, 234.073),
    (-127.9106, 0.0, 234.0783),
    (-127.9106, 0.0, 185.9824),
    (-127.9106, -10.0, 186.0003),
]
_loft4_d = [
    (-132.9684, -10.0, 231.3484),
    (-132.9684, -4.4688, 231.3575),
    (-132.9684, 0.0, 231.3545),
    (-132.9684, -0.0, 178.7483),
    (-132.9684, -5.38, 178.7573),
    (-132.9684, -10.0, 178.7526),
]

# Deduplicate and resample all to same point count and loft
def _dedupe_3d(pts):
    if not pts: return []
    out = [pts[0]]
    for p in pts[1:]:
        if math.hypot(p[0]-out[-1][0], p[1]-out[-1][1], p[2]-out[-1][2]) > 1e-6:
            out.append(p)
    return out

# Extract the "box" section of the main loft's end profile (x=-109.4) to bridge the gap
_loft4_conn = [
    _loft3_C[0],   # Top-left corner (y=-10, z~239)
    _loft3_C[1],   # Top-right corner (y=0, z~239)
    _loft3_C[72],  # Bottom-right corner (y=0, z~199)
    _loft3_C[73],  # Bottom-left corner (y=-10, z~199)
]

_l4a_clean = _dedupe_3d(_loft4_a)
_l4b_clean = _dedupe_3d(_loft4_b)
_l4c_clean = _dedupe_3d(_loft4_c)
_l4d_clean = _dedupe_3d(_loft4_d)

_LOFT4_N = 40
_l4conn = _resample_profile(_loft4_conn, _LOFT4_N)
_l4a = _resample_profile(_l4a_clean, _LOFT4_N)
_l4b = _resample_profile(_l4b_clean, _LOFT4_N)
_l4c = _resample_profile(_l4c_clean, _LOFT4_N)
_l4d = _resample_profile(_l4d_clean, _LOFT4_N)
print(f"Second loft: 5 profiles (bridging from x=-109.4) resampled to {_LOFT4_N} pts each")

_l4_wire_conn = Wire.make_polygon([Vector(*p) for p in _l4conn], close=True)
_l4_wire_a = Wire.make_polygon([Vector(*p) for p in _l4a], close=True)
_l4_wire_b = Wire.make_polygon([Vector(*p) for p in _l4b], close=True)
_l4_wire_c = Wire.make_polygon([Vector(*p) for p in _l4c], close=True)
_l4_wire_d = Wire.make_polygon([Vector(*p) for p in _l4d], close=True)

# Loft ordered from a (x=-116.11) → b (x=-118.11) → c (x=-127.91) → d (x=-132.97)
_l4_lateral = Solid.make_loft([_l4_wire_conn, _l4_wire_a, _l4_wire_b, _l4_wire_c, _l4_wire_d])
_l4_cap_conn = _make_planar_cap(_l4conn)
_l4_cap_d = _make_planar_cap(_l4d)

_l4_sew = BRepBuilderAPI_Sewing(1e-2)
_l4_sew.Add(_l4_lateral.wrapped)
_l4_sew.Add(_l4_cap_conn.wrapped)
_l4_sew.Add(_l4_cap_d.wrapped)
_l4_sew.Perform()
_l4_ocp = OCP_Solid()
_l4_bld = BRep_Builder()
_l4_bld.MakeSolid(_l4_ocp)
_l4_sewn = _l4_sew.SewedShape()
if _l4_sewn.ShapeType() == TopAbs_SHELL:
    _l4_bld.Add(_l4_ocp, _l4_sewn)
else:
    _l4_exp = TopExp_Explorer(_l4_sewn, TopAbs_SHELL)
    _l4_exp.Init(_l4_sewn, TopAbs_SHELL)
    _l4_bld.Add(_l4_ocp, _l4_exp.Current() if _l4_exp.More() else _l4_sewn)
_l4_solid = Solid(_l4_ocp)

with BuildPart() as loft_feature_strip:
    add(_l4_solid)
    
_l4p = loft_feature_strip.part
if _l4p:
    print(f"Second loft (strip) Is Manifold : {_l4p.is_manifold}")
    print(f"Second loft (strip) Volume      : {_l4p.volume:.2f} mm³")
else:
    print("Error: No second loft geometry created.")



# ── Third loft (User provided 3 profiles) ────────────────────────────
_loft5_a = [
    (-137.2145, 29.6313, 24.6544),
    (-137.2145, 34.4144, 43.6148),
    (-137.2145, 34.7287, 44.8608),
    (-137.2145, 35.0694, 46.2111),
    (-137.2145, 35.2448, 46.9064),
    (-137.2145, 35.3672, 47.5059),
    (-137.2145, 35.5216, 48.2707),
    (-137.2145, 36.1173, 51.2225),
    (-137.2145, 36.5633, 53.432),
    (-137.2145, 36.799, 55.6842),
    (-137.2145, 37.2506, 59.9616),
    (-137.2145, 32.4314, 110.3552),
    (-137.2145, 32.628, 108.752),
    (-137.2145, 33.3361, 101.8257),
    (-137.2145, 33.941, 96.0132),
    (-137.2145, 34.4311, 92.2031),
    (-137.2145, 35.0569, 86.5467),
    (-137.2145, 35.6116, 81.7424),
    (-137.2145, 36.4062, 76.0127),
    (-137.2145, 36.755, 73.183),
    (-137.2145, 37.237, 67.4878),
    (-137.2145, 37.3189, 66.6196),
    (-137.2145, 37.2881, 63.3222),
    (-137.2145, 20.5276, 64.6393),
    (-137.2145, 21.6461, 69.6494),
    (-137.2145, 22.7405, 71.7138),
    (-137.2145, 23.4871, 73.6638),
    (-137.2145, 24.245, 78.8207),
    (-137.2145, 24.5232, 79.6972),
    (-137.2145, 25.2696, 85.8297),
    (-137.2145, 25.2719, 85.8459),
    (-137.2145, 25.2718, 85.8485),
    (-137.2145, 25.9848, 91.8641),
    (-137.2145, 25.9569, 92.723),
    (-137.2145, 26.4484, 97.8913),
    (-137.2145, 26.3928, 99.3985),
    (-137.2145, 26.7204, 103.8496),
    (-137.2145, 26.6359, 105.8427),
    (-137.2145, 26.838, 109.6828),
    (-137.2145, 26.7214, 112.0326),
    (-137.2145, 26.8245, 115.3492),
    (-137.2145, 26.6707, 117.9512),
    (-137.2145, 31.7263, 117.5126),
    (-137.2145, 31.686, 117.8429),
    (-137.2145, 31.5114, 119.6596),
    (-137.2145, 30.9774, 125.3392),
    (-137.2145, 31.0748, 124.161),
    (-137.2145, 26.6612, 123.2168),
    (-137.2145, -10.0, 182.04458419766198),
    (-137.2145, -10.0, 228.35546514235745),
    (-137.2145, 0.0, 228.3442),
    (-137.2145, 0.358, 228.3258),
    (-137.2145, 3.4343, 228.0889),
    (-137.2145, 3.9206, 228.0088),
    (-137.2145, 6.543, 227.4052),
    (-137.2145, 7.0575, 227.2547),
    (-137.2145, 9.3104, 226.3055),
    (-137.2145, 9.877, 226.0584),
    (-137.2145, 11.7369, 224.8482),
    (-137.2145, 13.0297, 223.9515),
    (-137.2145, 14.5366, 222.5683),
    (-137.2145, 16.1063, 220.8135),
    (-137.2145, 17.211, 219.4443),
    (-137.2145, 19.2575, 216.0735),
    (-137.2145, 19.8455, 215.1142),
    (-137.2145, 20.2352, 214.3115),
    (-137.2145, 21.9168, 210.3788),
    (-137.2145, 22.2145, 208.4003),
    (-137.2145, 22.7647, 205.193),
    (-137.2145, 24.1216, 195.7765),
    (-137.2145, 24.3249, 193.947),
    (-137.2145, 24.4809, 192.8926),
    (-137.2145, 24.9735, 188.1012),
    (-137.2145, 25.5365, 183.8716),
    (-137.2145, 25.6651, 182.5152),
    (-137.2145, 25.9585, 180.0656),
    (-137.2145, 26.3214, 175.7521),
    (-137.2145, 26.4703, 174.6261),
    (-137.2145, 26.6692, 172.8464),
    (-137.2145, 27.3043, 165.507),
    (-137.2145, 27.7686, 160.9939),
    (-137.2145, 28.0926, 156.9505),
    (-137.2145, 28.2514, 154.6397),
    (-137.2145, 24.8114, 141.8247),
    (-137.2145, 29.9755, 135.9692),
    (-137.2145, 30.2643, 133.0939),
    (-137.2145, 30.5079, 130.1421),
    (-137.2145, 26.1961, 129.0329),
    (-137.2145, 26.0938, 131.1852),
    (-137.2145, 25.7785, 134.1166),
    (-137.2145, 25.6266, 135.9706),
    (-137.2145, 25.2382, 138.8869),
    (-137.2145, 29.5584, 140.9345),
    (-137.2145, 29.4796, 141.6832),
    (-137.2145, 29.3169, 143.5506),
    (-137.2145, 28.8388, 148.8426),
    (-137.2145, 28.8945, 147.9916),
    (-137.2145, 23.7653, 147.4685),
    (-137.2145, 23.4947, 148.638),
    (-137.2145, 22.8162, 151.2725),
    (-137.2145, 22.5124, 152.2684),
    (-137.2145, 21.712, 154.7482),
    (-137.2145, 21.2893, 155.7779),
    (-137.2145, 20.1021, 158.6162),
    (-137.2145, 19.7427, 159.2728),
    (-137.2145, 18.6637, 161.2559),
    (-137.2145, 18.2763, 161.8049),
    (-137.2145, 17.0452, 163.5844),
    (-137.2145, 16.1787, 164.4059),
    (-137.2145, 13.2207, 167.3033),
    (-137.2145, 12.7439, 167.5912),
    (-137.2145, 10.9903, 168.6918),
    (-137.2145, 10.4796, 168.9078),
    (-137.2145, 8.5314, 169.7655),
    (-137.2145, 7.9833, 169.9145),
    (-137.2145, 5.8313, 170.5236),
    (-137.2145, 5.2414, 170.6091),
    (-137.2145, 2.8774, 170.9657),
    (-137.2145, 1.499, 171.0444),
    (-137.2145, -0.0, 171.0248),
    (-137.2145, -10.0, 171.04241824935082),
    (-137.2145, -10.0, 85.15979049619807),
    (-137.2145, 19.5812, 24.6544),
]
_loft5_b = [
    (-220.9569, 26.8771, 24.6544),
    (-220.9569, 44.6269, 24.6544),
    (-220.9569, 47.55, 36.2417),
    (-220.9569, 49.5518, 44.1767),
    (-220.9569, 50.1728, 46.6384),
    (-220.9569, 50.6887, 48.9189),
    (-220.9569, 50.92, 49.95),
    (-220.9569, 51.0103, 50.483),
    (-220.9569, 51.4996, 53.2632),
    (-220.9569, 51.5589, 53.771),
    (-220.9569, 51.9187, 56.5964),
    (-220.9569, 51.9495, 57.0709),
    (-220.9569, 52.1777, 59.9406),
    (-220.9569, 52.1833, 60.3969),
    (-220.9569, 52.2781, 63.4672),
    (-220.9569, 52.2613, 63.8697),
    (-220.9569, 52.2025, 66.9859),
    (-220.9569, 52.1696, 67.3273),
    (-220.9569, 51.9523, 70.4866),
    (-220.9569, 51.9332, 70.6256),
    (-220.9569, 51.7623, 72.2272),
    (-220.9569, 51.742, 72.3511),
    (-220.9569, 51.4963, 73.8963),
    (-220.9569, 50.5452, 80.0866),
    (-220.9569, 50.5276, 80.1852),
    (-220.9569, 49.6031, 85.7812),
    (-220.9569, 49.0205, 89.0606),
    (-220.9569, 48.5789, 91.5031),
    (-220.9569, 47.5566, 96.7646),
    (-220.9569, 47.5395, 96.8545),
    (-220.9569, 47.5314, 96.894),
    (-220.9569, 46.5202, 101.69),
    (-220.9569, 46.049, 103.7773),
    (-220.9569, 45.4769, 106.193),
    (-220.9569, 44.5666, 109.7922),
    (-220.9569, 44.4119, 110.3653),
    (-220.9569, 44.0006, 111.7964),
    (-220.9569, 43.313, 114.2323),
    (-220.9569, 43.0735, 115.0286),
    (-220.9569, 42.1766, 117.7828),
    (-220.9569, 41.5616, 119.5656),
    (-220.9569, 41.0086, 121.0137),
    (-220.9569, 40.0242, 123.4638),
    (-220.9569, 39.7936, 123.9658),
    (-220.9569, 39.0071, 125.606),
    (-220.9569, 38.4792, 126.724),
    (-220.9569, 38.3142, 127.0622),
    (-220.9569, 37.1204, 129.1233),
    (-220.9569, 36.5521, 130.086),
    (-220.9569, 35.7049, 131.2499),
    (-220.9569, 34.7312, 132.5842),
    (-220.9569, 34.2297, 133.1171),
    (-220.9569, 32.842, 134.618),
    (-220.9569, 32.691, 134.7378),
    (-220.9569, 31.998, 135.3125),
    (-220.9569, 31.0035, 136.143),
    (-220.9569, 30.8731, 136.2607),
    (-220.9569, 29.1717, 137.3124),
    (-220.9569, 28.8168, 137.5616),
    (-220.9569, 27.245, 138.2479),
    (-220.9569, 26.6657, 138.5547),
    (-220.9569, 25.2155, 138.9733),
    (-220.9569, 24.411, 139.2828),
    (-220.9569, 23.0077, 139.5193),
    (-220.9569, 21.9148, 139.8067),
    (-220.9569, 20.6277, 139.9045),
    (-220.9569, 19.2827, 140.1251),
    (-220.9569, 18.1016, 140.1378),
    (-220.9569, 16.5039, 140.2801),
    (-220.9569, 15.4164, 140.247),
    (-220.9569, 13.5669, 140.3129),
    (-220.9569, 12.5574, 140.2615),
    (-220.9569, 10.4603, 140.2644),
    (-220.9569, 9.5085, 140.2133),
    (-220.9569, 7.172, 140.1758),
    (-220.9569, 6.2517, 140.1377),
    (-220.9569, 3.6896, 140.0896),
    (-220.9569, 2.2472, 140.0613),
    (-220.9569, 0.0, 140.0418),
    (-220.9569, -10.0, 140.0423375),
    (-220.9569, -10.0, 110.0468),
    (-220.9569, 0.0, 110.0468),
    (-220.9569, 1.4118, 110.0442),
    (-220.9569, 3.6014, 110.0577),
    (-220.9569, 4.784, 110.0447),
    (-220.9569, 6.8901, 110.0124),
    (-220.9569, 7.9063, 109.9847),
    (-220.9569, 9.8721, 109.9162),
    (-220.9569, 10.7126, 109.871),
    (-220.9569, 12.5537, 109.7519),
    (-220.9569, 13.2188, 109.6892),
    (-220.9569, 14.941, 109.502),
    (-220.9569, 15.4394, 109.4252),
    (-220.9569, 17.0404, 109.1494),
    (-220.9569, 17.3872, 109.0654),
    (-220.9569, 18.8587, 108.677),
    (-220.9569, 19.0742, 108.5966),
    (-220.9569, 20.4027, 108.0683),
    (-220.9569, 20.5111, 108.0059),
    (-220.9569, 21.6794, 107.3068),
    (-220.9569, 21.708, 107.2811),
    (-220.9569, 22.176, 106.8528),
    (-220.9569, 22.6813, 106.3917),
    (-220.9569, 22.6961, 106.3777),
    (-220.9569, 23.097, 105.8614),
    (-220.9569, 23.1094, 105.8455),
    (-220.9569, 23.4442, 105.2911),
    (-220.9569, 23.4603, 105.2655),
    (-220.9569, 23.938, 104.048),
    (-220.9569, 23.9791, 103.9548),
    (-220.9569, 24.2237, 102.579),
    (-220.9569, 24.2597, 102.4316),
    (-220.9569, 24.2892, 100.8932),
    (-220.9569, 24.3094, 100.6822),
    (-220.9569, 24.1401, 98.9765),
    (-220.9569, 24.1349, 98.6932),
    (-220.9569, 23.7815, 96.8144),
    (-220.9569, 23.7429, 96.4514),
    (-220.9569, 23.2448, 94.4891),
    (-220.9569, 23.1713, 94.0637),
    (-220.9569, 22.5249, 91.9282),
    (-220.9569, 22.4118, 91.422),
    (-220.9569, 21.623, 89.1019),
    (-220.9569, 21.4686, 88.5146),
    (-220.9569, 20.5404, 85.9948),
    (-220.9569, 20.3449, 85.3289),
    (-220.9569, 19.2774, 82.5899),
    (-220.9569, 19.0432, 81.8518),
    (-220.9569, 17.833, 78.8686),
    (-220.9569, 17.5648, 78.0691),
    (-220.9569, 16.2045, 74.8099),
    (-220.9569, 15.9095, 73.9655),
    (-220.9569, 15.4494, 72.7447),
    (-220.9569, 14.0757, 69.5238),
    (-220.9569, 8.9633, 49.2583),
    (-220.9569, 4.1616, 30.2244),
    (-220.9569, 23.8504, 25.3966),
]
_loft5_c = [
    (-314.96, 19.5895, 24.6544),
    (-314.96, 61.4598, 24.6544),
    (-314.96, 62.5879, 29.1259),
    (-314.96, 63.357, 32.1747),
    (-314.96, 64.2474, 35.7044),
    (-314.96, 64.3894, 36.3862),
    (-314.96, 64.6517, 37.68),
    (-314.96, 65.4492, 41.5841),
    (-314.96, 65.85, 45.1518),
    (-314.96, 66.1348, 47.58),
    (-314.96, 66.2127, 50.7883),
    (-314.96, 66.3031, 53.5925),
    (-314.96, 66.1306, 56.3693),
    (-314.96, 65.9672, 59.5727),
    (-314.96, 65.9179, 60.0125),
    (-314.96, 65.6044, 61.9727),
    (-314.96, 65.0831, 65.789),
    (-314.96, 65.0418, 65.9965),
    (-314.96, 64.6388, 67.6014),
    (-314.96, 63.6729, 71.8657),
    (-314.96, 63.2599, 73.065),
    (-314.96, 61.7554, 77.7541),
    (-314.96, 61.6239, 78.0647),
    (-314.96, 60.6125, 80.6135),
    (-314.96, 60.5552, 80.7345),
    (-314.96, 60.54, 80.7647),
    (-314.96, 59.3934, 83.3641),
    (-314.96, 57.9282, 86.4471),
    (-314.96, 56.921, 88.3857),
    (-314.96, 55.1691, 91.9701),
    (-314.96, 54.2928, 93.5701),
    (-314.96, 52.5037, 97.0723),
    (-314.96, 50.6678, 100.4867),
    (-314.96, 49.9082, 101.7932),
    (-314.96, 48.0381, 105.084),
    (-314.96, 47.3627, 106.1639),
    (-314.96, 45.2923, 109.581),
    (-314.96, 44.6849, 110.4735),
    (-314.96, 42.5725, 113.7113),
    (-314.96, 40.4025, 116.7738),
    (-314.96, 39.8627, 117.4948),
    (-314.96, 37.6459, 120.353),
    (-314.96, 37.1482, 120.948),
    (-314.96, 34.8812, 123.5942),
    (-314.96, 34.4161, 124.0841),
    (-314.96, 32.0943, 126.513),
    (-314.96, 29.7044, 128.7284),
    (-314.96, 29.2726, 129.122),
    (-314.96, 26.8192, 131.1115),
    (-314.96, 26.4041, 131.4318),
    (-314.96, 23.7461, 133.2813),
    (-314.96, 23.342, 133.5381),
    (-314.96, 20.6008, 135.1356),
    (-314.96, 20.2022, 135.3366),
    (-314.96, 17.3713, 136.6826),
    (-314.96, 14.4413, 137.7794),
    (-314.96, 14.0459, 137.9293),
    (-314.96, 11.0135, 138.7772),
    (-314.96, 10.6132, 138.8818),
    (-314.96, 7.4715, 139.4833),
    (-314.96, 7.0618, 139.5451),
    (-314.96, 3.804, 139.9035),
    (-314.96, 2.3404, 139.9605),
    (-314.96, -0.0, 140.0426),
    (-314.96, -10.0, 140.0426),
    (-314.96, -10.0, 110.04285108295872),
    (-314.96, -0.0, 110.0428),
    (-314.96, 1.7794, 110.0141),
    (-314.96, 4.4188, 109.9331),
    (-314.96, 6.0449, 109.8454),
    (-314.96, 8.3636, 109.6017),
    (-314.96, 9.8409, 109.4535),
    (-314.96, 11.1722, 109.2438),
    (-314.96, 13.2008, 108.8347),
    (-314.96, 14.3944, 108.562),
    (-314.96, 16.1562, 107.9846),
    (-314.96, 17.2179, 107.6477),
    (-314.96, 18.1513, 107.2474),
    (-314.96, 19.6736, 106.4957),
    (-314.96, 20.4887, 106.0303),
    (-314.96, 21.7916, 105.1001),
    (-314.96, 22.495, 104.5692),
    (-314.96, 23.0919, 103.975),
    (-314.96, 24.2003, 102.8578),
    (-314.96, 24.4623, 102.5362),
    (-314.96, 24.7005, 102.1989),
    (-314.96, 25.6345, 100.8889),
    (-314.96, 26.045, 100.1661),
    (-314.96, 26.8278, 98.655),
    (-314.96, 27.156, 97.8694),
    (-314.96, 27.4086, 97.0231),
    (-314.96, 28.0643, 95.3012),
    (-314.96, 28.2498, 94.3952),
    (-314.96, 28.374, 93.4299),
    (-314.96, 29.4013, 89.3164),
    (-314.96, 29.4725, 88.3467),
    (-314.96, 29.5006, 87.3254),
    (-314.96, 29.4909, 86.2525),
    (-314.96, 30.3198, 81.1806),
    (-314.96, 30.2833, 80.0754),
    (-314.96, 30.22, 78.9213),
    (-314.96, 30.5115, 71.1307),
    (-314.96, 30.8088, 69.1278),
    (-314.96, 24.6024, 44.5256),
]

_l5_cleans = [_dedupe_3d(p) for p in [_loft4_d, _loft5_a, _loft5_b, _loft5_c]]

_LOFT5_N = 150
_l5_res = [_resample_profile(p, _LOFT5_N) for p in _l5_cleans]

print(f"Third loft: 4 profiles resampled to {_LOFT5_N} pts each")

_l5_wires = [Wire.make_polygon([Vector(*pt) for pt in r_pts], close=True) for r_pts in _l5_res]
_l5_lateral = Solid.make_loft(_l5_wires, ruled=True)

_l5_cap_start = _make_planar_cap(_l5_res[0])
_l5_cap_end = _make_planar_cap(_l5_res[-1])

_l5_sew = BRepBuilderAPI_Sewing(1e-2)
_l5_sew.Add(_l5_lateral.wrapped)
_l5_sew.Add(_l5_cap_start.wrapped)
_l5_sew.Add(_l5_cap_end.wrapped)
_l5_sew.Perform()
_l5_ocp = OCP_Solid()
_l5_bld = BRep_Builder()
_l5_bld.MakeSolid(_l5_ocp)
_l5_sewn = _l5_sew.SewedShape()
if _l5_sewn.ShapeType() == TopAbs_SHELL:
    _l5_bld.Add(_l5_ocp, _l5_sewn)
else:
    _l5_exp = TopExp_Explorer(_l5_sewn, TopAbs_SHELL)
    _l5_exp.Init(_l5_sewn, TopAbs_SHELL)
    _l5_bld.Add(_l5_ocp, _l5_exp.Current() if _l5_exp.More() else _l5_sewn)
_l5_solid = Solid(_l5_ocp)

with BuildPart() as loft_feature_strip2:
    add(_l5_solid)
        # Smoothening: Try to fillet the base edges at Y=-10.0
    _base_edges = loft_feature_strip2.edges().filter_by(lambda e: abs(e.center().Y + 10.0) < 0.2)
    if _base_edges:
        try:
            fillet(_base_edges, radius=1.0)
            print(f"Smoothened base of loft_feature_strip2")
        except:
            print(f"Smoothing fillet at base of loft_feature_strip2 skipped")
    

_l5p = loft_feature_strip2.part
if _l5p:
    print(f"Third loft (strip2) Is Manifold : {_l5p.is_manifold}")
    print(f"Third loft (strip2) Volume      : {_l5p.volume:.2f} mm³")
else:
    print("Error: No third loft geometry created.")



# ── Fourth loft (User provided 3 profiles) ────────────────────────────
_loft6_a = [
    (-362.0893, 33.4102, 24.6544),
    (-362.0893, 39.5468, 48.5074),
    (-362.0893, 40.8551, 54.9938),
    (-362.0893, 40.9077, 55.262),
    (-362.0893, 40.993, 56.1772),
    (-362.0893, 41.5809, 62.2351),
    (-362.0893, 41.5684, 63.1239),
    (-362.0893, 41.5218, 69.452),
    (-362.0893, 41.5206, 69.6323),
    (-362.0893, 41.0697, 73.2643),
    (-362.0893, 40.7223, 76.554),
    (-362.0893, 40.2038, 79.5581),
    (-362.0893, 40.1601, 79.7911),
    (-362.0893, 39.3893, 84.4944),
    (-362.0893, 39.3329, 84.8111),
    (-362.0893, 38.5851, 89.4192),
    (-362.0893, 38.5533, 89.5973),
    (-362.0893, 37.8157, 94.1025),
    (-362.0893, 37.8065, 94.1529),
    (-362.0893, 37.4975, 95.9851),
    (-362.0893, 37.0769, 98.4909),
    (-362.0893, 36.3732, 102.458),
    (-362.0893, 36.3352, 102.7023),
    (-362.0893, 35.6278, 106.4107),
    (-362.0893, 35.0037, 109.3866),
    (-362.0893, 34.9175, 109.8666),
    (-362.0893, 34.133, 113.2151),
    (-362.0893, 34.0033, 113.801),
    (-362.0893, 33.1741, 116.8706),
    (-362.0893, 32.3072, 119.5927),
    (-362.0893, 32.0967, 120.3336),
    (-362.0893, 31.1599, 122.8063),
    (-362.0893, 30.8664, 123.5827),
    (-362.0893, 29.8475, 125.8138),
    (-362.0893, 28.7554, 127.7698),
    (-362.0893, 28.3333, 128.594),
    (-362.0893, 27.1382, 130.3355),
    (-362.0893, 26.5783, 131.1254),
    (-362.0893, 25.2693, 132.6606),
    (-362.0893, 24.5428, 133.3877),
    (-362.0893, 22.9972, 134.8164),
    (-362.0893, 21.3182, 136.0105),
    (-362.0893, 20.3726, 136.6645),
    (-362.0893, 18.5254, 137.6551),
    (-362.0893, 17.3548, 138.191),
    (-362.0893, 15.3251, 138.9867),
    (-362.0893, 13.1126, 139.591),
    (-362.0893, 11.6787, 139.9936),
    (-362.0893, 9.2509, 140.42),
    (-362.0893, 7.5519, 140.6675),
    (-362.0893, 4.8902, 140.9208),
    (-362.0893, 2.9161, 141.0044),
    (-362.0893, 0.0, 141.1097),
    (-362.0893, -10.0, 141.10662493334112),
    (-362.0893, -10.0, 110.96275829019153),
    (-362.0893, -4.6091, 110.9551),
    (-362.0893, -0.0, 110.9583),
    (-362.0893, -0.0, 112.0223),
    (-362.0893, 0.6245, 112.0072),
    (-362.0893, 1.2472, 111.9957),
    (-362.0893, 2.4855, 111.9729),
    (-362.0893, 5.7067, 111.8675),
    (-362.0893, 7.9142, 111.7232),
    (-362.0893, 10.9393, 111.3733),
    (-362.0893, 12.8041, 111.1479),
    (-362.0893, 14.3139, 110.865),
    (-362.0893, 17.3036, 110.1638),
    (-362.0893, 18.416, 109.8618),
    (-362.0893, 19.2283, 109.5526),
    (-362.0893, 22.1107, 108.3513),
    (-362.0893, 22.5408, 108.1267),
    (-362.0893, 23.9145, 107.2617),
    (-362.0893, 24.0855, 107.1379),
    (-362.0893, 25.198, 106.1941),
    (-362.0893, 25.3246, 106.1037),
    (-362.0893, 25.3251, 106.1033),
    (-362.0893, 26.3385, 105.0909),
    (-362.0893, 26.3434, 105.0856),
    (-362.0893, 27.083, 103.9826),
    (-362.0893, 27.0901, 103.9712),
    (-362.0893, 27.5598, 102.7794),
    (-362.0893, 27.567, 102.7603),
    (-362.0893, 27.771, 101.4814),
    (-362.0893, 27.7756, 101.4528),
    (-362.0893, 27.7187, 100.0889),
    (-362.0893, 27.7178, 100.0489),
    (-362.0893, 27.4051, 98.6018),
    (-362.0893, 27.3954, 98.5484),
    (-362.0893, 26.8585, 97.0805),
    (-362.0893, 26.8368, 97.0119),
    (-362.0893, 26.0734, 95.4706),
    (-362.0893, 25.9861, 95.2917),
    (-362.0893, 24.6213, 93.1172),
    (-362.0893, 24.5306, 92.9746),
    (-362.0893, 22.7869, 90.6797),
    (-362.0893, 22.6557, 90.4998),
    (-362.0893, 20.5479, 88.0898),
    (-362.0893, 20.3679, 87.8674),
    (-362.0893, 17.9115, 85.348),
    (-362.0893, 17.6741, 85.078),
    (-362.0893, 14.8849, 82.4545),
    (-362.0893, 14.5814, 82.1316),
    (-362.0893, 11.4755, 79.4096),
    (-362.0893, 11.0969, 79.0284),
    (-362.0893, 7.6905, 76.2131),
    (-362.0893, 7.228, 75.7685),
    (-362.0893, 3.5369, 72.8649),
    (-362.0893, 2.9818, 72.3516),
    (-362.0893, 1.8688, 68.0253),
    (-362.0893, -3.5755, 46.8632),
    (-362.0893, -3.9293, 24.6544),
]

_loft6_b = [
    (-426.6227, 50.8858, 24.6544),
    (-426.6227, 52.7023, 31.7153),
    (-426.6227, 56.7579, 47.4793),
    (-426.6227, 57.2909, 50.1657),
    (-426.6227, 57.6638, 52.0632),
    (-426.6227, 58.2106, 54.8229),
    (-426.6227, 58.552, 58.8617),
    (-426.6227, 58.846, 62.3385),
    (-426.6227, 58.7272, 66.5362),
    (-426.6227, 58.626, 70.1136),
    (-426.6227, 58.4386, 71.5284),
    (-426.6227, 57.5641, 77.712),
    (-426.6227, 57.4832, 77.9423),
    (-426.6227, 57.3415, 78.6129),
    (-426.6227, 56.9974, 80.2654),
    (-426.6227, 55.8352, 85.8213),
    (-426.6227, 55.6148, 86.8436),
    (-426.6227, 54.3391, 92.7083),
    (-426.6227, 54.2308, 93.1889),
    (-426.6227, 52.8437, 99.2701),
    (-426.6227, 52.8381, 99.2936),
    (-426.6227, 52.7531, 99.6461),
    (-426.6227, 51.423, 105.1791),
    (-426.6227, 51.3411, 105.5183),
    (-426.6227, 49.986, 110.8037),
    (-426.6227, 49.8229, 111.4373),
    (-426.6227, 48.5188, 116.1639),
    (-426.6227, 48.2791, 117.0302),
    (-426.6227, 47.0139, 121.2531),
    (-426.6227, 46.7017, 122.2939),
    (-426.6227, 45.6021, 125.6616),
    (-426.6227, 45.269, 126.6839),
    (-426.6227, 43.9584, 130.3238),
    (-426.6227, 43.4852, 131.6457),
    (-426.6227, 42.1794, 134.8717),
    (-426.6227, 41.6349, 136.2317),
    (-426.6227, 40.3201, 139.0913),
    (-426.6227, 39.7082, 140.4451),
    (-426.6227, 38.3702, 142.9797),
    (-426.6227, 37.6946, 144.2904),
    (-426.6227, 36.319, 146.5352),
    (-426.6227, 35.5834, 147.7736),
    (-426.6227, 34.1554, 149.7581),
    (-426.6227, 33.3639, 150.9019),
    (-426.6227, 31.8681, 152.6502),
    (-426.6227, 31.0246, 153.6841),
    (-426.6227, 29.4454, 155.2155),
    (-426.6227, 28.5538, 156.13),
    (-426.6227, 26.7448, 157.552),
    (-426.6227, 25.7376, 158.397),
    (-426.6227, 23.7945, 159.6002),
    (-426.6227, 22.7397, 160.3039),
    (-426.6227, 20.6435, 161.2993),
    (-426.6227, 19.544, 161.8673),
    (-426.6227, 17.2769, 162.6637),
    (-426.6227, 16.1336, 163.1047),
    (-426.6227, 13.6801, 163.7096),
    (-426.6227, 12.4915, 164.0345),
    (-426.6227, 9.8389, 164.4554),
    (-426.6227, 8.6002, 164.6753),
    (-426.6227, 5.74, 164.9205),
    (-426.6227, 4.4422, 165.0461),
    (-426.6227, 2.8502, 165.056),
    (-426.6227, 0.0, 165.185),
    (-426.6227, -10.0, 165.16552840172014),
    (-426.6227, -10.0, 130.13999740524815),
    (-426.6227, -5.9004, 130.1371),
    (-426.6227, 0.0, 130.1382),
    (-426.6227, 1.5164, 130.0816),
    (-426.6227, 4.4166, 130.0924),
    (-426.6227, 6.955, 129.9972),
    (-426.6227, 8.4469, 129.9356),
    (-426.6227, 10.6845, 129.7711),
    (-426.6227, 12.1014, 129.6567),
    (-426.6227, 14.0375, 129.421),
    (-426.6227, 15.3908, 129.2406),
    (-426.6227, 17.0345, 128.9346),
    (-426.6227, 18.3262, 128.6719),
    (-426.6227, 19.6937, 128.2988),
    (-426.6227, 20.9185, 127.9352),
    (-426.6227, 22.0318, 127.5003),
    (-426.6227, 23.1789, 127.0149),
    (-426.6227, 24.0635, 126.5247),
    (-426.6227, 25.1189, 125.8954),
    (-426.6227, 25.7024, 125.4493),
    (-426.6227, 26.5028, 124.7961),
    (-426.6227, 26.9618, 124.3211),
    (-426.6227, 27.6774, 123.5385),
    (-426.6227, 28.021, 123.0504),
    (-426.6227, 28.629, 122.1497),
    (-426.6227, 28.8833, 121.6399),
    (-426.6227, 29.3933, 120.5949),
    (-426.6227, 29.571, 120.0642),
    (-426.6227, 29.9762, 118.8672),
    (-426.6227, 30.0884, 118.3144),
    (-426.6227, 30.3833, 116.9607),
    (-426.6227, 30.4387, 116.3816),
    (-426.6227, 30.6198, 114.8696),
    (-426.6227, 30.6247, 114.2567),
    (-426.6227, 30.6906, 112.5895),
    (-426.6227, 30.5602, 111.2146),
    (-426.6227, 30.3523, 107.4469),
    (-426.6227, 30.1802, 106.4768),
    (-426.6227, 29.8547, 103.9995),
    (-426.6227, 29.5982, 102.8618),
    (-426.6227, 29.1417, 100.2679),
    (-426.6227, 28.7794, 98.8996),
    (-426.6227, 28.2161, 96.2526),
    (-426.6227, 27.7172, 94.5748),
    (-426.6227, 27.0792, 91.956),
    (-426.6227, 26.4027, 89.8733),
    (-426.6227, 25.7307, 87.3827),
    (-426.6227, 24.8246, 84.7819),
    (-426.6227, 24.1684, 82.539),
    (-426.6227, 22.969, 79.289),
    (-426.6227, 22.3887, 77.4331),
    (-426.6227, 21.0648, 74.1248),
    (-426.6227, 20.3861, 72.0744),
    (-426.6227, 13.7701, 46.358),
    (-426.6227, 8.1866, 24.6544),
]

_loft6_c = [
    (-478.3154, 64.8842, 24.6544),
    (-478.3154, 70.8142, 47.7047),
    (-478.3154, 71.9154, 53.2039),
    (-478.3154, 72.2369, 54.8273),
    (-478.3154, 72.7256, 60.1767),
    (-478.3154, 72.8901, 62.0995),
    (-478.3154, 72.7954, 67.6509),
    (-478.3154, 72.7426, 69.6266),
    (-478.3154, 71.9422, 75.7729),
    (-478.3154, 71.7562, 77.0999),
    (-478.3154, 71.6212, 77.769),
    (-478.3154, 71.4634, 78.587),
    (-478.3154, 70.8555, 81.8195),
    (-478.3154, 70.5731, 83.3343),
    (-478.3154, 69.3874, 89.3554),
    (-478.3154, 68.6792, 93.0341),
    (-478.3154, 67.8784, 96.9437),
    (-478.3154, 66.7642, 102.5135),
    (-478.3154, 66.3535, 104.4309),
    (-478.3154, 64.8196, 111.7723),
    (-478.3154, 64.809, 111.8193),
    (-478.3154, 64.7669, 112.0115),
    (-478.3154, 63.2434, 118.9446),
    (-478.3154, 62.8397, 120.837),
    (-478.3154, 61.6528, 125.9584),
    (-478.3154, 60.8173, 129.6809),
    (-478.3154, 60.0332, 132.8722),
    (-478.3154, 58.7413, 138.3167),
    (-478.3154, 58.3814, 139.6912),
    (-478.3154, 57.0133, 145.1194),
    (-478.3154, 56.6454, 146.5747),
    (-478.3154, 56.5955, 146.7972),
    (-478.3154, 54.7639, 153.6702),
    (-478.3154, 54.4243, 155.0355),
    (-478.3154, 52.2248, 162.7963),
    (-478.3154, 50.2808, 169.0598),
    (-478.3154, 49.3132, 172.579),
    (-478.3154, 47.4551, 178.1107),
    (-478.3154, 45.9399, 182.9722),
    (-478.3154, 44.2163, 187.6743),
    (-478.3154, 42.6581, 191.4733),
    (-478.3154, 40.479, 197.6299),
    (-478.3154, 39.0981, 200.6732),
    (-478.3154, 36.1618, 207.7457),
    (-478.3154, 35.0103, 210.0131),
    (-478.3154, 33.9412, 211.8392),
    (-478.3154, 30.3297, 219.1735),
    (-478.3154, 29.4567, 220.4675),
    (-478.3154, 24.985, 227.766),
    (-478.3154, 24.4162, 228.4771),
    (-478.3154, 23.8709, 229.028),
    (-478.3154, 18.8198, 235.4),
    (-478.3154, 18.5612, 235.6074),
    (-478.3154, 12.7216, 240.7422),
    (-478.3154, 12.7142, 240.7466),
    (-478.3154, 12.7067, 240.7497),
    (-478.3154, 12.4352, 240.8901),
    (-478.3154, 6.7236, 241.4001),
    (-478.3154, 6.3439, 241.417),
    (-478.3154, 0.5452, 241.5799),
    (-478.3154, 0.0, 241.5874),
    (-478.3154, 0.0, 244.9038),
    (-478.3154, -10.0, 244.9320974375098),
    (-478.3154, -10.0, 227.68627083791856),
    (-478.3154, 0.0, 184.4808),
    (-478.3154, 0.0273, 184.4788),
    (-478.3154, 3.8479, 184.3563),
    (-478.3154, 4.4146, 184.3592),
    (-478.3154, 4.829, 184.3056),
    (-478.3154, 8.4459, 183.6421),
    (-478.3154, 8.8752, 183.5377),
    (-478.3154, 12.1149, 182.4424),
    (-478.3154, 12.584, 182.2619),
    (-478.3154, 15.4421, 180.7651),
    (-478.3154, 15.9675, 180.4708),
    (-478.3154, 18.4478, 178.6213),
    (-478.3154, 19.0376, 178.1631),
    (-478.3154, 21.1514, 176.0266),
    (-478.3154, 21.8063, 175.3429),
    (-478.3154, 23.5715, 173.0),
    (-478.3154, 24.2859, 172.0184),
    (-478.3154, 25.7261, 169.5618),
    (-478.3154, 26.3646, 168.4474),
    (-478.3154, 27.3354, 166.3874),
    (-478.3154, 27.9982, 164.9579),
    (-478.3154, 28.7803, 162.9517),
    (-478.3154, 29.4417, 161.1998),
    (-478.3154, 30.0411, 159.3548),
    (-478.3154, 30.7063, 157.1994),
    (-478.3154, 31.1613, 155.5293),
    (-478.3154, 31.82, 152.9166),
    (-478.3154, 32.1481, 151.4841),
    (-478.3154, 32.79, 148.3591),
    (-478.3154, 33.0081, 147.2278),
    (-478.3154, 33.6233, 143.5339),
    (-478.3154, 33.7477, 142.768),
    (-478.3154, 34.3265, 138.4477),
    (-478.3154, 34.3724, 138.1117),
    (-478.3154, 34.6838, 134.8417),
    (-478.3154, 35.3012, 128.271),
    (-478.3154, 35.3778, 126.7399),
    (-478.3154, 35.6617, 122.0492),
    (-478.3154, 35.7517, 118.8198),
    (-478.3154, 35.8825, 115.5704),
    (-478.3154, 35.9321, 110.3349),
    (-478.3154, 35.9669, 108.8447),
    (-478.3154, 35.9372, 103.8381),
    (-478.3154, 35.9169, 101.8848),
    (-478.3154, 35.8936, 101.3477),
    (-478.3154, 35.7328, 94.7006),
    (-478.3154, 35.5792, 91.9708),
    (-478.3154, 35.4101, 87.2932),
    (-478.3154, 35.052, 81.9878),
    (-478.3154, 34.944, 79.6735),
    (-478.3154, 34.9236, 78.8287),
    (-478.3154, 34.3273, 71.8523),
    (-478.3154, -10.0, 92.61669390398951),
    (-478.3154, -10.0, 62.10289064190013),
    (-478.3154, 22.1849, 24.6544),
]

_loft6_d = [
    (-508.3819, 30.3269, 24.6544),
    (-508.3819, 32.4209, 32.7941),
    (-508.3819, 73.0262, 24.6544),
    (-508.3819, 79.0158, 47.9363),
    (-508.3819, 79.7585, 51.6455),
    (-508.3819, 80.414, 54.8753),
    (-508.3819, 80.7399, 58.2029),
    (-508.3819, 81.0815, 61.9393),
    (-508.3819, 81.0599, 65.411),
    (-508.3819, 80.9944, 69.2541),
    (-508.3819, 80.6065, 72.6966),
    (-508.3819, 80.1142, 76.5264),
    (-508.3819, 79.2491, 81.2587),
    (-508.3819, 77.6376, 90.6694),
    (-508.3819, 76.6799, 95.7613),
    (-508.3819, 74.3694, 108.499),
    (-508.3819, 73.6776, 112.0616),
    (-508.3819, 73.0932, 114.9349),
    (-508.3819, 70.1006, 130.4485),
    (-508.3819, 69.8308, 131.7226),
    (-508.3819, 69.5982, 132.7589),
    (-508.3819, 67.2739, 143.9831),
    (-508.3819, 63.036, 147.2377),
    (-508.3819, 60.3962, 149.1633),
    (-508.3819, 54.5811, 153.4604),
    (-508.3819, 50.4217, 156.3845),
    (-508.3819, 50.0018, 156.6745),
    (-508.3819, 43.426, 161.1547),
    (-508.3819, 40.7928, 162.9091),
    (-508.3819, 35.9171, 165.927),
    (-508.3819, 37.2544, 155.8921),
    (-508.3819, 37.8207, 150.769),
    (-508.3819, 38.6704, 143.1557),
    (-508.3819, 39.2337, 137.0458),
    (-508.3819, 39.5142, 134.1926),
    (-508.3819, 40.0532, 127.4479),
    (-508.3819, 40.1143, 126.6867),
    (-508.3819, 40.1236, 126.5146),
    (-508.3819, 40.7461, 117.7806),
    (-508.3819, 40.8461, 115.5322),
    (-508.3819, 41.2797, 108.7766),
    (-508.3819, 41.4433, 104.327),
    (-508.3819, 41.7168, 99.6717),
    (-508.3819, 41.9239, 92.897),
    (-508.3819, 42.0571, 90.4626),
    (-508.3819, 42.2936, 81.2384),
    (-508.3819, 42.2984, 81.147),
    (-508.3819, 42.3025, 81.0432),
    (-508.3819, 42.436, 71.7232),
]


_l6_cleans = [_dedupe_3d(p) for p in [_loft5_c, _loft6_a, _loft6_b, _loft6_c, _loft6_d]]

_LOFT6_N = 150
_l6_res = [_resample_profile(p, _LOFT6_N) for p in _l6_cleans]

print(f"Fourth loft: 5 profiles resampled to {_LOFT6_N} pts each")

_l6_wires = [Wire.make_polygon([Vector(*pt) for pt in r_pts], close=True) for r_pts in _l6_res]
_l6_lateral = Solid.make_loft(_l6_wires, ruled=True)

_l6_cap_start = _make_planar_cap(_l6_res[0])
_l6_cap_end = _make_planar_cap(_l6_res[-1])

_l6_sew = BRepBuilderAPI_Sewing(1e-2)
_l6_sew.Add(_l6_lateral.wrapped)
_l6_sew.Add(_l6_cap_start.wrapped)
_l6_sew.Add(_l6_cap_end.wrapped)
_l6_sew.Perform()
_l6_ocp = OCP_Solid()
_l6_bld = BRep_Builder()
_l6_bld.MakeSolid(_l6_ocp)
_l6_sewn = _l6_sew.SewedShape()
if _l6_sewn.ShapeType() == TopAbs_SHELL:
    _l6_bld.Add(_l6_ocp, _l6_sewn)
else:
    _l6_exp = TopExp_Explorer(_l6_sewn, TopAbs_SHELL)
    _l6_exp.Init(_l6_sewn, TopAbs_SHELL)
    _l6_bld.Add(_l6_ocp, _l6_exp.Current() if _l6_exp.More() else _l6_sewn)
_l6_solid = Solid(_l6_ocp)

with BuildPart() as loft_feature_strip3:
    add(_l6_solid)
        # Smoothening: Try to fillet the base edges at Y=-10.0
    _base_edges = loft_feature_strip3.edges().filter_by(lambda e: abs(e.center().Y + 10.0) < 0.2)
    if _base_edges:
        try:
            fillet(_base_edges, radius=1.0)
            print(f"Smoothened base of loft_feature_strip3")
        except:
            print(f"Smoothing fillet at base of loft_feature_strip3 skipped")
    

_l6p = loft_feature_strip3.part
if _l6p:
    print(f"Fourth loft (strip3) Is Manifold : {_l6p.is_manifold}")
    print(f"Fourth loft (strip3) Volume      : {_l6p.volume:.2f} mm³")
else:
    print("Error: No fourth loft geometry created.")



# ── Fifth loft (User provided 5 profiles) ────────────────────────────
_loft7_a = [
    (-701.6346, 38.7447, 286.4312),
    (-701.6346, 38.624, 286.6787),
    (-701.6346, 37.1605, 289.4416),
    (-701.6346, 37.1028, 289.6115),
    (-701.6346, 37.0633, 289.7278),
    (-701.6346, 36.3762, 291.757),
    (-701.6346, 35.8437, 293.3312),
    (-701.6346, 35.6476, 293.9123),
    (-701.6346, 35.3534, 294.7843),
    (-701.6346, 35.0418, 296.1309),
    (-701.6346, 34.868, 296.8878),
    (-701.6346, 34.3584, 299.1095),
    (-701.6346, 33.8953, 301.146),
    (-701.6346, 33.6853, 302.5477),
    (-701.6346, 33.4877, 303.896),
    (-701.6346, 33.1711, 306.0624),
    (-701.6346, 32.8187, 308.5227),
    (-701.6346, 32.7285, 309.6121),
    (-701.6346, 32.6598, 310.4817),
    (-701.6346, 32.4407, 313.2714),
    (-701.6346, 32.1675, 316.9112),
    (-701.6346, 32.1664, 317.0099),
    (-701.6346, 32.1717, 318.1261),
    (-701.6346, 25.4683, 318.6488),
    (-701.6346, 18.9078, 319.1423),
    (-701.6346, 17.2917, 319.2998),
    (-701.6346, 10.1413, 319.8342),
    (-701.6346, 9.4225, 319.9075),
    (-701.6346, 2.5929, 320.4004),
    (-701.6346, 2.1491, 320.4444),
    (-701.6346, -6.3457, 321.0196),
    (-701.6346, -6.0321, 320.9907),
    (-701.6346, -10.0, 321.2405151964769),
    (-701.6346, -10.0, 287.241),
    (-701.6346, -0.6962, 287.241),
    (-701.6346, 1.5561, 287.1418),
    (-701.6346, 4.3735, 287.0142),
    (-701.6346, 11.8023, 287.0127),
    (-701.6346, 14.6437, 286.8793),
    (-701.6346, 17.6588, 286.7382),
    (-701.6346, 24.8673, 286.7347),
    (-701.6346, 27.0949, 286.6296),
    (-701.6346, 33.0245, 286.4813),
    (-701.6346, 35.9688, 286.4776),
]

_loft7_b = [
    (-801.3863, 62.165, 328.1709),
    (-801.3863, 62.3548, 328.4389),
    (-801.3863, 80.1287, 354.8877),
    (-801.3863, 83.3671, 359.3833),
    (-801.3863, 79.3691, 360.1089),
    (-801.3863, 68.6171, 361.9578),
    (-801.3863, 63.9344, 362.7871),
    (-801.3863, 51.7368, 364.9124),
    (-801.3863, 48.8914, 365.4172),
    (-801.3863, 35.2501, 367.8073),
    (-801.3863, 33.3422, 368.1463),
    (-801.3863, 19.0742, 370.6569),
    (-801.3863, 18.0463, 370.8397),
    (-801.3863, 11.0812, 372.0675),
    (-801.3863, 9.7357, 372.3068),
    (-801.3863, 3.1366, 373.4707),
    (-801.3863, 0.9421, 373.8611),
    (-801.3863, -5.9214, 375.0717),
    (-801.3863, -9.2289, 375.6598),
    (-801.3863, -10.0, 375.7957267296641),
    (-801.3863, -10.0, 337.8742679948076),
    (-801.3863, -8.7663, 337.7018),
    (-801.3863, -4.1735, 337.0844),
    (-801.3863, -1.3086, 336.6883),
    (-801.3863, 3.2398, 336.0828),
    (-801.3863, 6.3944, 335.6507),
    (-801.3863, 10.7128, 335.0805),
    (-801.3863, 14.3558, 334.5853),
    (-801.3863, 18.2556, 334.074),
    (-801.3863, 22.5856, 333.4891),
    (-801.3863, 25.8761, 333.0602),
    (-801.3863, 31.0923, 332.3594),
    (-801.3863, 33.5829, 332.0364),
    (-801.3863, 48.5787, 330.0372),
    (-801.3863, 49.3098, 329.9426),
    (-801.3863, 49.6672, 329.897),
    (-801.3863, 58.7631, 328.6444),
    (-801.3863, 62.0028, 328.1955),
]

_loft7_c = [
    (-861.2342, 102.7411, 367.8628),
    (-861.2342, 111.5762, 376.3369),
    (-861.2342, 116.8453, 381.0835),
    (-861.2342, 122.6392, 386.488),
    (-861.2342, 131.2258, 394.8785),
    (-861.2342, 130.4952, 395.3013),
    (-861.2342, 129.501, 395.8201),
    (-861.2342, 129.2077, 395.9441),
    (-861.2342, 117.7419, 398.292),
    (-861.2342, 113.3998, 399.1794),
    (-861.2342, 112.3843, 399.383),
    (-861.2342, 93.6976, 403.1239),
    (-861.2342, 90.1829, 403.8167),
    (-861.2342, 73.5714, 407.0635),
    (-861.2342, 68.5133, 408.0374),
    (-861.2342, 52.2397, 411.1259),
    (-861.2342, 47.2605, 412.0576),
    (-861.2342, 38.7593, 413.63),
    (-861.2342, 36.7612, 413.9943),
    (-861.2342, 27.4967, 415.6745),
    (-861.2342, 26.3259, 415.8839),
    (-861.2342, 16.6202, 417.6036),
    (-861.2342, 14.4167, 417.9909),
    (-861.2342, 14.1135, 418.0434),
    (-861.2342, 2.7789, 419.9781),
    (-861.2342, 0.3198, 420.3924),
    (-861.2342, -8.2382, 421.8085),
    (-861.2342, -10.0, 422.09705853736597),
    (-861.2342, -10.0, 387.9698871736938),
    (-861.2342, -7.7282, 387.6065),
    (-861.2342, -3.1481, 386.8941),
    (-861.2342, 1.3036, 386.1658),
    (-861.2342, 6.4235, 385.3502),
    (-861.2342, 10.664, 384.6408),
    (-861.2342, 16.0402, 383.7638),
    (-861.2342, 20.3895, 383.0202),
    (-861.2342, 25.7173, 382.1311),
    (-861.2342, 30.5127, 381.2946),
    (-861.2342, 35.4726, 380.4491),
    (-861.2342, 41.0588, 379.4563),
    (-861.2342, 45.3209, 378.7153),
    (-861.2342, 52.0468, 377.4994),
    (-861.2342, 55.2736, 376.9282),
    (-861.2342, 63.4924, 375.4187),
    (-861.2342, 65.3438, 375.0855),
    (-861.2342, 75.7221, 373.1418),
    (-861.2342, 84.9806, 371.3775),
    (-861.2342, 89.849, 370.4122),
]

_loft7_d = [
    (-911.6974, 84.4743, 372.1264),
    (-911.6974, 101.228, 372.7573),
    (-911.6974, 127.5066, 373.7469),
    (-911.6974, 131.0135, 373.879),
    (-911.6974, 119.7713, 401.9796),
    (-911.6974, 117.973, 405.5912),
    (-911.6974, 116.7832, 407.9707),
    (-911.6974, 116.0499, 409.4388),
    (-911.6974, 113.2888, 413.6352),
    (-911.6974, 113.1889, 413.7871),
    (-911.6974, 113.158, 413.8343),
    (-911.6974, 110.6945, 416.9719),
    (-911.6974, 109.8606, 418.0376),
    (-911.6974, 108.0286, 420.0149),
    (-911.6974, 106.2125, 421.9831),
    (-911.6974, 105.204, 422.9067),
    (-911.6974, 102.2647, 425.6124),
    (-911.6974, 102.2199, 425.6454),
    (-911.6974, 102.1272, 425.7143),
    (-911.6974, 95.9195, 430.3273),
    (-911.6974, 87.9493, 434.6542),
    (-911.6974, 87.7648, 434.7546),
    (-911.6974, 87.6993, 434.7778),
    (-911.6974, 82.481, 436.5695),
    (-911.6974, 79.5145, 437.5878),
    (-911.6974, 79.4623, 437.6),
    (-911.6974, 75.6873, 438.4984),
    (-911.6974, 73.8357, 438.9224),
    (-911.6974, 66.4336, 440.6443),
    (-911.6974, 61.1531, 441.8062),
    (-911.6974, 55.2927, 443.1138),
    (-911.6974, 47.4523, 444.7677),
    (-911.6974, 43.8222, 445.5437),
    (-911.6974, 34.4432, 447.442),
    (-911.6974, 32.0799, 447.9268),
    (-911.6974, 22.1139, 449.8629),
    (-911.6974, 20.1308, 450.2538),
    (-911.6974, 11.9878, 451.7786),
    (-911.6974, 10.1155, 452.135),
    (-911.6974, 2.558, 453.5059),
    (-911.6974, 0.0607, 453.9675),
    (-911.6974, -6.0318, 455.0419),
    (-911.6974, -9.0994, 455.595),
    (-911.6974, -10.0, 455.7498718623962),
    (-911.6974, -10.0, 423.49142162216623),
    (-911.6974, -1.4686, 421.9966),
    (-911.6974, -0.553, 421.8342),
    (-911.6974, 8.196, 420.2223),
    (-911.6974, 8.8381, 420.1023),
    (-911.6974, 17.5763, 418.4),
    (-911.6974, 17.7242, 418.3709),
    (-911.6974, 19.5203, 417.9994),
    (-911.6974, 26.1744, 416.628),
    (-911.6974, 26.6093, 416.5332),
    (-911.6974, 34.1475, 414.8809),
    (-911.6974, 35.2435, 414.6226),
    (-911.6974, 41.6112, 413.1316),
    (-911.6974, 43.434, 412.662),
    (-911.6974, 48.6961, 411.3369),
    (-911.6974, 51.3732, 410.5727),
    (-911.6974, 55.2895, 409.5028),
    (-911.6974, 58.7794, 408.3832),
    (-911.6974, 60.1798, 407.9717),
    (-911.6974, 62.2764, 407.2378),
    (-911.6974, 63.3133, 406.9181),
    (-911.6974, 65.6331, 406.0497),
    (-911.6974, 66.3069, 405.8311),
    (-911.6974, 68.8475, 404.8111),
    (-911.6974, 69.4043, 404.5986),
    (-911.6974, 71.9174, 403.513),
    (-911.6974, 77.6555, 389.1704),
    (-911.6974, 79.2835, 385.101),
    (-911.6974, 80.0258, 383.2456),
    (-911.6974, 81.0184, 380.7645),
]

_loft7_e = [
    (-956.1491, 50.1523, 404.0852),
    (-956.1491, 54.7443, 404.2582),
    (-956.1491, 58.7962, 408.6097),
    (-956.1491, 60.7725, 410.7321),
    (-956.1491, 74.6389, 411.2543),
    (-956.1491, 79.1041, 411.4225),
    (-956.1491, 79.6109, 410.1558),
    (-956.1491, 81.5663, 405.2682),
    (-956.1491, 86.9381, 405.4705),
    (-956.1491, 96.6915, 405.8378),
    (-956.1491, 86.135, 432.2246),
    (-956.1491, 84.8704, 435.3854),
    (-956.1491, 82.0681, 441.0598),
    (-956.1491, 81.2671, 442.6748),
    (-956.1491, 79.5368, 445.3343),
    (-956.1491, 78.5313, 446.8744),
    (-956.1491, 78.203, 447.298),
    (-956.1491, 75.4338, 450.8562),
    (-956.1491, 74.0526, 452.3707),
    (-956.1491, 72.0108, 454.5981),
    (-956.1491, 69.3979, 457.0388),
    (-956.1491, 68.3082, 458.0504),
    (-956.1491, 64.3671, 461.0364),
    (-956.1491, 62.3548, 462.5426),
    (-956.1491, 60.9793, 463.3062),
    (-956.1491, 54.6899, 466.7655),
    (-956.1491, 50.8661, 468.1332),
    (-956.1491, 46.862, 469.5414),
    (-956.1491, 43.021, 470.5636),
    (-956.1491, 42.1141, 470.79),
    (-956.1491, 40.9239, 471.0829),
    (-956.1491, 33.2883, 472.932),
    (-956.1491, 30.2931, 473.6327),
    (-956.1491, 24.1296, 475.0515),
    (-956.1491, 19.7099, 476.036),
    (-956.1491, 14.689, 477.1373),
    (-956.1491, 9.2424, 478.2943),
    (-956.1491, 5.0214, 479.1787),
    (-956.1491, 0.0, 564.8082),
    (-956.1491, -10.0, 482.17825658307214),
    (-956.1491, -8.3503, 481.8568),
    (-956.1491, -3.1264, 480.8307),
    (-956.1491, 0.0054, 480.2001),
    (-956.1491, -10.0, 453.61869012089386),
    (-956.1491, -9.1635, 453.4357),
    (-956.1491, -5.9033, 452.6399),
    (-956.1491, -2.8539, 451.9289),
    (-956.1491, 1.0907, 450.8958),
    (-956.1491, 2.9831, 450.418),
    (-956.1491, 7.6851, 449.078),
    (-956.1491, 8.3568, 448.8916),
    (-956.1491, 10.942, 448.0794),
    (-956.1491, 13.342, 447.3155),
    (-956.1491, 13.845, 447.1358),
    (-956.1491, 17.966, 445.6791),
    (-956.1491, 19.5382, 445.0546),
    (-956.1491, 22.2195, 443.9914),
    (-956.1491, 24.8906, 442.8003),
    (-956.1491, 26.0658, 442.2728),
    (-956.1491, 28.8835, 440.8507),
    (-956.1491, 29.376, 440.6028),
    (-956.1491, 29.698, 440.4305),
    (-956.1491, 31.1165, 439.6447),
    (-956.1491, 31.889, 439.196),
    (-956.1491, 32.6305, 438.7595),
    (-956.1491, 33.937, 437.9538),
    (-956.1491, 34.0395, 437.8897),
    (-956.1491, 34.2558, 437.7479),
    (-956.1491, 35.3935, 437.0057),
    (-956.1491, 35.8391, 436.7051),
    (-956.1491, 36.4283, 436.3035),
    (-956.1491, 37.5955, 435.4718),
    (-956.1491, 39.4668, 430.7942),
    (-956.1491, 42.5344, 423.1267),
]


_l7_cleans = [_dedupe_3d(p) for p in [_loft6_d, _loft7_a, _loft7_b, _loft7_c, _loft7_d, _loft7_e]]

_LOFT7_N = 150
_l7_res = [_resample_profile(p, _LOFT7_N) for p in _l7_cleans]

print(f"Fifth loft: 6 profiles resampled to {_LOFT7_N} pts each")

_l7_wires = [Wire.make_polygon([Vector(*pt) for pt in r_pts], close=True) for r_pts in _l7_res]
_l7_lateral = Solid.make_loft(_l7_wires, ruled=True)

_l7_cap_start = _make_planar_cap(_l7_res[0])
_l7_cap_end = _make_planar_cap(_l7_res[-1])

_l7_sew = BRepBuilderAPI_Sewing(1e-2)
_l7_sew.Add(_l7_lateral.wrapped)
_l7_sew.Add(_l7_cap_start.wrapped)
_l7_sew.Add(_l7_cap_end.wrapped)
_l7_sew.Perform()

_l7_ocp = OCP_Solid()
_l7_bld = BRep_Builder()
_l7_bld.MakeSolid(_l7_ocp)
_l7_sewn = _l7_sew.SewedShape()

if _l7_sewn.ShapeType() == TopAbs_SHELL:
    _l7_bld.Add(_l7_ocp, _l7_sewn)
else:
    _l7_exp = TopExp_Explorer(_l7_sewn, TopAbs_SHELL)
    _l7_exp.Init(_l7_sewn, TopAbs_SHELL)
    _l7_bld.Add(_l7_ocp, _l7_exp.Current() if _l7_exp.More() else _l7_sewn)

_l7_solid = Solid(_l7_ocp)

with BuildPart() as loft_feature_strip4:
    add(_l7_solid)
        # Smoothening: Try to fillet the base edges at Y=-10.0
    _base_edges = loft_feature_strip4.edges().filter_by(lambda e: abs(e.center().Y + 10.0) < 0.2)
    if _base_edges:
        try:
            fillet(_base_edges, radius=1.0)
            print(f"Smoothened base of loft_feature_strip4")
        except:
            print(f"Smoothing fillet at base of loft_feature_strip4 skipped")
    

_l7p = loft_feature_strip4.part
if _l7p:
    print(f"Fifth loft (strip4) Is Manifold : {_l7p.is_manifold}")
    print(f"Fifth loft (strip4) Volume      : {_l7p.volume:.2f} mm³")
else:
    print("Error: No fifth loft geometry created.")




# ── Combined Sixth and Seventh loft (Continuous segment) ─────────────
# Bridging from _loft8_a (X=77) through _loft4_d (X=-132) to _loft9_d (X=-349)
_loft8_a = [
    (77.2118, 74.16, -470.6162), (77.2118, 75.0384, -475.6955), (77.2118, 75.1686, -477.5402), (77.2118, 75.2075, -480.2351), (77.2118, 75.1487, -484.1724), (77.2118, 74.8144, -487.0744), (77.2118, 74.3969, -489.7848), (77.2118, 73.8361, -492.538), (77.2118, 73.4244, -494.0512), (77.2118, 72.7752, -496.1149), (77.2118, 71.8021, -498.9084), (77.2118, 71.2415, -500.2347), (77.2118, 70.1829, -502.4323), (77.2118, 68.4215, -505.8282), (77.2118, 65.4128, -510.1415), (77.2118, 64.1365, -511.9982), (77.2118, 64.0121, -512.1497), (77.2118, 63.568, -512.7081), (77.2118, 62.6352, -513.776), (77.2118, 57.1114, -520.0841), (77.2118, 56.5806, -520.6423), (77.2118, 56.0566, -521.1546), (77.2118, 50.3518, -526.7963), (77.2118, 49.6402, -527.4268), (77.2118, 43.6428, -532.3462), (77.2118, 42.9897, -532.8534), (77.2118, 42.338, -533.2897), (77.2118, 36.0, -537.3849), (77.2118, 35.391, -537.7299), (77.2118, 34.8928, -537.9678), (77.2118, 28.0965, -541.2031), (77.2118, 27.6225, -541.394), (77.2118, 26.9953, -541.596), (77.2118, 19.634, -543.9345), (77.2118, 19.0197, -544.0814), (77.2118, 18.3839, -544.186), (77.2118, 10.3548, -545.6001), (77.2118, 9.7134, -545.6636), (77.2118, 8.4345, -545.7135), (77.2118, -0.0, -546.1672), (77.2118, -8.5307, -546.1672), (77.2118, -10.0, -515.2658399376711), (77.2118, -8.6698, -506.1671), (77.2118, -10.0, -485.21237268486243), (77.2118, 0.2793, -506.1637), (77.2118, -0.0, -506.1671), (77.2118, 6.0235, -506.0031), (77.2118, 6.2779, -505.9928), (77.2118, 11.2523, -505.5109), (77.2118, 11.4823, -505.4932), (77.2118, 11.6917, -505.4674), (77.2118, 15.9461, -504.6645), (77.2118, 16.1334, -504.6305), (77.2118, 19.7228, -503.5061), (77.2118, 19.8889, -503.4637), (77.2118, 20.0384, -503.4112), (77.2118, 23.0117, -501.9663), (77.2118, 23.1422, -501.9046), (77.2118, 25.5551, -500.1376), (77.2118, 25.647, -500.0813), (77.2118, 25.7439, -500.0047), (77.2118, 27.6462, -497.917), (77.2118, 27.7287, -497.8317), (77.2118, 27.8006, -497.7339), (77.2118, 29.3036, -495.2194), (77.2118, 29.2441, -495.3264), (77.2118, 30.343, -492.4875), (77.2118, 30.3914, -492.3715), (77.2118, 30.4316, -492.2408), (77.2118, 31.1166, -489.1888), (77.2118, 31.1478, -489.049), (77.2118, 31.5318, -485.6697), (77.2118, 31.5554, -485.5212), (77.2118, 31.5732, -485.3562), (77.2118, 31.7066, -481.6557), (77.2118, 31.7191, -481.4823), (77.2118, 31.7276, -481.2914), (77.2118, 31.4499, -472.9023), (77.2118, 31.455, -472.715), (77.2118, 31.4582, -472.5102), (77.2118, 31.4601, -472.2874), (77.2118, 31.3464, -470.6162)
]
_loft8_b = [
    (12.3625, 12.2359, -570.6775), (12.3625, 12.6375, -570.6272), (12.3625, 13.0645, -570.5759), (12.3625, 17.5137, -569.8059), (12.3625, 18.0664, -569.6961), (12.3625, 18.591, -569.5855), (12.3625, 23.4619, -568.1751), (12.3625, 24.1611, -567.9255), (12.3625, 24.5028, -567.7946), (12.3625, 28.6186, -565.8672), (12.3625, 29.5786, -565.3423), (12.3625, 30.1258, -565.0194), (12.3625, 34.0264, -562.3391), (12.3625, 35.1278, -561.4398), (12.3625, 35.7335, -560.8879), (12.3625, 39.5168, -556.7645), (12.3625, 40.5401, -555.4414), (12.3625, 41.1347, -554.5801), (12.3625, 45.4312, -547.4586), (12.3625, 46.1264, -546.0693), (12.3625, 46.5405, -545.1481), (12.3625, 52.8809, -531.0663), (12.3625, 53.3082, -530.1039), (12.3625, 53.7291, -529.136), (12.3625, 59.2062, -516.3262), (12.3625, 59.601, -515.3626), (12.3625, 66.8617, -497.0272), (12.3625, 67.2415, -496.0315), (12.3625, 67.6366, -494.9575), (12.3625, 75.1444, -473.1369), (12.3625, 75.4855, -472.0722), (12.3625, 75.8407, -470.9238), (12.3625, 83.2873, -470.6162), (12.3625, 31.3464, -470.6162), (12.3625, 31.428, -472.2874), (12.3625, 31.43, -472.5102), (12.3625, 31.4323, -472.715), (12.3625, 31.436, -472.9023), (12.3625, 31.2366, -481.2914), (12.3625, 31.2427, -481.4823), (12.3625, 31.2517, -481.6557), (12.3625, 31.3475, -485.3562), (12.3625, 31.3604, -485.5212), (12.3625, 31.3773, -485.6697), (12.3625, 31.652, -489.049), (12.3625, 31.6744, -489.1888), (12.3625, 32.1643, -492.2408), (12.3625, 32.193, -492.3715), (12.3625, 32.2276, -492.4875), (12.3625, 33.0135, -495.3264), (12.3625, 32.9709, -495.2194), (12.3625, 34.0401, -497.7339), (12.3625, 34.0915, -497.8317), (12.3625, 34.1506, -497.917), (12.3625, 35.5117, -500.0047), (12.3625, 35.581, -500.0813), (12.3625, 35.6468, -500.1376), (12.3625, 37.3739, -501.9046), (12.3625, 37.4673, -501.9663), (12.3625, 39.594, -503.4112), (12.3625, 39.7011, -503.4637), (12.3625, 39.8199, -503.5061), (12.3625, 42.3879, -504.6305), (12.3625, 42.5219, -504.6645), (12.3625, 45.429, -505.4674), (12.3625, 45.5788, -505.4932), (12.3625, 45.7433, -505.5109), (12.3625, 49.3, -505.9928), (12.3625, 49.482, -506.0031), (12.3625, -0.0, -506.1671), (12.3625, 0.2, -506.1637), (12.3625, -10.0, -485.21237268486243), (12.3625, -6.199, -506.1671), (12.3625, -10.0, -515.2658399376711), (12.3625, -6.1, -546.1672), (12.3625, -0.0, -546.3361), (12.3625, 3.8613, -546.2762), (12.3625, 4.1645, -546.2757), (12.3625, 4.4519, -546.2733), (12.3625, 8.2111, -546.2592), (12.3625, 8.5096, -546.2519), (12.3625, 14.0323, -546.1611)
]
_loft8_c = [
    (-25.0349, -0.0, -571.218), (-25.0349, 14.3642, -570.6775), (-25.0349, 14.8344, -570.6272), (-25.0349, 15.3343, -570.5759), (-25.0349, 20.5422, -569.8059), (-25.0349, 21.1892, -569.6961), (-25.0349, 21.8033, -569.5855), (-25.0349, 27.505, -568.1751), (-25.0349, 28.3235, -567.9255), (-25.0349, 28.7235, -567.7946), (-25.0349, 33.5413, -565.8672), (-25.0349, 34.6652, -565.3423), (-25.0349, 35.3057, -565.0194), (-25.0349, 39.8719, -562.3391), (-25.0349, 41.1612, -561.4398), (-25.0349, 41.8701, -560.8879), (-25.0349, 46.2988, -556.7645), (-25.0349, 47.4968, -555.4414), (-25.0349, 48.1929, -554.5801), (-25.0349, 53.2223, -547.4586), (-25.0349, 54.0362, -546.0693), (-25.0349, 54.5209, -545.1481), (-25.0349, 61.9431, -531.0663), (-25.0349, 62.4433, -530.1039), (-25.0349, 62.936, -529.136), (-25.0349, 69.347, -516.3262), (-25.0349, 69.8091, -515.3626), (-25.0349, 78.3075, -497.0272), (-25.0349, 78.7521, -496.0315), (-25.0349, 79.2148, -494.9575), (-25.0349, 87.9976, -473.1369), (-25.0349, 88.397, -472.0722), (-25.0349, 88.8128, -470.9238), (-25.0349, 83.2873, -470.6162), (-25.0349, 31.3464, -470.6162), (-25.0349, 31.396, -472.2874), (-25.0349, 31.401, -472.5102), (-25.0349, 31.4087, -472.715), (-25.0349, 31.42, -472.9023), (-25.0349, 30.7454, -481.2914), (-25.0349, 30.762, -481.4823), (-25.0349, 30.785, -481.6557), (-25.0349, 30.8664, -485.3562), (-25.0349, 30.8923, -485.5212), (-25.0349, 30.9324, -485.6697), (-25.0349, 31.156, -489.049), (-25.0349, 31.2033, -489.1888), (-25.0349, 31.897, -492.2408), (-25.0349, 31.942, -492.3715), (-25.0349, 32.0006, -492.4875), (-25.0349, 33.273, -495.3264), (-25.0349, 33.2083, -495.2194), (-25.0349, 34.62, -497.7339), (-25.0349, 34.71, -497.8317), (-25.0349, 34.815, -497.917), (-25.0349, 36.41, -500.0047), (-25.0349, 36.52, -500.0813), (-25.0349, 36.6234, -500.1376), (-25.0349, 38.647, -501.9046), (-25.0349, 38.783, -501.9663), (-25.0349, 41.2673, -503.4112), (-25.0349, 41.4243, -503.4637), (-25.0349, 41.601, -503.5061), (-25.0349, 44.606, -504.6305), (-25.0349, 44.801, -504.6645), (-25.0349, 48.33, -505.4674), (-25.0349, 48.544, -505.4932), (-25.0349, 48.783, -505.5109), (-25.0349, 53.064, -505.9928), (-25.0349, 53.308, -506.0031), (-25.0349, 1.4, -506.1637), (-25.0349, -10.0, -485.21237268486243), (-25.0349, -4.5, -506.1671), (-25.0349, -10.0, -515.2658399376711), (-25.0349, -4.0, -546.1672), (-25.0349, -0.0, -546.3361), (-25.0349, 3.8613, -546.2762), (-25.0349, 4.1645, -546.2757), (-25.0349, 4.4519, -546.2733), (-25.0349, 8.2111, -546.2592), (-25.0349, 8.5096, -546.2519), (-25.0349, 14.0323, -546.1611)
]
_loft8_d = [
    (-76.822, -0.0, -661.1299), (-76.822, 2.7, -661.0134), (-76.822, 3.4607, -680.9206), (-76.822, 4.4723, -680.6635), (-76.822, 5.7404, -680.2324), (-76.822, 6.7983, -679.8662), (-76.822, 8.3071, -679.2592), (-76.822, 9.6433, -678.6626), (-76.822, 13.3422, -676.9025), (-76.822, 14.8728, -676.1019), (-76.822, 19.4234, -673.5681), (-76.822, 20.9803, -672.6501), (-76.822, 25.2486, -669.9822), (-76.822, 26.8322, -668.9479), (-76.822, 31.577, -665.6837), (-76.822, 33.1141, -664.5788), (-76.822, 38.297, -660.6647), (-76.822, 39.7173, -659.5431), (-76.822, 45.2945, -654.9168), (-76.822, 46.5336, -653.8409), (-76.822, 51.3874, -649.4442), (-76.822, 52.4209, -648.4539), (-76.822, 57.4432, -643.5101), (-76.822, 58.2384, -642.6646), (-76.822, 63.38, -637.1425), (-76.822, 68.8054, -630.7495), (-76.822, 68.9744, -630.5528), (-76.822, 69.1094, -630.3917), (-76.822, 69.1653, -630.3053), (-76.822, 73.04, -624.771), (-76.822, 73.5268, -623.9433), (-76.822, 75.2574, -620.4846), (-76.822, 75.3278, -620.3052), (-76.822, 75.9424, -618.9279), (-76.822, 76.0015, -618.7891), (-76.822, 77.6301, -613.2487), (-76.822, 77.7931, -612.6712), (-76.822, 78.6933, -607.1358), (-76.822, 78.7709, -606.6603), (-76.822, 78.8352, -606.2324), (-76.822, 78.9341, -604.7402), (-76.822, 79.1449, -601.4837), (-76.822, 79.1357, -599.9078), (-76.822, 79.107, -596.6218), (-76.822, 78.9992, -595.1649), (-76.822, 78.7208, -591.6794), (-76.822, 78.5586, -590.5305), (-76.822, 77.9814, -586.6991), (-76.822, 77.9642, -586.6029), (-76.822, 77.2516, -582.72), (-76.822, 76.8125, -580.4075), (-76.822, 75.5102, -575.6394), (-76.822, 32.0728, -575.6394), (-76.822, -10.0, -640.9462), (-76.822, -80.0, -681.1299)
]
_loft8_e = [
    (-128.608, 18.0864, -575.6394), (-128.608, 19.1621, -579.5781), (-128.608, 19.0441, -583.8718), (-128.608, 18.22, -588.0969), (-128.608, 17.9553, -592.81), (-128.608, 17.0987, -595.8721), (-128.608, 16.7213, -600.1094), (-128.608, 15.8197, -602.9075), (-128.608, 15.3382, -606.6795), (-128.608, 14.4044, -609.2068), (-128.608, 13.8265, -612.5232), (-128.608, 12.8736, -614.7743), (-128.608, 12.2069, -617.6438), (-128.608, 11.248, -619.6143), (-128.608, 10.4994, -622.0449), (-128.608, 9.5475, -623.7315), (-128.608, 8.7239, -625.7307), (-128.608, 7.7921, -627.131), (-128.608, 6.8207, -628.8203), (-128.608, 5.8471, -630.014), (-128.608, 5.3336, -630.6737), (-128.608, 4.8665, -631.1429), (-128.608, 4.3356, -631.6805), (-128.608, 3.8844, -632.0645), (-128.608, 3.3372, -632.481), (-128.608, 2.9038, -632.7797), (-128.608, 2.3411, -633.0761), (-128.608, 1.9276, -633.2894), (-128.608, 1.3504, -633.4666), (-128.608, 0.9587, -633.5945), (-128.608, 0.1478, -633.6813), (-128.608, 0.0, -633.696), (-128.608, -3.8579, -633.6961), (-128.608, -9.0472, -633.6963), (-128.608, -12.9334, -633.6964), (-128.608, -15.9527, -633.6965), (-128.608, -18.3659, -633.6966), (-128.608, -20.339, -633.6966), (-128.608, -21.9822, -633.6966), (-128.608, -23.372, -633.6966), (-128.608, -39.4702, -633.6966), (-128.608, -38.7691, -673.8804), (-128.608, -37.9098, -673.8804), (-128.608, -37.0139, -673.8804), (-128.608, -34.1482, -673.8804), (-128.608, -30.8731, -673.8804), (-128.608, -27.0941, -673.8804), (-128.608, -22.6853, -673.8804), (-128.608, -17.4748, -673.8804), (-128.608, -11.2223, -673.8804), (-128.608, -3.5804, -673.8804), (-128.608, 0.0, -673.8804), (-128.608, 0.0965, -673.8719), (-128.608, 0.2642, -673.857), (-128.608, 0.3727, -673.8473), (-128.608, 0.4487, -673.8406), (-128.608, 0.5048, -673.8356), (-128.608, 0.8953, -673.7974), (-128.608, 1.4988, -673.6657), (-128.608, 2.0155, -673.5482), (-128.608, 3.5217, -673.0208), (-128.608, 4.8798, -672.5496), (-128.608, 6.7404, -671.6707), (-128.608, 8.4925, -670.8812), (-128.608, 10.6532, -669.6529), (-128.608, 12.7545, -668.5393), (-128.608, 15.0412, -667.0478), (-128.608, 17.3273, -665.6789), (-128.608, 19.7889, -663.8783), (-128.608, 22.3151, -662.1997), (-128.608, 24.9043, -660.096), (-128.608, 27.6333, -658.0977), (-128.608, 30.3035, -655.6983), (-128.608, 33.1973, -653.3682), (-128.608, 35.9027, -650.6819), (-128.608, 38.9221, -648.0065), (-128.608, 41.6181, -645.0435), (-128.608, 44.722, -642.0076), (-128.608, 47.3655, -638.7797), (-128.608, 50.5106, -635.3664), (-128.608, 53.6967, -631.3018), (-128.608, 54.9178, -629.8445), (-128.608, 56.544, -627.3316), (-128.608, 58.2251, -624.6892), (-128.608, 58.9754, -623.511), (-128.608, 61.6335, -617.5913), (-128.608, 62.0627, -616.6213), (-128.608, 63.6467, -611.1128), (-128.608, 63.8449, -610.4122), (-128.608, 64.8159, -602.9589), (-128.608, 64.9137, -602.1896), (-128.608, 64.6311, -594.6473), (-128.608, 64.6018, -593.9033), (-128.608, 64.3976, -592.8329), (-128.608, 63.2656, -587.0437), (-128.608, 60.1507, -575.6394), (-128.608, 32.0728, -575.6394), (-128.608, -80.0, -673.8804), (-128.608, -80.0, -633.6966)
]
_loft9_a = [
    (-142.4469, 42.4942, -575.6394), (-142.4469, 45.1912, -585.5135), (-142.4469, 45.8239, -587.8302), (-142.4469, 45.895, -588.0903), (-142.4469, 45.9095, -588.1647), (-142.4469, 45.9168, -588.2007), (-142.4469, 45.9385, -588.3087), (-142.4469, 47.206, -594.6136), (-142.4469, 47.5971, -602.0711), (-142.4469, 47.6114, -602.3707), (-142.4469, 46.8302, -609.8524), (-142.4469, 46.803, -610.0948), (-142.4469, 45.3678, -615.8416), (-142.4469, 45.3361, -615.963), (-142.4469, 42.7374, -622.5124), (-142.4469, 42.7295, -622.5315), (-142.4469, 42.2876, -623.3036), (-142.4469, 39.3428, -628.4576), (-142.4469, 39.2133, -628.6747), (-142.4469, 37.7149, -630.8606), (-142.4469, 37.22, -631.548), (-142.4469, 33.3533, -636.7972), (-142.4469, 29.9488, -640.9733), (-142.4469, 28.7581, -642.405), (-142.4469, 25.5717, -645.9478), (-142.4469, 24.0358, -647.5494), (-142.4469, 21.1068, -650.4986), (-142.4469, 18.531, -652.8404), (-142.4469, 16.6775, -654.5091), (-142.4469, 14.3642, -656.3986), (-142.4469, 12.4311, -657.8743), (-142.4469, 10.3878, -659.3538), (-142.4469, 8.5448, -660.5046), (-142.4469, 6.6727, -661.6672), (-142.4469, 5.1185, -662.4602), (-142.4469, 3.4831, -663.1976), (-142.4469, 2.2196, -663.6612), (-142.4469, 1.5365, -663.8503), (-142.4469, 1.0158, -663.9648), (-142.4469, 0.4447, -664.0283), (-142.4469, 0.0, -664.0663), (-142.4469, 0.0, -623.6088), (-142.4469, 0.1467, -623.5992), (-142.4469, 0.8525, -623.5306), (-142.4469, 0.9891, -623.5019), (-142.4469, 1.6288, -623.2963), (-142.4469, 1.7558, -623.2483), (-142.4469, 1.8728, -623.1809), (-142.4469, 2.4474, -622.8393), (-142.4469, 2.5553, -622.7518), (-142.4469, 3.0645, -622.276), (-142.4469, 3.1636, -622.1677), (-142.4469, 3.6082, -621.5603), (-142.4469, 3.779, -621.2796), (-142.4469, 4.1611, -620.5413), (-142.4469, 4.2949, -620.1717), (-142.4469, 4.7716, -618.5765), (-142.4469, 4.874, -618.3183), (-142.4469, 4.9718, -617.8459), (-142.4469, 5.2127, -616.3287), (-142.4469, 5.2953, -615.719), (-142.4469, 5.3727, -614.4455), (-142.4469, 5.5058, -612.0281), (-142.4469, 5.5074, -611.9211), (-142.4469, 5.4661, -609.0813), (-142.4469, 5.3675, -607.2554), (-142.4469, 5.163, -604.0754), (-142.4469, 4.8667, -601.6563), (-142.4469, 4.5095, -598.1616), (-142.4469, 4.0288, -594.4713), (-142.4469, 3.4937, -591.264), (-142.4469, 2.8913, -587.3323), (-142.4469, 2.1003, -583.2918), (-142.4469, 1.3908, -579.1575), (-142.4469, 0.4299, -575.6394), (-142.4469, -40.0, -589.8478), (-142.4469, -40.0, -595.7161), (-142.4469, -40.0, -596.3359), (-142.4469, -40.0, -596.8286), (-142.4469, -40.0, -604.1451), (-142.4469, -40.0, -605.436), (-142.4469, -40.0, -611.2627), (-142.4469, -40.0, -613.4722), (-142.4469, -40.0, -617.7906), (-142.4469, -40.0, -620.3557), (-142.4469, -40.0, -623.5829), (-142.4469, -11.926, -623.6151), (-142.4469, -41.6904, -664.0663), (-142.4469, -80.0, -664.0663), (-142.4469, -80.0, -654.5044), (-142.4469, -80.0, -651.3711), (-142.4469, -80.0, -641.0278), (-142.4469, -80.0, -634.8531), (-142.4469, -80.0, -628.6882), (-142.4469, -80.0, -624.8433), (-142.4469, -80.0, -617.3897), (-142.4469, -80.0, -615.1306), (-142.4469, -80.0, -613.2607), (-142.4469, -80.0, -605.6374), (-142.4469, -80.0, -604.8119), (-142.4469, -80.0, -604.1583), (-142.4469, -80.0, -596.4455), (-142.4469, -80.0, -596.2049), (-142.4469, -80.0, -589.8478), (-142.4469, -71.8873, -589.8478)
]
_loft9_b = [
    (-242.283, 60.1507, -575.6394), (-242.283, 60.1507, -575.6394), (-242.283, 63.2656, -587.0437), (-242.283, 64.3976, -592.8329), (-242.283, 64.6018, -593.9033), (-242.283, 64.6311, -594.6473), (-242.283, 64.9137, -602.1896), (-242.283, 64.8159, -602.9589), (-242.283, 63.8449, -610.4122), (-242.283, 63.6467, -611.1128), (-242.283, 62.0627, -616.6213), (-242.283, 61.6335, -617.5913), (-242.283, 58.9754, -623.511), (-242.283, 58.2251, -624.6892), (-242.283, 56.544, -627.3316), (-242.283, 54.9178, -629.8445), (-242.283, 53.6967, -631.3018), (-242.283, 50.5106, -635.3664), (-242.283, 47.3655, -638.7797), (-242.283, 44.722, -642.0076), (-242.283, 41.6181, -645.0435), (-242.283, 38.9221, -648.0065), (-242.283, 35.9027, -650.6819), (-242.283, 33.1973, -653.3682), (-242.283, 30.3035, -655.6983), (-242.283, 27.6333, -658.0977), (-242.283, 24.9043, -660.096), (-242.283, 22.3151, -662.1997), (-242.283, 19.7889, -663.8783), (-242.283, 17.3273, -665.6789), (-242.283, 15.0412, -667.0478), (-242.283, 12.7545, -668.5393), (-242.283, 10.6532, -669.6529), (-242.283, 8.4925, -670.8812), (-242.283, 6.7404, -671.6707), (-242.283, 4.8798, -672.5496), (-242.283, 3.5217, -673.0208), (-242.283, 2.0155, -673.5482), (-242.283, 1.4988, -673.6657), (-242.283, 0.8953, -673.7974), (-242.283, 0.5048, -673.8356), (-242.283, 0.4487, -673.8406), (-242.283, 0.3727, -673.8473), (-242.283, 0.2642, -673.857), (-242.283, 0.0965, -673.8719), (-242.283, 0.0, -673.8804), (-242.283, -3.5804, -673.8804), (-242.283, -11.2223, -673.8804), (-242.283, -17.4748, -673.8804), (-242.283, -22.6853, -673.8804), (-242.283, -27.0941, -673.8804), (-242.283, -30.8731, -673.8804), (-242.283, -34.1482, -673.8804), (-242.283, -37.0139, -673.8804), (-242.283, -37.9098, -673.8804), (-242.283, -38.7691, -673.8804), (-242.283, -39.4702, -633.6966), (-242.283, -23.372, -633.6966), (-242.283, -21.9822, -633.6966), (-242.283, -20.339, -633.6966), (-242.283, -18.3659, -633.6966), (-242.283, -15.9527, -633.6965), (-242.283, -12.9334, -633.6964), (-242.283, -9.0472, -633.6963), (-242.283, -3.8579, -633.6961), (-242.283, 0.0, -633.696), (-242.283, 0.1478, -633.6813), (-242.283, 0.9587, -633.5945), (-242.283, 1.3504, -633.4666), (-242.283, 1.9276, -633.2894), (-242.283, 2.3411, -633.0761), (-242.283, 2.9038, -632.7797), (-242.283, 3.3372, -632.481), (-242.283, 3.8844, -632.0645), (-242.283, 4.3356, -631.6805), (-242.283, 4.8665, -631.1429), (-242.283, 5.3336, -630.6737), (-242.283, 5.8471, -630.014), (-242.283, 6.8207, -628.8203), (-242.283, 7.7921, -627.131), (-242.283, 8.7239, -625.7307), (-242.283, 9.5475, -623.7315), (-242.283, 10.4994, -622.0449), (-242.283, 11.248, -619.6143), (-242.283, 12.2069, -617.6438), (-242.283, 12.8736, -614.7743), (-242.283, 13.8265, -612.5232), (-242.283, 14.4044, -609.2068), (-242.283, 15.3382, -606.6795), (-242.283, 15.8197, -602.9075), (-242.283, 16.7213, -600.1094), (-242.283, 17.0987, -595.8721), (-242.283, 17.9553, -592.81), (-242.283, 18.22, -588.0969), (-242.283, 19.0441, -583.8718), (-242.283, 19.1621, -579.5781), (-242.283, 18.0864, -575.6394), (-242.283, -70.9938, -633.6966), (-242.283, -80.0, -633.6966), (-242.283, -80.0, -640.802), (-242.283, -80.0, -673.8804)
]
_loft9_c = [
    (-317.8223, 73.5102, -575.6394), (-317.8223, 74.8125, -580.4075), (-317.8223, 75.2516, -582.72), (-317.8223, 75.9642, -586.6029), (-317.8223, 75.9814, -586.6991), (-317.8223, 76.5586, -590.5305), (-317.8223, 76.7208, -591.6794), (-317.8223, 76.9992, -595.1649), (-317.8223, 77.107, -596.6218), (-317.8223, 77.1357, -599.9078), (-317.8223, 77.1449, -601.4837), (-317.8223, 76.9341, -604.7402), (-317.8223, 76.8352, -606.2324), (-317.8223, 76.7709, -606.6603), (-317.8223, 76.6933, -607.1358), (-317.8223, 75.7931, -612.6712), (-317.8223, 75.6301, -613.2487), (-317.8223, 74.0015, -618.7891), (-317.8223, 73.9424, -618.9279), (-317.8223, 73.3278, -620.3052), (-317.8223, 73.2574, -620.4846), (-317.8223, 71.5268, -623.9433), (-317.8223, 71.04, -624.771), (-317.8223, 67.1653, -630.3053), (-317.8223, 67.1094, -630.3917), (-317.8223, 66.9744, -630.5528), (-317.8223, 66.8054, -630.7495), (-317.8223, 61.38, -637.1425), (-317.8223, 56.2384, -642.6646), (-317.8223, 55.4432, -643.5101), (-317.8223, 50.4209, -648.4539), (-317.8223, 49.3874, -649.4442), (-317.8223, 44.5336, -653.8409), (-317.8223, 43.2945, -654.9168), (-317.8223, 37.7173, -659.5431), (-317.8223, 36.297, -660.6647), (-317.8223, 31.1141, -664.5788), (-317.8223, 29.577, -665.6837), (-317.8223, 24.8322, -668.9479), (-317.8223, 23.2486, -669.9822), (-317.8223, 18.9803, -672.6501), (-317.8223, 17.4234, -673.5681), (-317.8223, 12.8728, -676.1019), (-317.8223, 11.3422, -676.9025), (-317.8223, 7.6433, -678.6626), (-317.8223, 6.3071, -679.2592), (-317.8223, 4.7983, -679.8662), (-317.8223, 3.7404, -680.2324), (-317.8223, 2.4723, -680.6635), (-317.8223, 1.4607, -680.9206), (-317.8223, 1.0552, -681.0134), (-317.8223, 0.3119, -681.0989), (-317.8223, 0.0, -681.1299), (-317.8223, -2.7014, -681.1299), (-317.8223, -4.3124, -681.13), (-317.8223, -5.8575, -681.1299), (-317.8223, -1.7507, -640.9461), (-317.8223, 0.0, -640.9462), (-317.8223, 0.4203, -640.8972), (-317.8223, 0.696, -640.865), (-317.8223, 0.7879, -640.8543), (-317.8223, 0.9752, -640.8327), (-317.8223, 1.779, -640.5741), (-317.8223, 2.0296, -640.4922), (-317.8223, 2.2944, -640.3557), (-317.8223, 4.0751, -639.3223), (-317.8223, 4.3535, -639.1305), (-317.8223, 4.6439, -638.8834), (-317.8223, 6.624, -637.1642), (-317.8223, 6.9262, -636.8605), (-317.8223, 9.0664, -634.4617), (-317.8223, 9.7021, -633.6819), (-317.8223, 11.8716, -630.7469), (-317.8223, 12.488, -629.8204), (-317.8223, 14.7367, -626.2729), (-317.8223, 15.3762, -625.1394), (-317.8223, 17.6705, -620.9887), (-317.8223, 18.3267, -619.6396), (-317.8223, 20.6332, -614.8967), (-317.8223, 21.2993, -613.3219), (-317.8223, 23.5851, -607.9994), (-317.8223, 24.2539, -606.1874), (-317.8223, 26.4867, -600.2995), (-317.8223, 27.1504, -598.2372), (-317.8223, 29.2986, -591.8001), (-317.8223, 29.9486, -589.473), (-317.8223, 31.9816, -582.5043), (-317.8223, 32.6086, -579.8963), (-317.8223, 31.4458, -575.6394), (-317.8223, -38.7814, -640.9462), (-317.8223, -59.283, -640.9462), (-317.8223, -80.0, -640.9461), (-317.8223, -80.0, -657.2905), (-317.8223, -80.0, -681.1299)
]
_loft9_d = [
    (-349.8096, 26.7057, -575.6394), (-349.8096, 1.7233, -583.1492), (-349.8096, 0.383, -575.6394), (-349.8096, 5.444, -605.1079), (-349.8096, 5.9455, -606.8755), (-349.8096, 6.132, -607.9872), (-349.8096, 6.6559, -609.9365), (-349.8096, 6.7619, -610.7436), (-349.8096, 7.2544, -612.6954), (-349.8096, 7.3326, -613.3787), (-349.8096, 7.7881, -615.3237), (-349.8096, 7.843, -615.8943), (-349.8096, 8.2562, -617.8237), (-349.8096, 8.292, -618.2923), (-349.8096, 8.6579, -620.1975), (-349.8096, 8.6785, -620.5742), (-349.8096, 8.9922, -622.4471), (-349.8096, 9.0013, -622.7417), (-349.8096, 9.2584, -624.5747), (-349.8096, 9.2595, -624.7963), (-349.8096, 9.4937, -627.0159), (-349.8096, 9.4887, -627.1961), (-349.8096, 9.6449, -630.3461), (-349.8096, 9.6375, -630.4416), (-349.8096, 9.5752, -632.4643), (-349.8096, 9.5531, -633.3234), (-349.8096, 9.5533, -633.3494), (-349.8096, 9.2302, -635.8489), (-349.8096, 9.2174, -635.9375), (-349.8096, 8.6646, -638.0251), (-349.8096, 8.6458, -638.0885), (-349.8096, 8.2896, -638.984), (-349.8096, 8.261, -639.0519), (-349.8096, 7.8527, -639.8578), (-349.8096, 7.8128, -639.9274), (-349.8096, 7.3537, -640.6469), (-349.8096, 7.3012, -640.7158), (-349.8096, 6.7924, -641.352), (-349.8096, 6.7263, -641.4181), (-349.8096, 6.1687, -641.9736), (-349.8096, 6.0883, -642.035), (-349.8096, 5.4826, -642.5122), (-349.8096, 5.2811, -642.6127), (-349.8096, 3.9239, -643.3414), (-349.8096, 3.6577, -643.4076), (-349.8096, 2.0874, -643.8475), (-349.8096, 1.9311, -643.8647), (-349.8096, 1.0748, -643.9738), (-349.8096, 0.905, -643.9796), (-349.8096, 0.0, -644.016), (-349.8096, -0.22, -644.016), (-349.8096, -1.6028, -644.016), (-349.8096, -2.8867, -644.016), (-349.8096, -4.0819, -644.016), (-349.8096, -5.1974, -644.016), (-349.8096, -6.2408, -644.016), (-349.8096, -7.2189, -644.016), (-349.8096, -38.4896, -644.016), (-349.8096, -54.324, -644.016), (-349.8096, -80.0, -644.016), (-349.8096, -80.0, -664.2726), (-349.8096, -80.0, -684.1997), (-349.8096, -8.1746, -684.1998), (-349.8096, -6.6596, -684.1997), (-349.8096, -5.0794, -684.1998), (-349.8096, -3.4295, -684.1998), (-349.8096, -1.7053, -684.1997), (-349.8096, 0.0, -684.1997), (-349.8096, 0.961, -684.1291), (-349.8096, 2.2918, -684.0184), (-349.8096, 3.3403, -683.8086), (-349.8096, 4.7472, -683.4705), (-349.8096, 5.8607, -683.1284), (-349.8096, 7.3677, -682.5447), (-349.8096, 8.5189, -682.082), (-349.8096, 9.6895, -681.5112), (-349.8096, 11.3083, -680.6558), (-349.8096, 12.4618, -680.0001), (-349.8096, 14.5249, -678.6016), (-349.8096, 15.6181, -677.8931), (-349.8096, 16.6681, -677.1326), (-349.8096, 18.9051, -675.2472), (-349.8096, 19.8314, -674.5012), (-349.8096, 20.6907, -673.7405), (-349.8096, 23.0702, -671.3148), (-349.8096, 24.3641, -669.9975), (-349.8096, 26.8257, -667.0035), (-349.8096, 27.379, -666.3213), (-349.8096, 30.7814, -660.7771), (-349.8096, 30.8663, -660.6475), (-349.8096, 31.438, -659.1604), (-349.8096, 34.1532, -653.631), (-349.8096, 34.2216, -653.4894), (-349.8096, 34.3162, -653.2177), (-349.8096, 36.1602, -647.4502), (-349.8096, 36.2334, -647.1164), (-349.8096, 37.2763, -641.4146), (-349.8096, 37.3319, -640.9619), (-349.8096, 37.6218, -635.3377), (-349.8096, 37.6357, -634.8619), (-349.8096, 37.6046, -634.3763), (-349.8096, 37.1689, -628.7798), (-349.8096, 37.1125, -628.4087), (-349.8096, 35.9331, -622.8005), (-349.8096, 35.8679, -622.5218), (-349.8096, 35.3955, -620.8254), (-349.8096, 31.5412, -606.9854)
]

# Profiles: _loft8_a, _loft8_b, _loft8_c, _loft8_d, _loft8_e, _loft4_d, _loft9_a, _loft9_b, _loft9_c, _loft9_d
# We'll reverse the order to maintain positive normals (increasing X).
_loft10_a = [(-429.9859, 52.2748, -604.1361), (-429.9859, 56.2768, -618.5065), (-429.9859, 56.588, -619.6238), (-429.9859, 56.6479, -619.884), (-429.9859, 56.689, -620.0618), (-429.9859, 56.9383, -621.1419), (-429.9859, 57.5296, -623.7028), (-429.9859, 57.6727, -624.708), (-429.9859, 58.1112, -627.8361), (-429.9859, 58.1695, -629.0383), (-429.9859, 58.3397, -632.8065), (-429.9859, 58.2763, -634.0329), (-429.9859, 58.0655, -637.7759), (-429.9859, 57.871, -639.049), (-429.9859, 57.2921, -642.6939), (-429.9859, 56.9466, -644.0338), (-429.9859, 56.0234, -647.5105), (-429.9859, 55.1458, -649.6986), (-429.9859, 53.1497, -654.5504), (-429.9859, 52.3667, -655.969), (-429.9859, 50.9254, -658.515), (-429.9859, 49.9167, -660.0075), (-429.9859, 48.3486, -662.2692), (-429.9859, 47.0817, -663.8244), (-429.9859, 45.4432, -665.7831), (-429.9859, 42.4631, -668.7853), (-429.9859, 42.2289, -669.0285), (-429.9859, 37.9941, -672.5658), (-429.9859, 36.9803, -673.5342), (-429.9859, 33.9668, -675.7284), (-429.9859, 32.7607, -676.8674), (-429.9859, 29.8131, -678.8217), (-429.9859, 28.6106, -679.8567), (-429.9859, 25.7518, -681.5774), (-429.9859, 24.5599, -682.5058), (-429.9859, 21.8126, -683.9992), (-429.9859, 20.6384, -684.8186), (-429.9859, 18.025, -686.0903), (-429.9859, 16.8757, -686.7988), (-429.9859, 14.4188, -687.8537), (-429.9859, 13.3015, -688.4497), (-429.9859, 10.6634, -689.4126), (-429.9859, 9.4105, -689.9638), (-429.9859, 7.0523, -690.6446), (-429.9859, 5.8643, -691.0393), (-429.9859, 3.8226, -691.4441), (-429.9859, 2.7111, -691.6813), (-429.9859, 1.2582, -691.7971), (-429.9859, 0.8721, -691.8269), (-429.9859, 0.1728, -691.8809), (-429.9859, 0.0, -691.8943), (-429.9859, -10.0, -691.8943), (-429.9859, -10.0, -651.705), (-429.9859, 0.0, -651.7013), (-429.9859, 1.9395, -651.5957), (-429.9859, 2.0359, -651.5905), (-429.9859, 2.051, -651.5895), (-429.9859, 3.9586, -651.253), (-429.9859, 3.9637, -651.2521), (-429.9859, 3.9697, -651.2502), (-429.9859, 5.5413, -650.7468), (-429.9859, 5.7387, -650.6835), (-429.9859, 5.7401, -650.6831), (-429.9859, 7.3575, -649.886), (-429.9859, 7.373, -649.8778), (-429.9859, 8.123, -649.3929), (-429.9859, 8.134, -649.3851), (-429.9859, 8.8427, -648.8436), (-429.9859, 8.8571, -648.8315), (-429.9859, 9.524, -648.234), (-429.9859, 9.5414, -648.2164), (-429.9859, 10.1662, -647.5636), (-429.9859, 10.1863, -647.5394), (-429.9859, 11.3042, -646.0765), (-429.9859, 11.3546, -645.9972), (-429.9859, 12.3012, -644.3249), (-429.9859, 12.3553, -644.2007), (-429.9859, 13.2197, -642.067), (-429.9859, 13.2783, -641.8531), (-429.9859, 13.9243, -639.4782), (-429.9859, 13.9676, -639.1689), (-429.9859, 14.3988, -636.5707), (-429.9859, 14.4131, -636.1432), (-429.9859, 14.6351, -633.3418), (-429.9859, 14.6045, -632.7719), (-429.9859, 14.6249, -629.7894), (-429.9859, 14.5315, -629.0511), (-429.9859, 14.3598, -625.9117), (-429.9859, 14.1838, -624.9778), (-429.9859, 13.8316, -621.7071), (-429.9859, 13.5514, -620.5492), (-429.9859, 13.1071, -618.8076), (-429.9859, 12.6242, -615.7631), (-429.9859, 9.3862, -604.1361)]
_loft10_b = [(-530.4893, 71.3238, -575.6394), (-530.4893, 81.5598, -612.3949), (-530.4893, 82.5629, -615.9971), (-530.4893, 82.7876, -616.9929), (-530.4893, 83.7751, -621.7237), (-530.4893, 83.8883, -622.621), (-530.4893, 84.278, -627.5924), (-530.4893, 84.301, -628.6276), (-530.4893, 84.0732, -633.5994), (-530.4893, 83.9754, -634.7029), (-530.4893, 83.1894, -639.4576), (-530.4893, 82.9336, -640.7292), (-530.4893, 81.933, -644.2215), (-530.4893, 81.5731, -645.3458), (-530.4893, 80.2323, -648.7853), (-530.4893, 79.4069, -650.4739), (-530.4893, 77.8524, -653.5961), (-530.4893, 75.802, -656.9038), (-530.4893, 75.3148, -657.6339), (-530.4893, 74.2956, -658.8995), (-530.4893, 70.4095, -663.6672), (-530.4893, 70.2284, -663.8538), (-530.4893, 68.9942, -665.0732), (-530.4893, 65.6076, -668.0656), (-530.4893, 63.8008, -669.5611), (-530.4893, 62.7092, -670.3389), (-530.4893, 53.5638, -676.7105), (-530.4893, 52.8687, -677.148), (-530.4893, 44.0693, -682.5006), (-530.4893, 43.2782, -682.9584), (-530.4893, 34.9863, -687.3702), (-530.4893, 34.1302, -687.8256), (-530.4893, 33.2911, -688.2343), (-530.4893, 25.6096, -691.7924), (-530.4893, 24.7508, -692.1732), (-530.4893, 17.8974, -694.8919), (-530.4893, 17.0504, -695.2302), (-530.4893, 16.1426, -695.5468), (-530.4893, 10.2434, -697.4727), (-530.4893, 9.4219, -697.7081), (-530.4893, 6.891, -698.3777), (-530.4893, 4.6707, -698.8538), (-530.4893, 3.9733, -698.9985), (-530.4893, 2.0806, -699.2839), (-530.4893, 0.5406, -699.3789), (-530.4893, 0.0, -699.4036), (-530.4893, -10.0, -699.4036), (-530.4893, 2.5, -658.3658), (-530.4893, 3.1528, -658.2524), (-530.4893, 2.5, -657.1791), (-530.4893, 2.5, -656.6374), (-530.4893, 2.5, -653.7106), (-530.4893, 2.5, -650.283), (-530.4893, 2.5, -644.1843), (-530.4893, 2.5, -640.0166), (-530.4893, 2.5, -633.9035), (-530.4893, 2.5, -625.6278), (-530.4893, 2.5, -618.8003), (-530.4893, 2.5, -610.2989), (-530.4893, -2.2933, -610.3061), (-530.4893, -10.0, -610.2937), (-530.4893, 35.8408, -602.2317), (-530.4893, 39.7294, -616.1948), (-530.4893, 38.3068, -620.1578), (-530.4893, 37.7688, -621.314), (-530.4893, 37.5714, -621.7354), (-530.4893, 36.2848, -624.6813), (-530.4893, 35.2428, -626.8939), (-530.4893, 34.436, -628.5239), (-530.4893, 32.7677, -631.6724), (-530.4893, 32.4802, -632.1868), (-530.4893, 31.3582, -634.0742), (-530.4893, 30.4327, -635.6462), (-530.4893, 30.2515, -635.9394), (-530.4893, 30.1711, -636.0611), (-530.4893, 28.3044, -638.8788), (-530.4893, 27.4775, -640.048), (-530.4893, 26.095, -641.8997), (-530.4893, 25.3105, -642.8991), (-530.4893, 24.7105, -643.6423), (-530.4893, 23.8134, -644.704), (-530.4893, 21.8933, -646.8458), (-530.4893, 21.4687, -647.2868), (-530.4893, 19.1823, -649.5309), (-530.4893, 19.0709, -649.6398), (-530.4893, 19.0252, -649.6788), (-530.4893, 16.5834, -651.7485), (-530.4893, 14.7579, -653.1042), (-530.4893, 13.3118, -654.0421), (-530.4893, 12.4461, -654.608), (-530.4893, 11.6856, -655.0678), (-530.4893, 11.0037, -655.4443), (-530.4893, 10.381, -655.7537), (-530.4893, 7.9943, -656.8269), (-530.4893, 6.9225, -657.2726), (-530.4893, 5.9983, -657.5547), (-530.4893, 28.4352, -575.6394)]

# ── Helper for robust manifold lofts ───────────────────────────────────────
def _make_solid_loft(profiles, ruled=True, n=150):
    res = [_resample_profile(_dedupe_3d(p), n) for p in profiles]
    wires = [Wire.make_polygon([Vector(*pt) for pt in r_pts], close=True) for r_pts in res]
    
    try:
        # Try direct solid generation first
        lateral = Solid.make_loft(wires, ruled=ruled)
        if lateral.is_manifold:
            return lateral
    except:
        pass

    # Fallback to manual sewing with high tolerance
    try:
        lateral = Solid.make_loft(wires, ruled=ruled)
        cap_s = _make_planar_cap(res[0])
        cap_e = _make_planar_cap(res[-1])
        sew = BRepBuilderAPI_Sewing(0.5) # Increased tolerance for complex joins
        sew.Add(lateral.wrapped)
        sew.Add(cap_s.wrapped)
        sew.Add(cap_e.wrapped)
        sew.Perform()
        return Solid(sew.SewedShape())
    except:
        # Last resort: just return the lateral shell
        return Solid.make_loft(wires, ruled=ruled)



# ── Loft 6, 7, 8 construction (Extended to meet each other) ─────────────────
# Loft 6: Original segment
loft_6 = _make_solid_loft([_loft8_a, _loft8_b, _loft8_c, _loft8_d, _loft8_e])

# Loft 7: Extended to meet Loft 6 (at _loft8_e) via bridge _loft4_d
loft_7 = _make_solid_loft([_loft8_e, _loft4_d, _loft9_a, _loft9_b, _loft9_c, _loft9_d])

# Loft 8: Extended to meet Loft 7 (at _loft9_d)
loft_8 = _make_solid_loft([_loft9_d, _loft10_a, _loft10_b])

print(f"Loft 6 Is Manifold: {loft_6.is_manifold}")
print(f"Loft 7 (Extended) Is Manifold: {loft_7.is_manifold}")
print(f"Loft 8 (Extended) Is Manifold: {loft_8.is_manifold}")

# ── Final Enclosure Joining ────────────────────────────────────────────────
# The three lofts only share boundary wires (1D curves), not faces, so OCCT's
# boolean fuse silently discards the operands it can't connect — here it
# returned only loft_8 even though is_manifold reported True. Validate that
# the fuse actually contains all three operands (by volume); otherwise fall
# back to a Compound so loft_6 and loft_7 remain present in the export.
enclosure_final = None
_expected_vol = loft_6.volume + loft_7.volume + loft_8.volume
try:
    fused = loft_6.fuse(loft_7).fuse(loft_8)
    if fused.is_manifold and fused.volume >= 0.95 * _expected_vol:
        enclosure_final = fused
        print(f"Joined all extended segments into final enclosure. Manifold: {enclosure_final.is_manifold}")
    else:
        print(f"Warning: Fusion dropped operands (got {fused.volume:.1f} mm³, expected ~{_expected_vol:.1f} mm³). Falling back to Compound.")
        enclosure_final = Compound([loft_6, loft_7, loft_8])
except:
    print("Warning: Fusion failed, falling back to Compound.")
    enclosure_final = Compound([loft_6, loft_7, loft_8])


# ── Multi-profile Loft with smooth curvy structure ──────────────────────────
# Creates a loft between 6 profile sections with fine curves
# X=-897.897 → X=-553.2745, following base periphery edge

_loft_profile_1 = [
    (-897.897, 277.7199, -575.6394),
    (-897.897, 277.7024, -578.3821),
    (-897.897, 277.673, -583.0216),
    (-897.897, 277.6393, -588.309),
    (-897.897, 277.6139, -592.3206),
    (-897.897, 277.5877, -596.4264),
    (-897.897, 277.5835, -597.0833),
    (-897.897, 276.5348, -600.5154),
    (-897.897, 276.4476, -600.767),
    (-897.897, 273.3952, -607.3398),
    (-897.897, 272.373, -609.3233),
    (-897.897, 270.8034, -611.7739),
    (-897.897, 270.1882, -612.6499),
    (-897.897, 268.4888, -614.911),
    (-897.897, 267.8851, -615.6215),
    (-897.897, 264.6366, -619.127),
    (-897.897, 263.859, -619.602),
    (-897.897, 260.4233, -621.894),
    (-897.897, 258.4003, -623.0952),
    (-897.897, 257.3425, -623.8017),
    (-897.897, 254.9847, -625.3752),
    (-897.897, 249.1912, -629.3438),
    (-897.897, 247.5536, -630.4229),
    (-897.897, 246.5615, -631.1281),
    (-897.897, 240.9925, -633.2051),
    (-897.897, 239.4648, -633.7114),
    (-897.897, 237.5288, -634.2353),
    (-897.897, 220.3214, -638.8031),
    (-897.897, 212.7312, -640.7243),
    (-897.897, 203.0293, -643.0933),
    (-897.897, 191.496, -645.7963),
    (-897.897, 187.1542, -646.7573),
    (-897.897, 172.6848, -649.865),
    (-897.897, 170.9907, -650.1954),
    (-897.897, 159.9397, -652.3356),
    (-897.897, 156.0632, -653.0938),
    (-897.897, 147.4932, -654.5728),
    (-897.897, 139.528, -655.9596),
    (-897.897, 134.7869, -656.6762),
    (-897.897, 122.9017, -658.4867),
    (-897.897, 117.1203, -659.24),
    (-897.897, 106.5442, -660.6155),
    (-897.897, 94.7296, -661.9436),
    (-897.897, 82.8819, -663.0971),
    (-897.897, 70.8617, -664.0913),
    (-897.897, 58.6484, -664.9251),
    (-897.897, 49.5166, -665.4407),
    (-897.897, 44.3278, -665.8863),
    (-897.897, 40.0, -666.6625),
    (-897.897, 8.6985, -666.6625),
    (-897.897, 24.0, -630.3553),
    (-897.897, 28.8898, -626.3345),
    (-897.897, 33.6555, -626.2255),
    (-897.897, 40.2182, -626.0167),
    (-897.897, 49.4527, -625.6637),
    (-897.897, 58.4367, -625.2335),
    (-897.897, 72.1322, -624.374),
    (-897.897, 87.4817, -623.1526),
    (-897.897, 102.1592, -621.6713),
    (-897.897, 116.2459, -619.9314),
    (-897.897, 129.8216, -617.9343),
    (-897.897, 142.9647, -615.6814),
    (-897.897, 155.7532, -613.1737),
    (-897.897, 168.2633, -610.4126),
    (-897.897, 181.1177, -607.2545),
    (-897.897, 193.7081, -603.8658),
    (-897.897, 202.6325, -601.2947),
    (-897.897, 205.57, -600.4053),
    (-897.897, 208.768, -599.4413),
    (-897.897, 211.7924, -598.4938),
    (-897.897, 214.9169, -597.5231),
    (-897.897, 218.0365, -596.5145),
    (-897.897, 221.0883, -595.5399),
    (-897.897, 224.3121, -594.4675),
    (-897.897, 227.2814, -593.527),
    (-897.897, 230.6303, -592.4634),
    (-897.897, 236.9967, -590.1697),
]

_loft_profile_2 = [
    (-827.718, 245.528, -575.6394),
    (-827.718, 246.2982, -579.1175),
    (-827.718, 246.5501, -580.2489),
    (-827.718, 246.6427, -580.6633),
    (-827.718, 246.6844, -581.1621),
    (-827.718, 247.1497, -586.8998),
    (-827.718, 247.0962, -587.5933),
    (-827.718, 246.4532, -595.4589),
    (-827.718, 246.3604, -595.8815),
    (-827.718, 245.5027, -599.6717),
    (-827.718, 245.3461, -600.1644),
    (-827.718, 244.1615, -603.798),
    (-827.718, 243.6036, -604.9757),
    (-827.718, 240.3629, -611.667),
    (-827.718, 239.8581, -612.4456),
    (-827.718, 237.9334, -615.3484),
    (-827.718, 237.2453, -616.2286),
    (-827.718, 235.1706, -618.8221),
    (-827.718, 233.4048, -620.5962),
    (-827.718, 229.4549, -624.4717),
    (-827.718, 227.7086, -625.8377),
    (-827.718, 225.0469, -627.8632),
    (-827.718, 222.8136, -629.28),
    (-827.718, 220.3, -630.8251),
    (-827.718, 217.5301, -632.2237),
    (-827.718, 215.2648, -633.3249),
    (-827.718, 213.3988, -634.0074),
    (-827.718, 210.0096, -635.3355),
    (-827.718, 201.1737, -638.1199),
    (-827.718, 190.2282, -641.3998),
    (-827.718, 186.5655, -642.5271),
    (-827.718, 178.0098, -644.87),
    (-827.718, 172.5468, -646.4139),
    (-827.718, 165.7049, -648.1146),
    (-827.718, 158.5125, -649.9567),
    (-827.718, 153.2514, -651.127),
    (-827.718, 144.3791, -653.1565),
    (-827.718, 140.5827, -653.9002),
    (-827.718, 130.0627, -656.0142),
    (-827.718, 127.6276, -656.4264),
    (-827.718, 115.4795, -658.5307),
    (-827.718, 100.5899, -660.7004),
    (-827.718, 90.2527, -661.9748),
    (-827.718, 79.6777, -663.0969),
    (-827.718, 68.7927, -664.0701),
    (-827.718, 57.5664, -664.8919),
    (-827.718, 50.2841, -665.3428),
    (-827.718, 48.0737, -665.4877),
    (-827.718, 44.2854, -665.8614),
    (-827.718, 40.0, -666.6625),
    (-827.718, 30.0, -646.6625),
    (-827.718, 18.0439, -646.6625),
    (-827.718, 4.3698, -646.6625),
    (-827.718, -10.0, -646.6625),
    (-827.718, -10.0, -666.6625),
    (-827.718, 30.0, -626.5704),
    (-827.718, 30.5353, -626.302),
    (-827.718, 33.1609, -626.233),
    (-827.718, 37.4814, -626.0947),
    (-827.718, 42.261, -625.9372),
    (-827.718, 46.429, -625.7629),
    (-827.718, 51.0237, -625.5641),
    (-827.718, 55.0452, -625.3539),
    (-827.718, 59.4637, -625.1136),
    (-827.718, 66.5501, -624.629),
    (-827.718, 74.3035, -624.073),
    (-827.718, 80.9456, -623.4657),
    (-827.718, 88.1781, -622.7644),
    (-827.718, 94.415, -622.0355),
    (-827.718, 101.1805, -621.1864),
    (-827.718, 107.0511, -620.3373),
    (-827.718, 113.4029, -619.3378),
    (-827.718, 118.9463, -618.3704),
    (-827.718, 124.9374, -617.2173),
    (-827.718, 130.1925, -616.1339),
    (-827.718, 135.8758, -614.8237),
    (-827.718, 140.8808, -613.6271),
    (-827.718, 146.3094, -612.1556),
    (-827.718, 151.1019, -610.8493),
    (-827.718, 156.3292, -609.2119),
    (-827.718, 161.3049, -607.6835),
    (-827.718, 166.7818, -605.7249),
    (-827.718, 171.2486, -604.2032),
    (-827.718, 176.232, -602.2041),
    (-827.718, 178.4228, -601.4073),
    (-827.718, 180.9009, -600.3391),
    (-827.718, 183.0697, -599.519),
    (-827.718, 185.5473, -598.404),
    (-827.718, 187.698, -597.5613),
    (-827.718, 190.1826, -596.3987),
    (-827.718, 192.3189, -595.5343),
    (-827.718, 194.818, -594.3232),
    (-827.718, 199.8898, -575.6394),
]

_loft_profile_3 = [
    (-742.2929, 205.6799, -575.6394),
    (-742.2929, 206.4184, -578.1315),
    (-742.2929, 206.8483, -580.0611),
    (-742.2929, 207.5635, -583.0087),
    (-742.2929, 207.9772, -585.9662),
    (-742.2929, 208.1557, -588.0244),
    (-742.2929, 208.2758, -591.0031),
    (-742.2929, 208.0744, -593.8197),
    (-742.2929, 207.9028, -596.7906),
    (-742.2929, 207.479, -599.5557),
    (-742.2929, 207.1905, -600.9203),
    (-742.2929, 206.5781, -603.433),
    (-742.2929, 206.176, -604.7409),
    (-742.2929, 205.4388, -606.9876),
    (-742.2929, 204.6389, -608.9607),
    (-742.2929, 203.55, -611.3346),
    (-742.2929, 200.8671, -616.2786),
    (-742.2929, 200.1671, -617.3924),
    (-742.2929, 199.3967, -618.4094),
    (-742.2929, 195.8487, -622.9854),
    (-742.2929, 194.2896, -624.6107),
    (-742.2929, 189.9161, -628.7359),
    (-742.2929, 188.7666, -629.6729),
    (-742.2929, 187.5929, -630.4511),
    (-742.2929, 182.4365, -634.0299),
    (-742.2929, 181.2582, -634.6591),
    (-742.2929, 176.2564, -637.1906),
    (-742.2929, 175.0548, -637.6503),
    (-742.2929, 174.43, -637.8735),
    (-742.2929, 173.9891, -638.0379),
    (-742.2929, 170.2511, -639.4673),
    (-742.2929, 163.1873, -642.2845),
    (-742.2929, 160.9455, -643.0673),
    (-742.2929, 152.3518, -646.1905),
    (-742.2929, 151.528, -646.4489),
    (-742.2929, 145.6446, -648.3646),
    (-742.2929, 141.2686, -649.7884),
    (-742.2929, 131.5659, -652.5636),
    (-742.2929, 128.5011, -653.4043),
    (-742.2929, 119.3032, -655.6809),
    (-742.2929, 116.5947, -656.2969),
    (-742.2929, 108.4274, -658.0051),
    (-742.2929, 104.8121, -658.716),
    (-742.2929, 97.4374, -659.9872),
    (-742.2929, 92.9836, -660.7325),
    (-742.2929, 88.147, -661.4228),
    (-742.2929, 82.8826, -662.157),
    (-742.2929, 78.3117, -662.7047),
    (-742.2929, 72.2245, -663.3956),
    (-742.2929, 67.8865, -663.822),
    (-742.2929, 60.9759, -664.4407),
    (-742.2929, 56.8291, -664.7649),
    (-742.2929, 50.5269, -665.2029),
    (-742.2929, 49.7273, -665.2553),
    (-742.2929, 49.4722, -665.2721),
    (-742.2929, 49.3467, -665.2804),
    (-742.2929, 48.8962, -665.3083),
    (-742.2929, 48.4653, -665.3529),
    (-742.2929, 44.4168, -665.7851),
    (-742.2929, 40.4236, -666.58),
    (-742.2929, 40.0, -666.6625),
    (-742.2929, 30.0, -646.6625),
    (-742.2929, 30.0, -634.0184),
    (-742.2929, 30.0, -632.5146),
    (-742.2929, 30.0, -631.8591),
    (-742.2929, 30.0, -630.2983),
    (-742.2929, 30.0, -626.6504),
    (-742.2929, 30.2436, -626.2846),
    (-742.2929, 30.4654, -626.2782),
    (-742.2929, 31.9958, -626.2252),
    (-742.2929, 44.2779, -625.7843),
    (-742.2929, 45.7218, -625.7182),
    (-742.2929, 56.985, -625.0821),
    (-742.2929, 58.3411, -625.0034),
    (-742.2929, 60.6703, -624.8228),
    (-742.2929, 70.9653, -623.9915),
    (-742.2929, 73.0412, -623.7698),
    (-742.2929, 82.4562, -622.7404),
    (-742.2929, 91.1032, -621.5047),
    (-742.2929, 92.9034, -621.2507),
    (-742.2929, 100.8177, -619.809),
    (-742.2929, 102.3967, -619.5233),
    (-742.2929, 109.6486, -617.8718),
    (-742.2929, 111.0258, -617.5593),
    (-742.2929, 117.6852, -615.6935),
    (-742.2929, 118.8803, -615.36),
    (-742.2929, 125.0171, -613.275),
    (-742.2929, 126.0495, -612.9268),
    (-742.2929, 131.734, -610.6172),
    (-742.2929, 132.6221, -610.2613),
    (-742.2929, 137.9255, -607.7213),
    (-742.2929, 138.7437, -607.3363),
    (-742.2929, 143.7338, -604.5618),
    (-742.2929, 144.3784, -604.2127),
    (-742.2929, 153.7943, -597.8883),
    (-742.2929, 154.038, -597.7436),
    (-742.2929, 154.2748, -597.5956),
    (-742.2929, 154.5057, -597.4446),
    (-742.2929, 154.7315, -597.2906),
    (-742.2929, 163.8336, -589.921),
    (-742.2929, 161.9828, -575.6394),
]

_loft_profile_4 = [
    (-709.8999, 188.5226, -575.6394),
    (-709.8999, 189.113, -578.309),
    (-709.8999, 189.1799, -578.7679),
    (-709.8999, 189.219, -579.0809),
    (-709.8999, 189.6235, -581.9069),
    (-709.8999, 189.7489, -584.8123),
    (-709.8999, 189.7681, -586.1871),
    (-709.8999, 189.6499, -589.5401),
    (-709.8999, 189.4384, -591.0052),
    (-709.8999, 188.3347, -597.8764),
    (-709.8999, 187.7737, -599.6913),
    (-709.8999, 185.6854, -605.7094),
    (-709.8999, 184.4754, -608.1661),
    (-709.8999, 183.0188, -610.9599),
    (-709.8999, 181.1307, -613.8275),
    (-709.8999, 179.5191, -616.0911),
    (-709.8999, 177.6044, -618.4279),
    (-709.8999, 174.5782, -621.5191),
    (-709.8999, 173.1428, -622.9504),
    (-709.8999, 171.6666, -624.1604),
    (-709.8999, 168.7303, -626.4863),
    (-709.8999, 166.7428, -627.8677),
    (-709.8999, 165.5199, -628.6389),
    (-709.8999, 164.1581, -629.4476),
    (-709.8999, 162.2111, -630.5414),
    (-709.8999, 159.3247, -631.8865),
    (-709.8999, 159.2088, -631.932),
    (-709.8999, 158.2562, -632.3809),
    (-709.8999, 157.9711, -632.5191),
    (-709.8999, 157.4674, -632.785),
    (-709.8999, 139.0489, -639.8066),
    (-709.8999, 137.7544, -640.3839),
    (-709.8999, 119.6337, -646.4397),
    (-709.8999, 118.047, -647.0955),
    (-709.8999, 100.4613, -652.1571),
    (-709.8999, 98.5429, -652.826),
    (-709.8999, 81.4732, -656.9432),
    (-709.8999, 79.2664, -657.5701),
    (-709.8999, 62.6946, -660.784),
    (-709.8999, 60.1801, -661.3349),
    (-709.8999, 44.0876, -663.6794),
    (-709.8999, 41.3606, -664.1002),
    (-709.8999, 39.9608, -664.2496),
    (-709.8999, 24.3009, -665.7526),
    (-709.8999, 22.8423, -665.8627),
    (-709.8999, 13.3151, -666.2724),
    (-709.8999, 5.5833, -666.6139),
    (-709.8999, 5.5833, -655.0981),
    (-709.8999, 5.5833, -653.8154),
    (-709.8999, 5.5833, -652.8718),
    (-709.8999, 5.5833, -646.6625),
    (-709.8999, 7.362, -646.6625),
    (-709.8999, 30.0, -646.6625),
    (-709.8999, 30.0, -636.3341),
    (-709.8999, 30.0, -631.1957),
    (-709.8999, 30.0, -629.7139),
    (-709.8999, 30.0, -628.0284),
    (-709.8999, 30.0, -626.9552),
    (-709.8999, 30.0, -625.4389),
    (-709.8999, 30.7499, -625.3727),
    (-709.8999, 44.9244, -623.7338),
    (-709.8999, 46.4573, -623.537),
    (-709.8999, 47.4383, -623.3669),
    (-709.8999, 60.9438, -620.9764),
    (-709.8999, 62.257, -620.6834),
    (-709.8999, 75.2287, -617.5481),
    (-709.8999, 76.4795, -617.1869),
    (-709.8999, 77.677, -616.7535),
    (-709.8999, 90.2539, -612.8799),
    (-709.8999, 91.4129, -612.367),
    (-709.8999, 92.5434, -611.7711),
    (-709.8999, 104.861, -607.1668),
    (-709.8999, 105.978, -606.4779),
    (-709.8999, 118.1677, -601.1572),
    (-709.8999, 119.2851, -600.3666),
    (-709.8999, 120.4129, -599.4742),
    (-709.8999, 132.6104, -593.4414),
    (-709.8999, 133.1457, -592.9842),
    (-709.8999, 133.6879, -592.5026),
    (-709.8999, 142.4318, -587.2961),
    (-709.8999, 147.1349, -584.7281),
    (-709.8999, 144.7082, -575.6394),
]

_loft_profile_5 = [
    (-663.3697, 171.684, -575.6394),
    (-663.3697, 171.7761, -575.9841),
    (-663.3697, 172.0566, -577.2527),
    (-663.3697, 172.4522, -579.787),
    (-663.3697, 172.7133, -581.2378),
    (-663.3697, 173.0534, -583.1132),
    (-663.3697, 173.1109, -585.6063),
    (-663.3697, 173.2051, -590.5267),
    (-663.3697, 172.8692, -593.1113),
    (-663.3697, 172.2178, -597.8923),
    (-663.3697, 171.4337, -600.6256),
    (-663.3697, 170.1258, -605.0455),
    (-663.3697, 168.7879, -607.9603),
    (-663.3697, 166.9656, -611.8242),
    (-663.3697, 164.9244, -614.9139),
    (-663.3697, 162.7808, -618.0721),
    (-663.3697, 159.8827, -621.2872),
    (-663.3697, 157.6766, -623.6648),
    (-663.3697, 153.7848, -626.8993),
    (-663.3697, 151.7954, -628.4993),
    (-663.3697, 148.6256, -630.5573),
    (-663.3697, 147.5098, -631.2532),
    (-663.3697, 144.9375, -632.6621),
    (-663.3697, 143.0846, -633.6402),
    (-663.3697, 139.2037, -635.1198),
    (-663.3697, 127.5584, -640.3132),
    (-663.3697, 120.9562, -642.5196),
    (-663.3697, 109.7293, -647.1594),
    (-663.3697, 103.5398, -648.9409),
    (-663.3697, 91.9487, -652.9822),
    (-663.3697, 86.1509, -654.3807),
    (-663.3697, 74.1913, -657.7779),
    (-663.3697, 68.7561, -658.832),
    (-663.3697, 56.1193, -661.6007),
    (-663.3697, 51.0123, -662.3447),
    (-663.3697, 37.9849, -664.3549),
    (-663.3697, 31.2975, -665.0685),
    (-663.3697, 26.5124, -665.5278),
    (-663.3697, 19.7551, -666.0381),
    (-663.3697, 13.8748, -666.291),
    (-663.3697, 13.3907, -666.3063),
    (-663.3697, 10.6176, -666.4085),
    (-663.3697, 8.9615, -666.4747),
    (-663.3697, 3.5443, -666.6338),
    (-663.3697, 2.9549, -666.6431),
    (-663.3697, 1.7779, -666.6557),
    (-663.3697, 0.0, -666.6625),
    (-663.3697, -2.9158, -666.6625),
    (-663.3697, -6.4004, -666.6625),
    (-663.3697, -10.0, -666.6625),
    (-663.3697, -10.0, -646.6625),
    (-663.3697, 3.279, -646.6625),
    (-663.3697, 6.5059, -646.6625),
    (-663.3697, 9.9192, -646.6625),
    (-663.3697, 13.0786, -646.6625),
    (-663.3697, 15.67, -646.6625),
    (-663.3697, 17.5408, -646.6625),
    (-663.3697, 18.6718, -646.6625),
    (-663.3697, 30.0, -646.6625),
    (-663.3697, 30.0, -642.4528),
    (-663.3697, 30.0, -627.8582),
    (-663.3697, 30.0, -625.5017),
    (-663.3697, 31.6324, -625.3469),
    (-663.3697, 33.3398, -625.1961),
    (-663.3697, 38.6715, -624.5796),
    (-663.3697, 45.4618, -623.7078),
    (-663.3697, 49.8073, -622.9543),
    (-663.3697, 54.8278, -622.0657),
    (-663.3697, 60.5998, -620.7778),
    (-663.3697, 65.3806, -619.6222),
    (-663.3697, 70.8226, -618.0507),
    (-663.3697, 76.0327, -616.1653),
    (-663.3697, 80.5713, -614.7675),
    (-663.3697, 85.5519, -612.5631),
    (-663.3697, 90.4099, -610.0025),
    (-663.3697, 94.7418, -608.3833),
    (-663.3697, 99.4729, -605.465),
    (-663.3697, 103.6985, -603.6206),
    (-663.3697, 108.3544, -600.3266),
    (-663.3697, 113.0537, -596.6079),
    (-663.3697, 117.1512, -594.5813),
    (-663.3697, 119.3417, -592.7104),
    (-663.3697, 121.56, -590.7398),
    (-663.3697, 126.2295, -587.9594),
    (-663.3697, 127.4752, -586.7867),
    (-663.3697, 130.1192, -584.2975),
    (-663.3697, 127.8075, -575.6394),
]

_loft_profile_6 = [
    (-553.2745, 131.6952, -575.6394),
    (-553.2745, 131.8231, -576.1187),
    (-553.2745, 132.188, -577.4853),
    (-553.2745, 132.594, -579.0057),
    (-553.2745, 133.0485, -580.7079),
    (-553.2745, 133.5609, -582.6271),
    (-553.2745, 133.7163, -583.2089),
    (-553.2745, 133.8287, -583.792),
    (-553.2745, 133.8785, -584.0495),
    (-553.2745, 133.9069, -584.1961),
    (-553.2745, 134.2808, -586.1323),
    (-553.2745, 134.4469, -586.9899),
    (-553.2745, 134.6097, -587.8321),
    (-553.2745, 134.9633, -589.6551),
    (-553.2745, 134.9721, -589.8127),
    (-553.2745, 134.9732, -589.8326),
    (-553.2745, 135.0126, -590.5382),
    (-553.2745, 135.0192, -590.6513),
    (-553.2745, 135.0539, -591.2698),
    (-553.2745, 135.1195, -592.392),
    (-553.2745, 135.2057, -593.9031),
    (-553.2745, 135.337, -596.1234),
    (-553.2745, 134.9431, -601.9019),
    (-553.2745, 134.9043, -602.5039),
    (-553.2745, 133.9494, -607.5157),
    (-553.2745, 133.7452, -608.6189),
    (-553.2745, 132.4463, -612.7358),
    (-553.2745, 131.9249, -614.4321),
    (-553.2745, 130.541, -617.5208),
    (-553.2745, 129.5294, -619.8332),
    (-553.2745, 128.1426, -622.1909),
    (-553.2745, 126.6406, -624.8085),
    (-553.2745, 125.4679, -626.3963),
    (-553.2745, 123.3502, -629.3343),
    (-553.2745, 122.6915, -630.0893),
    (-553.2745, 120.9647, -632.0992),
    (-553.2745, 119.546, -633.5319),
    (-553.2745, 118.7152, -634.3606),
    (-553.2745, 118.4315, -634.6524),
    (-553.2745, 118.4006, -634.6887),
    (-553.2745, 118.0967, -634.9985),
    (-553.2745, 117.2106, -635.8307),
    (-553.2745, 115.1222, -637.7722),
    (-553.2745, 114.292, -638.5849),
    (-553.2745, 112.5258, -640.2024),
    (-553.2745, 111.3954, -641.2922),
    (-553.2745, 109.9424, -642.6007),
    (-553.2745, 108.5182, -643.9502),
    (-553.2745, 107.37, -644.9652),
    (-553.2745, 105.6578, -646.5568),
    (-553.2745, 104.8067, -647.2939),
    (-553.2745, 102.8116, -649.1099),
    (-553.2745, 102.2504, -649.5851),
    (-553.2745, 99.9767, -651.6074),
    (-553.2745, 99.699, -651.8368),
    (-553.2745, 97.1506, -654.0471),
    (-553.2745, 97.1504, -654.0473),
    (-553.2745, 97.1486, -654.0488),
    (-553.2745, 94.5559, -656.2194),
    (-553.2745, 94.3311, -656.4175),
    (-553.2745, 91.9596, -658.3447),
    (-553.2745, 91.5146, -658.725),
    (-553.2745, 89.3593, -660.4215),
    (-553.2745, 88.6982, -660.9681),
    (-553.2745, 87.9114, -661.5938),
    (-553.2745, 85.8733, -663.1494),
    (-553.2745, 84.8415, -663.9395),
    (-553.2745, 83.054, -665.2542),
    (-553.2745, 81.5092, -666.3429),
    (-553.2745, 80.2202, -667.294),
    (-553.2745, 78.8675, -668.208),
    (-553.2745, 77.3746, -669.2629),
    (-553.2745, 76.4156, -669.8853),
    (-553.2745, 75.0493, -670.8111),
    (-553.2745, 74.2128, -671.3335),
    (-553.2745, 72.7125, -672.3107),
    (-553.2745, 71.9958, -672.7405),
    (-553.2745, 70.3626, -673.7611),
    (-553.2745, 69.7632, -674.1056),
    (-553.2745, 67.9978, -675.1616),
    (-553.2745, 67.5132, -675.4278),
    (-553.2745, 65.6166, -676.5115),
    (-553.2745, 65.2444, -676.7066),
    (-553.2745, 63.2172, -677.8103),
    (-553.2745, 62.9551, -677.9411),
    (-553.2745, 60.7979, -679.0574),
    (-553.2745, 60.6436, -679.1305),
    (-553.2745, 58.3571, -680.2522),
    (-553.2745, 58.2975, -680.2788),
    (-553.2745, 57.1754, -680.7958),
    (-553.2745, 55.4024, -681.6036),
    (-553.2745, 55.3244, -681.6404),
    (-553.2745, 52.4551, -682.8545),
    (-553.2745, 52.2535, -682.9427),
    (-553.2745, 49.463, -684.034),
    (-553.2745, 49.1412, -684.1639),
    (-553.2745, 46.4227, -685.1409),
    (-553.2745, 45.9843, -685.3032),
    (-553.2745, 43.3311, -686.1738),
    (-553.2745, 42.7798, -686.3599),
    (-553.2745, 40.1848, -687.1319),
    (-553.2745, 39.5246, -687.3336),
    (-553.2745, 36.9808, -688.0138),
    (-553.2745, 36.2155, -688.2236),
    (-553.2745, 33.7155, -688.8186),
    (-553.2745, 32.8497, -689.0295),
    (-553.2745, 30.3856, -689.5451),
    (-553.2745, 29.424, -689.7508),
    (-553.2745, 26.9879, -690.1925),
    (-553.2745, 25.9355, -690.3872),
    (-553.2745, 23.5189, -690.7596),
    (-553.2745, 22.3812, -690.9382),
    (-553.2745, 19.9753, -691.2455),
    (-553.2745, 18.7582, -691.4036),
    (-553.2745, 16.802, -691.6081),
    (-553.2745, 15.759, -691.7187),
    (-553.2745, 13.7983, -691.884),
    (-553.2745, 12.7108, -691.9768),
    (-553.2745, 10.7403, -692.1041),
    (-553.2745, 9.6123, -692.1778),
    (-553.2745, 7.6259, -692.268),
    (-553.2745, 6.462, -692.3215),
    (-553.2745, 4.4536, -692.3754),
    (-553.2745, 3.2584, -692.4078),
    (-553.2745, 0.8597, -692.4474),
    (-553.2745, 0.0, -692.4497),
    (-553.2745, 2.5, -645.5389),
    (-553.2745, 5.6826, -644.6264),
    (-553.2745, 9.9902, -644.5538),
    (-553.2745, 18.8493, -644.0286),
    (-553.2745, 28.8107, -642.9545),
    (-553.2745, 37.5585, -641.3488),
    (-553.2745, 44.1238, -639.5985),
    (-553.2745, 49.7476, -637.4958),
    (-553.2745, 53.2161, -635.9077),
    (-553.2745, 55.6666, -634.5716),
    (-553.2745, 59.0029, -632.5),
    (-553.2745, 60.9206, -631.1135),
    (-553.2745, 64.136, -628.4801),
    (-553.2745, 65.6053, -627.1078),
    (-553.2745, 68.7166, -623.8305),
    (-553.2745, 69.8127, -622.5423),
    (-553.2745, 72.8463, -618.5365),
    (-553.2745, 73.6324, -617.4064),
    (-553.2745, 76.6274, -612.5858),
    (-553.2745, 77.1521, -611.6917),
    (-553.2745, 80.1628, -605.9691),
    (-553.2745, 80.4584, -605.3922),
    (-553.2745, 83.5558, -598.6794),
    (-553.2745, 83.5951, -598.5963),
    (-553.2745, 84.8295, -595.7217),
    (-553.2745, 85.1117, -595.0663),
    (-553.2745, 85.1246, -595.0318),
    (-553.2745, 86.6834, -591.2699),
    (-553.2745, 86.7722, -591.0223),
    (-553.2745, 88.0949, -587.3389),
    (-553.2745, 89.8583, -583.2786),
    (-553.2745, 88.2586, -577.2872),
    (-553.2745, 87.8186, -575.6394),
]

# Build smooth loft with multiple profiles (pass raw point lists; _make_solid_loft dedupes, resamples, and builds wires itself)
_loft_wires_smooth = [
    _loft_profile_1,
    _loft_profile_2,
    _loft_profile_3,
    _loft_profile_4,
    _loft_profile_5,
    _loft_profile_6,
]

loft_smooth = _make_solid_loft(_loft_wires_smooth)

print(f"Smooth multi-profile loft created:")
print(f"  Is Manifold: {loft_smooth.is_manifold}")
print(f"  Volume: {loft_smooth.volume:.2f} mm³")

# ── Tail multi-profile loft (10 sections, Z = 299 → -119) ───────────────────
# 10 planar profiles (each at a constant Z), lofted through _make_solid_loft
# which handles dedup, resample to uniform point count, wire build, and loft.
_tail_profile_1 = [
    (-1086.6036,  76.1503, 299.2014), (-1089.4499,  74.7713, 299.2014), (-1089.7954,  74.2652, 299.2014),
    (-1090.3780,  73.3834, 299.2014), (-1098.8364,  60.5331, 299.2014), (-1101.1633,  56.7034, 299.2014),
    (-1103.1880,  53.4693, 299.2014), (-1105.9768,  48.6716, 299.2014), (-1107.3123,  46.4471, 299.2014),
    (-1110.1220,  41.3971, 299.2014), (-1110.8154,  40.1937, 299.2014), (-1113.8727,  34.4592, 299.2014),
    (-1114.1389,  33.9781, 299.2014), (-1116.3101,  29.6329, 299.2014), (-1120.2229,  21.6732, 299.2014),
    (-1120.8953,  20.0753, 299.2014), (-1125.5853,   9.5473, 299.2014), (-1127.0014,   5.7740, 299.2014),
    (-1130.2640,  -2.3872, 299.2014), (-1131.5022,  -6.0628, 299.2014), (-1132.9281, -10.0000, 299.2014),
    (-1110.9131, -10.0000, 299.2014), (-1110.4789,  -8.9055, 299.2014), (-1110.0533,  -7.9466, 299.2014),
    (-1107.6449,  -2.4822, 299.2014), (-1107.3012,  -1.7783, 299.2014), (-1106.0019,   0.9734, 299.2014),
    (-1105.8581,   1.2495, 299.2014), (-1104.4820,   4.0432, 299.2014), (-1104.3738,   4.2430, 299.2014),
    (-1101.3265,  10.0754, 299.2014), (-1101.2916,  10.1366, 299.2014), (-1100.7806,  11.0601, 299.2014),
    (-1098.9879,  14.3066, 299.2014), (-1096.7505,  18.1792, 299.2014), (-1094.6169,  21.7212, 299.2014),
    (-1094.0429,  22.8393, 299.2014), (-1093.2180,  24.1829, 299.2014), (-1091.8876,  26.3499, 299.2014),
    (-1086.6036,  28.9099, 299.2014),
]
_tail_profile_2 = [
    (-1046.3776, 106.2042, 272.7934), (-1051.2489, 103.8441, 272.7934), (-1052.7190, 103.1319, 272.7934),
    (-1058.7128,  94.5478, 272.7934), (-1064.6110,  86.0944, 272.7934), (-1068.0435,  80.9507, 272.7934),
    (-1070.2562,  77.6331, 272.7934), (-1074.1007,  71.6835, 272.7934), (-1075.6913,  69.2152, 272.7934),
    (-1075.9984,  68.7216, 272.7934), (-1080.3661,  61.7160, 272.7934), (-1081.5569,  59.7380, 272.7934),
    (-1084.8630,  54.2577, 272.7934), (-1089.7160,  45.7654, 272.7934), (-1093.3150,  39.4766, 272.7934),
    (-1100.2114,  26.4578, 272.7934), (-1101.0233,  24.9041, 272.7934), (-1101.3191,  24.2805, 272.7934),
    (-1107.8700,  10.5850, 272.7934), (-1109.4927,   6.8671, 272.7934), (-1112.0762,   1.0026, 272.7934),
    (-1116.4183, -10.0000, 272.7934), (-1092.1524, -10.0000, 272.7934), (-1091.7896,  -9.0699, 272.7934),
    (-1090.3737,  -5.8052, 272.7934), (-1088.0576,  -0.4819, 272.7934), (-1086.6114,   2.5403, 272.7934),
    (-1084.0112,   7.9577, 272.7934), (-1082.5254,  10.8026, 272.7934), (-1079.6683,  16.2569, 272.7934),
    (-1078.9101,  17.6251, 272.7934), (-1077.3919,  20.3561, 272.7934), (-1076.6162,  21.7056, 272.7934),
    (-1075.0487,  24.4231, 272.7934), (-1073.4240,  27.1101, 272.7934), (-1070.1736,  32.4644, 272.7934),
    (-1068.9553,  34.3771, 272.7934), (-1066.5721,  38.1006, 272.7934), (-1065.2811,  40.0508, 272.7934),
    (-1062.8627,  43.6840, 272.7934), (-1061.4829,  45.6974, 272.7934), (-1059.0544,  49.2191, 272.7934),
    (-1057.9605,  50.7711, 272.7934), (-1055.8561,  53.7278, 272.7934), (-1055.1567,  54.7105, 272.7934),
    (-1046.3776,  58.9638, 272.7934),
]
_tail_profile_3 = [
    ( -968.5874, 164.5233, 221.7251), ( -977.8836, 160.0338, 221.7251), ( -980.4006, 156.6022, 221.7251),
    ( -981.2438, 155.3648, 221.7251), ( -983.2751, 152.4175, 221.7251), ( -991.9405, 139.9319, 221.7251),
    ( -994.5309, 136.2276, 221.7251), ( -995.3085, 135.0918, 221.7251), (-1004.7320, 120.4231, 221.7251),
    (-1010.1361, 112.0223, 221.7251), (-1016.7239, 100.9215, 221.7251), (-1021.7286,  92.6814, 221.7251),
    (-1026.1136,  85.0374, 221.7251), (-1028.4538,  80.8329, 221.7251), (-1030.3956,  77.2436, 221.7251),
    (-1034.7065,  69.0265, 221.7251), (-1036.0238,  66.4353, 221.7251), (-1041.9007,  54.4952, 221.7251),
    (-1045.4906,  46.7775, 221.7251), (-1049.3735,  38.3025, 221.7251), (-1051.7067,  32.7878, 221.7251),
    (-1055.0872,  24.9251, 221.7251), (-1056.2866,  21.8355, 221.7251), (-1062.0466,   6.9273, 221.7251),
    (-1063.4460,   3.3266, 221.7251), (-1064.3103,   0.8623, 221.7251), (-1067.1753,  -7.3471, 221.7251),
    (-1068.0370, -10.0000, 221.7251), (-1042.0966, -10.0000, 221.7251), (-1042.0294,  -9.7834, 221.7251),
    (-1040.1039,  -3.5046, 221.7251), (-1038.2919,   2.4059, 221.7251), (-1038.1143,   2.9358, 221.7251),
    (-1033.9910,  14.7934, 221.7251), (-1033.4262,  16.3081, 221.7251), (-1029.2925,  26.8997, 221.7251),
    (-1028.3249,  29.2518, 221.7251), (-1024.2143,  38.7363, 221.7251), (-1022.8379,  41.7979, 221.7251),
    (-1018.7834,  50.3182, 221.7251), (-1016.9994,  53.9804, 221.7251), (-1014.9919,  57.9366, 221.7251),
    (-1012.8411,  62.0387, 221.7251), (-1010.8403,  65.8264, 221.7251), (-1006.9661,  72.7629, 221.7251),
    (-1001.4456,  82.3029, 221.7251), ( -998.7540,  86.8464, 221.7251), ( -996.1401,  91.0695, 221.7251),
    ( -993.5862,  95.0286, 221.7251), ( -986.5752, 105.8930, 221.7251), ( -984.1255, 109.5549, 221.7251),
    ( -978.8036, 112.1333, 221.7251), ( -972.7855, 115.1158, 221.7251), ( -969.6449, 116.6722, 221.7251),
    ( -968.5874, 117.1737, 221.7251),
]
_tail_profile_4 = [
    ( -916.3170, 199.0448, 187.4101), ( -920.4778, 197.4199, 187.4101), ( -927.6234, 194.6293, 187.4101),
    ( -935.2396, 191.6550, 187.4101), ( -937.0441, 190.9503, 187.4101), ( -943.7854, 188.2997, 187.4101),
    ( -947.7501, 181.0469, 187.4101), ( -949.4914, 177.8534, 187.4101), ( -953.0091, 171.6075, 187.4101),
    ( -953.6715, 170.3998, 187.4101), ( -954.2882, 169.2429, 187.4101), ( -957.1578, 164.2552, 187.4101),
    ( -957.3540, 163.9111, 187.4101), ( -958.7803, 161.4816, 187.4101), ( -959.2826, 160.6810, 187.4101),
    ( -963.5576, 153.4886, 187.4101), ( -965.2497, 150.8241, 187.4101), ( -967.4422, 147.1133, 187.4101),
    ( -969.2552, 144.2264, 187.4101), ( -970.8402, 141.4577, 187.4101), ( -973.2417, 137.5047, 187.4101),
    ( -974.4111, 135.3776, 187.4101), ( -978.1831, 128.9151, 187.4101), ( -978.4394, 128.4256, 187.4101),
    ( -979.5431, 126.4362, 187.4101), ( -982.2912, 121.5923, 187.4101), ( -983.1908, 120.0964, 187.4101),
    ( -984.9117, 116.9933, 187.4101), ( -990.6234, 106.2435, 187.4101), ( -994.6444,  98.4390, 187.4101),
    ( -996.8728,  93.7210, 187.4101), ( -999.9134,  87.6676, 187.4101), (-1001.0719,  85.3547, 187.4101),
    (-1002.3159,  82.9167, 187.4101), (-1005.2225,  76.4756, 187.4101), (-1007.0673,  72.2936, 187.4101),
    (-1009.4407,  67.0363, 187.4101), (-1013.2553,  58.4227, 187.4101), (-1013.8056,  56.9804, 187.4101),
    (-1022.3267,  34.7018, 187.4101), (-1022.3286,  34.6957, 187.4101), (-1000.5739,  24.7650, 187.4101),
    (-1002.3332,  19.8594, 187.4101), (-1003.8997,  15.1005, 187.4101), (-1006.5163,   6.8626, 187.4101),
    (-1006.9863,   5.2316, 187.4101), (-1008.5266,  -0.3244, 187.4101), (-1009.8057,  -4.7329, 187.4101),
    (-1010.1834,  -6.0763, 187.4101), (-1011.1662, -10.0000, 187.4101), (-1035.8452, -10.0000, 187.4101),
    (-1030.5752,   8.9656, 187.4101), (-1029.8989,  11.5415, 187.4101), ( -997.5453,  32.9339, 187.4101),
    ( -997.0368,  34.2050, 187.4101), ( -995.7737,  37.2655, 187.4101), ( -993.2611,  43.5609, 187.4101),
    ( -992.2071,  46.1066, 187.4101), ( -989.2817,  52.7446, 187.4101), ( -986.2317,  59.4291, 187.4101),
    ( -985.1591,  61.6573, 187.4101), ( -983.1404,  65.7140, 187.4101), ( -980.8659,  70.4237, 187.4101),
    ( -979.5888,  72.9690, 187.4101), ( -976.4200,  79.0188, 187.4101), ( -972.1750,  86.8280, 187.4101),
    ( -971.9157,  87.2880, 187.4101), ( -971.5385,  87.9048, 187.4101), ( -962.6893, 104.1927, 187.4101),
    ( -958.8534, 111.2591, 187.4101), ( -953.6078, 119.8567, 187.4101), ( -948.7569, 128.1899, 187.4101),
    ( -948.3754, 128.7678, 187.4101), ( -947.6293, 129.9397, 187.4101), ( -944.7027, 134.8450, 187.4101),
    ( -944.0929, 135.8305, 187.4101), ( -941.6025, 140.0412, 187.4101), ( -940.6117, 141.7502, 187.4101),
    ( -939.4364, 143.7597, 187.4101), ( -939.0209, 144.5122, 187.4101), ( -928.8109, 148.4995, 187.4101),
    ( -916.7435, 153.2122, 187.4101), ( -916.3170, 153.3787, 187.4101),
]
_tail_profile_5 = [
    ( -936.4068, 135.5577, 158.4698), ( -937.4857, 135.3948, 158.4698), ( -939.2220, 129.6079, 158.4698),
    ( -941.3352, 125.2734, 158.4698), ( -942.3034, 123.4394, 158.4698), ( -942.9555, 122.2040, 158.4698),
    ( -945.4781, 117.0467, 158.4698), ( -946.5024, 115.1216, 158.4698), ( -948.7245, 110.6155, 158.4698),
    ( -951.6802, 104.9005, 158.4698), ( -954.9566,  98.1238, 158.4698), ( -961.5891,  84.1964, 158.4698),
    ( -963.3031,  80.3810, 158.4698), ( -967.3331,  70.8811, 158.4698), ( -971.3372,  61.8913, 158.4698),
    ( -972.1405,  59.9788, 158.4698), ( -975.1954,  52.3369, 158.4698), ( -976.7879,  48.2253, 158.4698),
    ( -978.8796,  42.5610, 158.4698), ( -981.1950,  36.0115, 158.4698), ( -982.3633,  32.5515, 158.4698),
    ( -985.3131,  23.2973, 158.4698), ( -985.6187,  22.2952, 158.4698), ( -986.7390,  18.3506, 158.4698),
    ( -988.9297,  10.9908, 158.4698), ( -989.7409,   7.8289, 158.4698), ( -992.2330,  -1.6846, 158.4698),
    ( -994.0481, -10.0000, 158.4698), (-1019.4020,  -5.4939, 158.4698), (-1019.9996,  -8.0926, 158.4698),
    (-1020.3368, -10.0000, 158.4698), (-1017.9995,   1.0360, 158.4698), (-1016.9127,   6.0596, 158.4698),
    (-1016.5895,   7.6683, 158.4698), (-1016.2417,   9.0956, 158.4698), (-1013.3637,  20.8290, 158.4698),
    (-1009.9476,  32.7224, 158.4698), (-1009.7465,  33.4158, 158.4698), (-1009.4554,  34.2943, 158.4698),
    (-1006.9140,  42.3843, 158.4698), (-1005.9766,  45.4901, 158.4698), (-1004.2719,  50.3030, 158.4698),
    (-1002.6120,  55.1927, 158.4698), (-1001.6244,  57.7241, 158.4698), ( -999.0627,  64.5980, 158.4698),
    ( -998.7332,  65.3402, 158.4698), ( -998.1414,  66.7559, 158.4698), ( -993.7885,  77.3446, 158.4698),
    ( -991.7770,  82.6129, 158.4698), ( -988.6280,  89.2825, 158.4698), ( -984.0851,  99.6936, 158.4698),
    ( -983.6358, 100.5714, 158.4698), ( -981.5734, 104.8914, 158.4698), ( -980.5438, 107.0731, 158.4698),
    ( -980.1583, 107.9291, 158.4698), ( -977.6805, 112.9318, 158.4698), ( -976.2392, 115.9832, 158.4698),
    ( -975.0236, 118.3469, 158.4698), ( -973.0777, 122.3065, 158.4698), ( -972.5705, 123.2624, 158.4698),
    ( -970.8008, 126.7544, 158.4698), ( -970.0948, 128.1537, 158.4698), ( -969.9142, 128.5301, 158.4698),
    ( -966.7794, 134.6841, 158.4698), ( -965.8911, 136.5233, 158.4698), ( -965.5656, 137.1696, 158.4698),
    ( -962.2879, 136.5091, 158.4698), ( -959.6248, 136.2600, 158.4698), ( -954.4257, 135.8371, 158.4698),
    ( -948.8800, 135.2898, 158.4698), ( -948.1337, 135.2363, 158.4698), ( -947.3124, 135.1607, 158.4698),
    ( -944.4492, 135.2083, 158.4698),
]
_tail_profile_6 = [
    ( -953.3171,  83.6690, 111.9376), ( -953.3590,  83.5386, 111.9376), ( -953.7459,  82.4308, 111.9376),
    ( -956.3690,  84.5532, 111.9376), ( -960.7903,  85.8284, 111.9376), ( -969.4831,  88.3023, 111.9376),
    ( -969.6628,  88.3520, 111.9376), ( -972.7441,  89.3139, 111.9376), ( -981.8253,  92.2197, 111.9376),
    ( -982.6922,  92.5291, 111.9376), ( -984.2577,  87.7947, 111.9376), ( -984.5429,  86.9392, 111.9376),
    ( -984.5670,  86.8500, 111.9376), ( -984.9205,  85.7647, 111.9376), ( -986.1733,  81.9408, 111.9376),
    ( -986.4055,  81.2875, 111.9376), ( -988.0025,  76.2241, 111.9376), ( -988.7206,  74.1313, 111.9376),
    ( -989.8589,  70.3563, 111.9376), ( -991.0034,  66.8660, 111.9376), ( -992.1232,  62.8151, 111.9376),
    ( -994.4956,  55.0441, 111.9376), ( -995.0991,  53.0302, 111.9376), ( -995.4461,  51.9717, 111.9376),
    ( -997.7586,  43.2078, 111.9376), ( -999.6825,  36.5286, 111.9376), (-1000.1202,  34.6840, 111.9376),
    (-1001.5814,  28.9354, 111.9376), (-1001.6462,  28.6802, 111.9376), (-1001.6746,  28.5739, 111.9376),
    (-1002.9640,  23.0118, 111.9376), (-1003.5891,  20.4484, 111.9376), (-1004.3131,  16.9393, 111.9376),
    (-1005.6525,  10.7712, 111.9376), (-1005.6966,  10.5675, 111.9376), (-1005.7235,  10.4489, 111.9376),
    (-1006.8994,   4.1799, 111.9376), (-1007.6876,   0.1554, 111.9376), (-1008.0559,  -2.1692, 111.9376),
    (-1008.6279,  -5.6446, 111.9376), (-1009.2248,  -9.2114, 111.9376), (-1009.3638, -10.0000, 111.9376),
    ( -977.8686, -10.0000, 111.9376), ( -976.7808,  -3.5465, 111.9376), ( -976.4327,  -1.6512, 111.9376),
    ( -975.0926,   5.2943, 111.9376), ( -974.2044,   9.5951, 111.9376), ( -973.2497,  13.9873, 111.9376),
    ( -971.7650,  20.4701, 111.9376), ( -971.2656,  22.5379, 111.9376), ( -969.2067,  30.7539, 111.9376),
    ( -969.1582,  30.9531, 111.9376), ( -969.1450,  31.0057, 111.9376), ( -966.9696,  39.2405, 111.9376),
    ( -965.4066,  44.8496, 111.9376), ( -962.3243,  55.4499, 111.9376), ( -957.6826,  70.3720, 111.9376),
    ( -957.4086,  71.2143, 111.9376), ( -956.9051,  72.7203, 111.9376),
]
_tail_profile_7 = [
    ( -971.4583,  45.4801,  44.3712), ( -971.5086,  44.0678,  44.3712), ( -973.6117,  45.5980,  44.3712),
    ( -974.5336,  45.6323,  44.3712), ( -977.8340,  45.7723,  44.3712), ( -980.3962,  45.9087,  44.3712),
    ( -982.3675,  46.0170,  44.3712), ( -986.2526,  46.1749,  44.3712), ( -987.4893,  46.2425,  44.3712),
    ( -992.1058,  46.4280,  44.3712), ( -992.5657,  46.4529,  44.3712), ( -995.4229,  46.5658,  44.3712),
    ( -997.6082,  46.6873,  44.3712), ( -997.9544,  46.7017,  44.3712), (-1000.3323,  46.8326,  44.3712),
    (-1000.7026,  46.8481,  44.3712), (-1003.0032,  46.9745,  44.3712), (-1003.6141,  47.0000,  44.3712),
    (-1004.3364,  47.0309,  44.3712), (-1006.5311,  47.1489,  44.3712), (-1007.5137,  47.1907,  44.3712),
    (-1009.4379,  47.2938,  44.3712), (-1011.0448,  47.3187,  44.3712), (-1011.4727,  47.3311,  44.3712),
    (-1011.7153,  46.5324,  44.3712), (-1011.7153,  46.5288,  44.3712), (-1011.7153,  17.4119,  44.3712),
    (-1011.7155,  17.3999,  44.3712), (-1011.7156,  17.3989,  44.3712), (-1011.7157,  17.3978,  44.3712),
    (-1011.7158,  17.3954,  44.3712), (-1011.7158, -10.0000,  44.3712), ( -972.1886, -10.0000,  44.3712),
    ( -972.1644,  -8.7301,  44.3712), ( -972.1644,  -0.4557,  44.3712), ( -972.0886,   3.3343,  44.3712),
    ( -972.0084,   6.9886,  44.3712), ( -971.8362,  13.9239,  44.3712), ( -971.6506,  20.4124,  44.3712),
    ( -971.6506,  39.1495,  44.3712),
]
_tail_profile_8 = [
    ( -883.2574, 252.3725, -21.9876), ( -888.9629, 255.0488, -21.9876), ( -891.1860, 256.0917, -21.9876),
    ( -895.8114, 257.8868, -21.9876), ( -897.3817, 258.4955, -21.9876), ( -898.4474, 258.9088, -21.9876),
    ( -902.9303, 259.9881, -21.9876), ( -906.0714, 260.7429, -21.9876), ( -909.4331, 261.1623, -21.9876),
    ( -911.4369, 261.4109, -21.9876), ( -915.0688, 261.5295, -21.9876), ( -916.8520, 261.5863, -21.9876),
    ( -920.7804, 261.3538, -21.9876), ( -922.2745, 261.2639, -21.9876), ( -926.5163, 260.6237, -21.9876),
    ( -927.6603, 260.4497, -21.9876), ( -932.2196, 259.3367, -21.9876), ( -932.9642, 259.1539, -21.9876),
    ( -937.8310, 257.4931, -21.9876), ( -938.1415, 257.3866, -21.9876), ( -941.5547, 255.8670, -21.9876),
    ( -943.1476, 255.1582, -21.9876), ( -943.2831, 255.0828, -21.9876), ( -947.9394, 252.4811, -21.9876),
    ( -949.0095, 251.6984, -21.9876), ( -955.8704, 246.6601, -21.9876), ( -957.4265, 245.1750, -21.9876),
    ( -960.7742, 241.9655, -21.9876), ( -962.7790, 239.6011, -21.9876), ( -965.1620, 236.7769, -21.9876),
    ( -967.4538, 233.4064, -21.9876), ( -968.9773, 231.1532, -21.9876), ( -969.7054, 229.8110, -21.9876),
    ( -972.1672, 225.1836, -21.9876), ( -974.7268, 219.2306, -21.9876), ( -979.4117, 208.6700, -21.9876),
    ( -980.7153, 205.3529, -21.9876), ( -986.0705, 191.7102, -21.9876), ( -987.9734, 186.3664, -21.9876),
    ( -990.0176, 180.5870, -21.9876), ( -992.2886, 173.5060, -21.9876), ( -993.9584, 168.4101, -21.9876),
    ( -996.5342, 159.4039, -21.9876), ( -998.9263, 151.1707, -21.9876), (-1001.6051, 140.4677, -21.9876),
    (-1002.4618, 137.0684, -21.9876), (-1003.1577, 133.9891, -21.9876), (-1005.7640, 121.9238, -21.9876),
    (-1006.0048, 120.7571, -21.9876), (-1006.2219, 119.6365, -21.9876), (-1008.7627, 105.7456, -21.9876),
    (-1009.0232, 104.2569, -21.9876), (-1009.0741, 103.9537, -21.9876), (-1009.8830,  98.9251, -21.9876),
    (-1010.0709,  97.6769, -21.9876), (-1010.8494,  92.4939, -21.9876), (-1011.4686,  88.0803, -21.9876),
    (-1013.0144,  75.6789, -21.9876), (-1013.0405,  75.4728, -21.9876), (-1013.6415,  69.6133, -21.9876),
    (-1014.2963,  63.2488, -21.9876), (-1014.3327,  62.8905, -21.9876), (-1014.3571,  62.6336, -21.9876),
    (-1014.3680,  62.5111, -21.9876), (-1014.3756,  62.4177, -21.9876), (-1015.3581,  50.3633, -21.9876),
    (-1015.3804,  50.0750, -21.9876), (-1015.3948,  49.8772, -21.9876), (-1015.4158,  49.6763, -21.9876),
    (-1015.4499,  49.4250, -21.9876), (-1016.0060,  44.8773, -21.9876), (-1016.0242,  44.7814, -21.9876),
    (-1016.2542,  43.5706, -21.9876), (-1016.3164,  43.2513, -21.9876), (-1016.4738,  42.6163, -21.9876),
    (-1017.1086,  40.0000, -21.9876), (-1017.1086,  16.5270, -21.9876), (-1017.1086, -10.0000, -21.9876),
    ( -976.9563, -10.0000, -21.9876), ( -976.9563,  -6.0724, -21.9876), ( -976.9563,   6.9398, -21.9876),
    ( -976.9563,  29.3834, -21.9876), ( -976.9563,  29.8764, -21.9876), ( -976.9563,  30.6162, -21.9876),
    ( -976.9563,  41.4864, -21.9876), ( -976.9563,  42.1854, -21.9876), ( -976.9563,  43.3057, -21.9876),
    ( -976.9563,  53.7959, -21.9876), ( -976.9563,  54.7395, -21.9876), ( -976.9563,  56.2150, -21.9876),
    ( -976.9563,  66.2757, -21.9876), ( -976.9563,  67.2934, -21.9876), ( -976.9563,  68.8165, -21.9876),
    ( -976.9563,  79.2527, -21.9876), ( -976.9563,  79.5764, -21.9876), ( -976.9563,  80.0099, -21.9876),
    ( -976.9563,  80.4763, -21.9876), ( -976.9563,  82.1673, -21.9876), ( -975.9491,  82.2628, -21.9876),
    ( -973.2953,  82.6975, -21.9876), ( -972.3409,  91.8911, -21.9876), ( -972.3183,  92.0769, -21.9876),
    ( -970.7023, 104.7007, -21.9876), ( -970.6243, 105.2106, -21.9876), ( -968.8100, 117.0014, -21.9876),
    ( -968.6561, 117.8158, -21.9876), ( -966.6628, 128.8081, -21.9876), ( -966.4081, 129.9213, -21.9876),
    ( -964.2566, 140.1507, -21.9876), ( -963.8744, 141.5559, -21.9876), ( -961.5870, 151.0593, -21.9876),
    ( -961.0490, 152.7509, -21.9876), ( -958.6491, 161.5662, -21.9876), ( -956.2599, 169.4733, -21.9876),
    ( -955.4381, 171.7048, -21.9876), ( -952.9902, 179.0258, -21.9876), ( -951.9481, 181.5153, -21.9876),
    ( -949.4648, 188.2888, -21.9876), ( -948.1738, 191.0414, -21.9876), ( -945.6802, 197.3130, -21.9876),
    ( -944.2061, 200.1664, -21.9876), ( -942.2450, 204.7474, -21.9876), ( -940.5286, 207.8172, -21.9876),
    ( -938.4141, 212.4907, -21.9876), ( -936.1778, 217.1832, -21.9876), ( -934.2080, 220.4585, -21.9876),
    ( -930.7248, 226.9145, -21.9876), ( -929.8420, 228.3286, -21.9876), ( -929.6509, 228.6058, -21.9876),
    ( -918.0540, 223.1658, -21.9876), ( -900.7378, 215.0430, -21.9876), ( -887.0308, 244.3144, -21.9876),
]
_tail_profile_9 = [
    ( -886.2950, 245.6983, -51.4576), ( -890.4415, 247.6433, -51.4576), ( -894.7110, 249.6461, -51.4576),
    ( -901.1165, 252.1320, -51.4576), ( -901.9245, 252.4471, -51.4576), ( -903.6567, 252.8684, -51.4576),
    ( -909.5071, 254.2798, -51.4576), ( -910.7765, 254.4429, -51.4576), ( -914.8438, 254.9532, -51.4576),
    ( -916.1777, 255.0031, -51.4576), ( -920.2302, 255.1393, -51.4576), ( -921.6306, 255.0646, -51.4576),
    ( -925.6249, 254.8329, -51.4576), ( -927.0916, 254.6218, -51.4576), ( -930.9843, 254.0400, -51.4576),
    ( -932.5142, 253.6792, -51.4576), ( -936.2643, 252.7703, -51.4576), ( -937.8510, 252.2444, -51.4576),
    ( -941.4206, 251.0338, -51.4576), ( -943.0540, 250.3258, -51.4576), ( -946.4093, 248.8405, -51.4576),
    ( -948.0759, 247.9326, -51.4576), ( -951.1880, 246.2024, -51.4576), ( -954.1887, 244.0534, -51.4576),
    ( -959.1077, 240.4594, -51.4576), ( -961.1576, 238.5462, -51.4576), ( -964.0137, 235.8229, -51.4576),
    ( -965.9933, 233.5445, -51.4576), ( -968.4123, 230.6946, -51.4576), ( -970.2663, 228.0427, -51.4576),
    ( -972.2478, 225.1320, -51.4576), ( -975.2869, 219.5034, -51.4576), ( -975.4654, 219.1800, -51.4576),
    ( -975.5781, 218.9460, -51.4576), ( -982.6566, 203.1279, -51.4576), ( -989.0925, 186.7531, -51.4576),
    ( -990.1004, 184.4118, -51.4576), ( -993.1736, 175.7334, -51.4576), ( -993.5144, 174.8550, -51.4576),
    ( -996.4921, 165.7951, -51.4576), ( -996.7115, 165.1840, -51.4576), (-1002.4594, 145.4383, -51.4576),
    (-1002.5157, 145.2600, -51.4576), (-1003.4300, 141.6392, -51.4576), (-1005.3221, 134.1321, -51.4576),
    (-1007.6443, 123.8551, -51.4576), (-1007.9644, 122.3734, -51.4576), (-1009.0243, 117.2389, -51.4576),
    (-1009.9794, 112.3074, -51.4576), (-1010.4246, 109.8736, -51.4576), (-1012.1440, 100.0448, -51.4576),
    (-1012.1551,  99.9790, -51.4576), (-1012.3305,  98.8885, -51.4576), (-1013.7406,  89.5193, -51.4576),
    (-1013.9353,  88.2231, -51.4576), (-1014.0902,  87.1194, -51.4576), (-1014.4767,  84.0181, -51.4576),
    (-1015.6492,  74.7626, -51.4576), (-1016.0204,  71.1437, -51.4576), (-1016.9548,  62.1420, -51.4576),
    (-1017.2776,  58.1763, -51.4576), (-1017.9615,  49.8756, -51.4576), (-1018.2988,  47.1181, -51.4576),
    (-1018.5756,  44.8772, -51.4576), (-1018.7651,  43.8795, -51.4576), (-1018.8875,  43.2514, -51.4576),
    (-1019.1971,  42.0022, -51.4576), (-1019.6829,  40.0000, -51.4576), ( -978.9986,  34.2099, -51.4576),
    ( -979.0847,  30.8682, -51.4576), ( -979.3977,  17.5613, -51.4576), ( -979.4256,  14.0702, -51.4576),
    ( -979.5306,   0.0000, -51.4576), ( -979.5306,  -7.2712, -51.4576), ( -979.5306, -10.0000, -51.4576),
    (-1019.6829, -10.0000, -51.4576), (-1019.6829,  -6.1288, -51.4576), ( -978.4821,  46.7602, -51.4576),
    ( -978.3316,  50.0200, -51.4576), ( -977.6188,  61.8240, -51.4576), ( -977.3942,  65.0557, -51.4576),
    ( -976.4944,  76.1270, -51.4576), ( -976.2152,  79.0746, -51.4576), ( -975.1421,  89.4110, -51.4576),
    ( -974.7771,  92.4117, -51.4576), ( -973.5458, 102.0308, -51.4576), ( -973.0346, 105.3745, -51.4576),
    ( -971.6590, 114.3145, -51.4576), ( -971.0106, 117.7446, -51.4576), ( -969.5082, 126.0296, -51.4576),
    ( -968.7002, 129.5615, -51.4576), ( -967.0900, 137.2171, -51.4576), ( -966.0981, 140.8645, -51.4576),
    ( -964.4005, 147.9175, -51.4576), ( -963.1988, 151.6957, -51.4576), ( -961.4352, 158.1739, -51.4576),
    ( -959.6794, 163.9847, -51.4576), ( -958.1901, 168.0286, -51.4576), ( -956.4120, 173.3468, -51.4576),
    ( -954.6597, 177.5324, -51.4576), ( -952.8802, 182.3862, -51.4576), ( -950.8395, 186.7374, -51.4576),
    ( -949.0812, 191.1597, -51.4576), ( -946.8676, 195.4445, -51.4576), ( -945.5094, 198.6171, -51.4576),
    ( -943.0081, 203.0908, -51.4576), ( -941.5746, 206.2591, -51.4576), ( -940.0586, 209.4403, -51.4576),
    ( -937.2944, 214.0366, -51.4576), ( -934.9927, 218.3026, -51.4576), ( -933.0986, 221.3368, -51.4576),
    ( -932.7430, 221.8525, -51.4576), ( -932.6885, 221.9316, -51.4576), ( -930.7650, 221.0294, -51.4576),
    ( -903.7754, 208.3688, -51.4576), ( -894.1820, 228.8555, -51.4576),
]
_tail_profile_10 = [
    ( -893.2323, 230.4558, -118.7609), ( -893.8185, 230.7308, -118.7609), ( -902.7613, 234.9258, -118.7609),
    ( -904.3389, 235.5380, -118.7609), ( -905.8823, 236.1397, -118.7609), ( -908.9569, 237.3395, -118.7609),
    ( -909.9060, 237.7099, -118.7609), ( -910.8251, 237.9417, -118.7609), ( -917.2945, 239.5235, -118.7609),
    ( -917.3839, 239.5387, -118.7609), ( -918.2009, 239.6464, -118.7609), ( -922.6471, 240.2250, -118.7609),
    ( -923.5736, 240.2636, -118.7609), ( -927.9606, 240.4370, -118.7609), ( -929.0141, 240.3864, -118.7609),
    ( -933.2846, 240.1698, -118.7609), ( -934.4788, 240.0055, -118.7609), ( -938.5772, 239.4287, -118.7609),
    ( -939.9206, 239.1220, -118.7609), ( -943.7957, 238.2229, -118.7609), ( -945.2907, 237.7408, -118.7609),
    ( -948.8980, 236.5617, -118.7609), ( -950.5398, 235.8677, -118.7609), ( -953.8416, 234.4545, -118.7609),
    ( -955.6182, 233.5096, -118.7609), ( -958.5859, 231.9124, -118.7609), ( -962.0288, 229.5050, -118.7609),
    ( -966.4740, 226.3615, -118.7609), ( -968.9910, 224.0711, -118.7609), ( -971.3824, 221.8683, -118.7609),
    ( -973.9341, 219.0113, -118.7609), ( -975.8049, 216.8887, -118.7609), ( -978.3012, 213.4280, -118.7609),
    ( -979.6874, 211.4772, -118.7609), ( -982.0215, 207.3890, -118.7609), ( -982.5158, 206.5233, -118.7609),
    ( -982.9726, 205.6957, -118.7609), ( -988.2104, 194.8229, -118.7609), ( -989.1295, 192.7691, -118.7609),
    ( -989.9651, 190.6430, -118.7609), ( -996.4655, 175.5432, -118.7609), ( -996.9895, 174.0636, -118.7609),
    ( -999.8281, 166.7463, -118.7609), (-1000.3875, 165.0443, -118.7609), (-1002.9818, 157.8184, -118.7609),
    (-1004.1515, 153.8002, -118.7609), (-1008.6883, 139.4228, -118.7609), (-1009.3499, 136.8027, -118.7609),
    (-1011.4592, 129.0533, -118.7609), (-1012.1115, 126.1675, -118.7609), (-1013.9834, 118.4367, -118.7609),
    (-1014.3030, 116.8891, -118.7609), (-1015.1527, 113.0301, -118.7609), (-1015.4649, 111.4189, -118.7609),
    (-1016.2601, 107.5550, -118.7609), (-1016.8551, 104.1545, -118.7609), (-1018.2890,  96.3903, -118.7609),
    (-1018.8377,  92.7437, -118.7609), (-1020.0694,  84.9248, -118.7609), (-1020.5617,  81.0417, -118.7609),
    (-1021.6010,  73.1406, -118.7609), (-1022.0277,  69.0324, -118.7609), (-1022.8833,  61.0195, -118.7609),
    (-1023.3217,  55.7448, -118.7609), (-1023.8193,  49.8763, -118.7609), (-1024.1634,  47.1479, -118.7609),
    (-1024.4440,  44.8770, -118.7609), (-1024.4995,  44.5851, -118.7609), (-1024.7593,  43.2515, -118.7609),
    (-1025.4166,  40.5998, -118.7609), (-1025.5621,  40.0000, -118.7609), ( -984.6617,  39.9179, -118.7609),
    ( -984.8269,  35.9038, -118.7609), ( -985.1630,  22.8484, -118.7609), ( -985.2641,  18.5501, -118.7609),
    ( -985.3758,   4.5596, -118.7609), ( -985.4098,   0.0000, -118.7609), ( -985.4098, -10.0000, -118.7609),
    (-1025.5621, -10.0000, -118.7609), ( -984.0968,  52.1530, -118.7609), ( -983.8726,  55.8651, -118.7609),
    ( -983.0720,  67.3818, -118.7609), ( -982.7960,  70.7788, -118.7609), ( -981.8590,  80.6689, -118.7609),
    ( -981.5394,  83.7470, -118.7609), ( -980.3927,  93.1763, -118.7609), ( -980.0398,  95.9332, -118.7609),
    ( -978.5392, 105.7487, -118.7609), ( -978.1654, 108.1780, -118.7609), ( -976.3878, 117.5821, -118.7609),
    ( -976.0066, 119.6842, -118.7609), ( -973.9349, 128.7398, -118.7609), ( -973.5611, 130.5173, -118.7609),
    ( -971.1767, 139.2856, -118.7609), ( -970.8260, 140.7423, -118.7609), ( -968.1085, 149.2858, -118.7609),
    ( -967.7980, 150.4266, -118.7609), ( -967.4888, 151.4498, -118.7609), ( -964.4752, 159.6330, -118.7609),
    ( -964.2264, 160.3771, -118.7609), ( -960.8526, 168.4362, -118.7609), ( -960.6804, 168.9059, -118.7609),
    ( -956.9274, 176.9080, -118.7609), ( -956.8483, 177.1070, -118.7609), ( -953.8386, 182.9327, -118.7609),
    ( -953.1749, 184.2217, -118.7609), ( -952.9530, 184.6411, -118.7609), ( -950.0610, 189.8245, -118.7609),
    ( -948.7529, 192.1033, -118.7609), ( -946.8163, 195.3484, -118.7609), ( -944.3098, 199.4284, -118.7609),
    ( -943.6929, 200.3494, -118.7609), ( -943.4068, 200.7699, -118.7609), ( -940.5359, 205.3689, -118.7609),
    ( -940.4298, 205.5227, -118.7609), ( -939.6263, 206.6883, -118.7609), ( -911.2716, 193.3886, -118.7609),
    ( -910.7127, 193.1263, -118.7609), ( -910.5140, 193.5507, -118.7609),
]

loft_tail = _make_solid_loft([
    _tail_profile_1, _tail_profile_2, _tail_profile_3, _tail_profile_4, _tail_profile_5,
    _tail_profile_6, _tail_profile_7, _tail_profile_8, _tail_profile_9, _tail_profile_10,
])

print(f"Tail multi-profile loft created:")
print(f"  Is Manifold: {loft_tail.is_manifold}")
print(f"  Volume: {loft_tail.volume:.2f} mm³")

# ── New connector loft ─────────────────────────────────────────────────────
_conn_profile_1 = [
    (-921.4195, 230.41803, -370.8729),
    (-921.4195, 234.81939, -395.31115),
    (-921.4195, 271.30896, -577.14053),
    (-921.4195, 273.61766, -591.76244),
    (-921.4195, 273.61766, -604.90959),
    (-921.4195, 267.77448, -616.23075),
    (-921.4195, 256.81852, -624.63033),
    (-921.4195, 245.104,   -629.04826),
    (-921.4195, 194.47942, -430.50496),
    (-921.4195, 191.24893, -407.03413),
    (-921.4195, 188.94235, -384.54495),
    (-921.4195, 186.05912, -362.63241),
    (-921.4195, 186.05912, -344.7564),
    (-921.4195, 186.05912, -277.76582),
    (-921.4195, 201.1014,  -274.83415),
    (-921.4195, 216.73694, -268.48222),
    (-921.4195, 230.41803, -261.15306),
]
_conn_profile_2 = [
    (-924.82528,  297.08645,  -68.42741),
    (-949.20546,  296.05641,  -97.48259),
    (-966.76681,  296.05641, -118.41139),
    (-996.49222,  296.05641, -153.83675),
    (-1026.21763, 296.05641, -189.26211),
    (-1070.41561, 285.53619, -241.93522),
    (-1077.72354, 273.86484, -250.64447),
    (-1086.80321, 256.07408, -261.46519),
    (-1094.31366, 239.25801, -270.41581),
    (-1101.68294, 219.88809, -279.19817),
    (-1071.22958, 229.37958, -242.90527),
    (-1032.75152, 238.02644, -197.0489),
    (-1002.91403, 238.02644, -161.48997),
    (-977.42485,  238.02644, -131.11314),
    (-939.93974,  238.02644,  -86.44013),
    (-932.55654,  256.07408,  -77.64117),
]
loft_connector = _make_solid_loft([_conn_profile_1, _conn_profile_2])
print(f"Connector loft: manifold={loft_connector.is_manifold}  volume={loft_connector.volume:.2f} mm³")

# ── New loft 2 (quad profiles) ─────────────────────────────────────────────
_loft_new2_a = [
    (-1019.61991, -80.0, -442.62558),
    (-1026.70596, -80.0, -386.31913),
    (-979.53248,  -80.0, -442.62558),
    (-984.34203,  -80.0, -382.13167),
]
_loft_new2_b = [
    (-1037.88419, 222.0302, -383.73012),
    (-1013.67316, 222.0302, -420.07424),
    (-988.2955,   222.0302, -402.32579),
    (-1008.43084, 222.0302, -368.76027),
]
loft_new2 = _make_solid_loft([_loft_new2_a, _loft_new2_b])
print(f"New loft2: manifold={loft_new2.is_manifold}  volume={loft_new2.volume:.2f} mm³")

# ── Cylinder feature ──────────────────────────────────────
cyl_main = {"pos": (-1125.9, 120.5, -27.1756), "rot": (295, -30, -70), "od": 520.0, "id": 410.6, "height": 50.8}
cyl_sub  = {"pos": (-1125.9, 120.5 , -27.1756), "rot": (295, -30, -70), "od": 450.4, "height": 40.0}

with BuildPart() as cyl_part:
    with Locations(Location(cyl_main["pos"], cyl_main["rot"])):
        Cylinder(radius=cyl_main["od"] / 2, height=cyl_main["height"])
        Cylinder(radius=cyl_main["id"] / 2, height=cyl_main["height"], mode=Mode.SUBTRACT)
    with Locations(Location(cyl_sub["pos"], cyl_sub["rot"])):
        Cylinder(radius=cyl_sub["od"] / 2, height=cyl_sub["height"], mode=Mode.SUBTRACT)

print(f"Cylinder OK | Volume: {cyl_part.part.volume:.2f} mm3")

cbb = cyl_part.part.bounding_box()
print(f"Cyl bbox: X:{cbb.min.X:.1f}→{cbb.max.X:.1f}  Y:{cbb.min.Y:.1f}→{cbb.max.Y:.1f}  Z:{cbb.min.Z:.1f}→{cbb.max.Z:.1f}")


# ── Normal-direction cuts ──   ← original code continues

# ── Normal-direction cuts ──────────────────────────────────────────────────
# Each entry = (4-point rectangular profile, depth mm). Cut is extruded along
# the profile's own normal; direction is chosen automatically per cut (tries
# +depth first, retries -depth if volume doesn't decrease).
_normal_cut_profiles = [
    # First batch: 9 cuts @ 27 mm across swept_combined
    ([(-429.1895,   6.8946,  -595.9571), (-477.4642,  19.9141, -596.1644), (-488.1176, -17.2958, -452.2793), (-439.8429, -30.3152, -452.0719)], 27.0),
    ([(-442.8641, -40.3051, -375.9587), (-491.1388, -27.2857, -376.1661), (-492.1108, -28.5164, -227.1743), (-443.8360, -41.5358, -226.9670)], 27.0),
    ([(-441.8091, -32.8047, -150.6662), (-490.0838, -19.7853, -150.8736), (-481.3124,  15.0416,   -6.2667), (-433.0377,   2.0222,   -6.0593)], 27.0),
    ([(-614.6955, 107.0061, -563.9471), (-661.6523, 124.1778, -564.3816), (-675.2888,  90.5453, -419.8690), (-628.3321,  73.3736, -419.4344)], 27.0),
    ([(-632.0966,  65.0098, -343.1585), (-679.0533,  82.1815, -343.5931), (-679.9023,  83.6308, -194.6025), (-632.9454,  66.4591, -194.1680)], 27.0),
    ([(-630.0493,  76.3055, -118.0312), (-677.0061,  93.4772, -118.4658), (-665.0132, 129.9160,   25.5113), (-618.0564, 112.7443,   25.9458)], 27.0),
    ([(-807.2899, 188.6986, -570.0942), (-852.5710, 209.9025, -570.2290), (-869.3063, 175.0791, -426.3254), (-824.0252, 153.8752, -426.1906)], 27.0),
    ([(-828.6202, 144.5465, -350.0722), (-873.9013, 165.7504, -350.2070), (-874.8473, 164.6776, -201.2139), (-829.5662, 143.4737, -201.0791)], 27.0),
    ([(-825.9389, 151.7050, -124.7820), (-871.2199, 172.9089, -124.9169), (-856.3164, 205.6550,   19.6742), (-811.0352, 184.4511,   19.8090)], 27.0),
    # Second batch: 6 cuts @ 26 mm (X ≈ 0..-48 column and X ≈ -208..-268 column)
    ([(   8.1830,  31.9508, -442.6343), ( -41.6191,  36.3853, -442.9191), ( -45.6493,   0.4078, -298.3841), (   4.1529,  -4.0267, -298.0992)], 26.0),
    ([(   2.9123, -13.0596, -221.8168), ( -46.8898,  -8.6251, -222.1016), ( -47.6333,  -7.4053,  -73.1084), (   2.1688, -11.8398,  -72.8236)], 26.0),
    ([(   2.6488,  -1.5592,    3.3093), ( -47.1533,   2.8753,    3.0245), ( -44.6497,  39.9280,  142.1533), (   5.1524,  35.4935,  142.4382)], 26.0),
    ([(-208.1528,   7.6786, -559.6337), (-257.3786,  16.4410, -559.8411), (-264.7484, -21.5557, -415.9560), (-215.5226, -30.3182, -415.7486)], 26.0),
    ([(-217.6616, -40.5334, -339.6354), (-266.8874, -31.7710, -339.8428), (-267.7484, -33.0816, -190.8510), (-218.5226, -41.8441, -190.6436)], 26.0),
    ([(-217.2644, -32.9696, -114.3429), (-266.4902, -24.2072, -114.5503), (-260.9789,  10.0618,   25.2040), (-211.7531,   1.2994,   25.4114)], 26.0),
]

def _apply_normal_cut(solid, pts, depth=27.0):
    """Cut `solid` by extruding the quad `pts` (a 4-point planar-ish profile) 27 mm along its normal.
    Reorders points to form a simple (non-self-intersecting) rectangle if needed.
    Picks the normal sign that actually reduces volume; returns the cut solid (unchanged if cut produced nothing)."""
    import math
    n = len(pts)
    cx = sum(p[0] for p in pts) / n
    cy = sum(p[1] for p in pts) / n
    cz = sum(p[2] for p in pts) / n
    cc = Vector(cx, cy, cz)
    # Build local plane from a robust estimate of the normal: cross of two diagonals.
    # This is direction-only (not handedness-aware) and is refined after reordering.
    d1 = Vector(pts[2][0]-pts[0][0], pts[2][1]-pts[0][1], pts[2][2]-pts[0][2])
    d2 = Vector(pts[3][0]-pts[1][0], pts[3][1]-pts[1][1], pts[3][2]-pts[1][2])
    nrm0 = d1.cross(d2).normalized()
    ref  = Vector(0, 0, 1) if abs(nrm0.Z) < 0.9 else Vector(1, 0, 0)
    xloc = nrm0.cross(ref).normalized()
    yloc = nrm0.cross(xloc).normalized()
    # Project points to 2D in the plane, sort by polar angle around centroid → simple polygon.
    raw_proj = [((Vector(*p) - cc).dot(xloc), (Vector(*p) - cc).dot(yloc)) for p in pts]
    order = sorted(range(n), key=lambda i: math.atan2(raw_proj[i][1], raw_proj[i][0]))
    pts    = [pts[i]      for i in order]
    proj   = [raw_proj[i] for i in order]
    # Recompute normal with Newell's formula on the reordered (simple) polygon, giving correct winding sign.
    nx = ny = nz = 0.0
    for i in range(n):
        p1, p2 = pts[i], pts[(i + 1) % n]
        nx += (p1[1] - p2[1]) * (p1[2] + p2[2])
        ny += (p1[2] - p2[2]) * (p1[0] + p2[0])
        nz += (p1[0] - p2[0]) * (p1[1] + p2[1])
    nrm  = Vector(nx, ny, nz).normalized()
    plane = Plane(origin=cc, x_dir=xloc, z_dir=nrm)
    with BuildSketch(plane, mode=Mode.PRIVATE) as sk:
        with BuildLine():
            Polyline(*proj, close=True)
        make_face()
    face = sk.sketch.face()
    v0 = solid.volume
    for signed in (depth, -depth):
        try:
            tool = Solid.extrude(face, nrm * signed)
            result = solid.cut(tool)
            if result.volume < v0 - 1e-3:
                return result
        except Exception:
            continue
    return solid  # no change if neither direction cuts material

# All solids the 9 cut profiles might pass through — checked by 3D bbox overlap.
_cut_targets = {
    'snap_fit':            snap_fit.part,
    'loft_part':           loft_part.part,
    'swept_combined':      swept_combined.part,
    'profile_part':        profile_part.part,
    'profile_part2':       profile_part2.part,
    'loft_part2':          loft_part2.part,
    'loft_feature1':       loft_feature1.part,
    'loft_feature_3prof':  loft_feature_3prof.part,
    'loft_feature_strip':  loft_feature_strip.part,
    'loft_feature_strip2': loft_feature_strip2.part,
    'loft_feature_strip3': loft_feature_strip3.part,
    'loft_feature_strip4': loft_feature_strip4.part,
    'loft_6':              loft_6,
    'loft_7':              loft_7,
    'loft_8':              loft_8,
    'loft_smooth':         loft_smooth,
}

def _bbox_overlap_3d(bb, pts, margin=5.0):
    xs = [p[0] for p in pts]; ys = [p[1] for p in pts]; zs = [p[2] for p in pts]
    return (min(xs) <= bb.max.X + margin and max(xs) >= bb.min.X - margin and
            min(ys) <= bb.max.Y + margin and max(ys) >= bb.min.Y - margin and
            min(zs) <= bb.max.Z + margin and max(zs) >= bb.min.Z - margin)

for _ci, (_cpts, _cdepth) in enumerate(_normal_cut_profiles, 1):
    _applied = []
    for _name in list(_cut_targets.keys()):
        _s = _cut_targets[_name]
        if _s is None:
            continue
        if not _bbox_overlap_3d(_s.bounding_box(), _cpts):
            continue
        _before = _s.volume
        _cut = _apply_normal_cut(_s, _cpts, depth=_cdepth)
        _after = _cut.volume
        if _after < _before - 1e-3:
            _cut_targets[_name] = _cut
            _applied.append(f"{_name} (Δ={_before - _after:.0f} mm³)")
    if _applied:
        print(f"Cut {_ci} ({_cdepth:.0f} mm) applied to: {', '.join(_applied)}")
    else:
        print(f"Cut {_ci} ({_cdepth:.0f} mm): no overlap with any target solid")

# Pull the cut versions back out for the Compound export.
loft_6      = _cut_targets['loft_6']
loft_7      = _cut_targets['loft_7']
loft_8      = _cut_targets['loft_8']
loft_smooth = _cut_targets['loft_smooth']
enclosure_final = Compound([loft_6, loft_7, loft_8])

# ── Centre-point slot cut (50 mm from body) ────────────────────────────────
# Slot on the plane Y=-12.5, centers at (-700.676, -666.663) and (-762.120, -666.663).
# Width 36.168 mm (= diameter of end caps). Cut 50 mm into body along normal (±Y).
_slot_c1 = Vector(-700.67643769, -12.5000003, -666.66252136)
_slot_c2 = Vector(-762.12016188, -12.5000003, -666.66252136)
_slot_w  = 36.16807954
_slot_r  = _slot_w / 2.0
_slot_len_vec = _slot_c1 - _slot_c2
_slot_L  = _slot_len_vec.length
_slot_dir = _slot_len_vec.normalized()  # direction from c2 to c1 in the slot plane
_slot_plane = Plane(origin=(_slot_c1 + _slot_c2) * 0.5, x_dir=_slot_dir, z_dir=Vector(0, 1, 0))
with BuildSketch(_slot_plane, mode=Mode.PRIVATE) as _slot_sk:
    SlotCenterToCenter(center_separation=_slot_L, height=_slot_w)
_slot_face = _slot_sk.sketch.face()

def _apply_axis_cut(targets, face, normal_vec, depth):
    """Try +normal*depth, then -normal*depth; apply cut to every target it reduces."""
    tool_plus  = Solid.extrude(face, normal_vec * depth)
    tool_minus = Solid.extrude(face, normal_vec * (-depth))
    hits = []
    for name in list(targets.keys()):
        s = targets[name]
        if s is None: continue
        best_vol = s.volume; best_result = None
        for tool in (tool_plus, tool_minus):
            try:
                r = s.cut(tool)
                if r.volume < best_vol - 1e-3:
                    best_vol = r.volume; best_result = r
            except Exception:
                continue
        if best_result is not None:
            targets[name] = best_result
            hits.append(f"{name} (Δ={s.volume - best_vol:.0f} mm³)")
    return hits

_cut_targets['loft_6'] = loft_6
_cut_targets['loft_7'] = loft_7
_cut_targets['loft_8'] = loft_8
_cut_targets['loft_smooth'] = loft_smooth
_cut_targets['loft_tail'] = loft_tail

_slot_hits = _apply_axis_cut(_cut_targets, _slot_face, Vector(0, 1, 0), 50.0)
print(f"Slot cut (50 mm) applied to: {', '.join(_slot_hits) if _slot_hits else 'no overlapping solid'}")

# ── Circle cut (Ø52 mm, 50 mm from body) ───────────────────────────────────
_circ_center = Vector(-890.35685252, -16.00000226, -666.66252136)
_circ_r      = 51.99998745 / 2.0
_circ_plane  = Plane(origin=_circ_center, x_dir=Vector(1, 0, 0), z_dir=Vector(0, 1, 0))
with BuildSketch(_circ_plane, mode=Mode.PRIVATE) as _circ_sk:
    Circle(radius=_circ_r)
_circ_face = _circ_sk.sketch.face()
_circ_hits = _apply_axis_cut(_cut_targets, _circ_face, Vector(0, 1, 0), 50.0)
print(f"Circle cut (Ø52, 50 mm) applied to: {', '.join(_circ_hits) if _circ_hits else 'no overlapping solid'}")

# Propagate cut versions back to the named variables used below.
loft_6      = _cut_targets['loft_6']
loft_7      = _cut_targets['loft_7']
loft_8      = _cut_targets['loft_8']
loft_smooth = _cut_targets['loft_smooth']
loft_tail   = _cut_targets['loft_tail']
enclosure_final = Compound([loft_6, loft_7, loft_8])

# ── 6-point polygon extrusion (+170 mm along +Z) ────────────────────────────
# User requested: join points in given order. Given order is self-intersecting
# (bowtie) — edge 1→2 crosses edge 5→0. Fall back to polar-sort reordering if
# the given order fails to form a valid face.
_poly_pts = [
    (-555.4731, 306.4489,  99.1526),
    (-574.5690, 299.7430, 102.3692),
    (-569.2477, 281.9537,  96.8730),
    (-544.5531, 269.9433,  87.8738),
    (-590.9301, 253.1042,  95.5565),
    (-596.8796, 272.2500, 101.5275),
]
# Build best-fit plane for the (planar) polygon.
_pcx = sum(p[0] for p in _poly_pts) / 6
_pcy = sum(p[1] for p in _poly_pts) / 6
_pcz = sum(p[2] for p in _poly_pts) / 6
_pc  = Vector(_pcx, _pcy, _pcz)
# Newell's normal (works even if ordering is a bowtie — we handle that below).
_pnx = _pny = _pnz = 0.0
for _i in range(6):
    _p1, _p2 = _poly_pts[_i], _poly_pts[(_i + 1) % 6]
    _pnx += (_p1[1] - _p2[1]) * (_p1[2] + _p2[2])
    _pny += (_p1[2] - _p2[2]) * (_p1[0] + _p2[0])
    _pnz += (_p1[0] - _p2[0]) * (_p1[1] + _p2[1])
_pnrm = Vector(_pnx, _pny, _pnz).normalized()
_pref = Vector(0, 0, 1) if abs(_pnrm.Z) < 0.9 else Vector(1, 0, 0)
_pxloc = _pnrm.cross(_pref).normalized()
_pyloc = _pnrm.cross(_pxloc).normalized()
_pplane = Plane(origin=_pc, x_dir=_pxloc, z_dir=_pnrm)
_pproj = [((Vector(*p) - _pc).dot(_pxloc), (Vector(*p) - _pc).dot(_pyloc)) for p in _poly_pts]

poly_extrude = None
for _attempt_pts in (_pproj, sorted(_pproj, key=lambda q: math.atan2(q[1], q[0]))):
    try:
        with BuildSketch(_pplane, mode=Mode.PRIVATE) as _psk:
            with BuildLine():
                Polyline(*_attempt_pts, close=True)
            make_face()
        _pface = _psk.sketch.face()
        if _pface.area > 1e-6:
            poly_extrude = Solid.extrude(_pface, Vector(0, 0, 170.0))
            if _attempt_pts is not _pproj:
                print("Polygon reordered (given order was self-intersecting).")
            break
    except Exception as _e:
        continue

if poly_extrude is not None:
    print(f"Polygon extrusion (+Z 170 mm): is_manifold={poly_extrude.is_manifold}  volume={poly_extrude.volume:.2f} mm³")
else:
    print("Polygon extrusion failed.")
   # ── New Feature — profile extrude 7.9mm ──────────────────
_new_pts = [
    (-150.5693, 0.0, -664.7693),
    (-50.551,   0.0, -654.9635),
    (18.7428,   0.0, -649.0801),
    (36.3832,   0.0, -647.2204),
    (56.4872,   0.0, -641.4764),
    (70.8471,   0.0, -631.0141),
    (80.8307,   0.0, -615.8221),
    (84.4286,   0.0, -599.032),
    (82.4329,   0.0, -576.7895),
    (79.9548,   0.0, -546.1238),
    (11.3009,   0.0, -546.1238),
    (1.7563,    0.0, -548.3945),
    (-6.3784,   0.0, -552.9499),
    (-13.8623,  0.0, -560.1084),
    (-21.0208,  0.0, -566.9415),
    (-28.5047,  0.0, -576.0523),
    (-42.8216,  0.0, -595.2501),
    (-56.1625,  0.0, -611.5194),
    (-71.1302,  0.0, -627.7887),
    (-87.7249,  0.0, -642.7565),
    (-100.9998, 0.0, -650.7012),
    (-116.1137, 0.0, -657.2431),
    (-132.1299, 0.0, -662.4314),
]
_new_vecs   = [Vector(x, y, z) for x, y, z in _new_pts]
_new_wire   = Wire.make_polygon(_new_vecs, close=True)
_new_face   = Face(_new_wire)
new_feature = Solid.extrude(_new_face, Vector(0, -79, 0))
print(f"New Feature DONE | Volume: {new_feature.volume:.4e} mm3")
_bb_nf = new_feature.bounding_box()
print(f"new_feature: X:{_bb_nf.min.X:.1f}→{_bb_nf.max.X:.1f}  Y:{_bb_nf.min.Y:.1f}→{_bb_nf.max.Y:.1f}  Z:{_bb_nf.min.Z:.1f}→{_bb_nf.max.Z:.1f}")
hole_pts_new = [
    (-451.5554, -34.0132, -693.9643),
    (-4.9532,   -33.7006, -651.1039),
]
for x, y, z in hole_pts_new:
    print(f"Point: X:{x:.1f} Y:{y:.1f} Z:{z:.1f}")
    print(f"  X ok: {_bb_nf.min.X <= x <= _bb_nf.max.X}")
    print(f"  Y ok: {_bb_nf.min.Y <= y <= _bb_nf.max.Y}")
    print(f"  Z ok: {_bb_nf.min.Z <= z <= _bb_nf.max.Z}")
# ── Hole cut — 55mm dia, 63mm deep, normal direction ──────────────────
import numpy as np

hole_pts_new = [
    (-451.5554, -34.0132, -693.9643),
    (-4.9532,   -33.7006, -651.1039),
]

# SVD plane from 2 points
_hp = np.array([[x, y, z] for x, y, z in hole_pts_new])
_hc = np.mean(_hp, axis=0)
_, _, _hVt = np.linalg.svd(_hp - _hc)
_hn = _hVt[-1] / np.linalg.norm(_hVt[-1])

for i, (x, y, z) in enumerate(hole_pts_new):
    _hole_plane = Plane(
        origin=(x, y, z),
        z_dir=Vector(float(_hn[0]), float(_hn[1]), float(_hn[2]))
    )
    _hole_face = Face(Wire(Edge.make_circle(15, _hole_plane)))  # r=27.5 = 55mm dia
    _hole_tool = Solid.extrude(
        _hole_face.moved(Location(Vector(
            float(-_hn[0] * 31.5),
            float(-_hn[1] * 31.5),
            float(-_hn[2] * 31.5),
        ))),
        Vector(float(_hn[0] * 63.0),
               float(_hn[1] * 63.0),
               float(_hn[2] * 63.0))
    )

    new_feature = new_feature.cut(_hole_tool)
    try:    new_feature = new_feature.solids()[0]
    except: pass
    print(f"Hole {i+1} DONE | Volume: {new_feature.volume:.4e} mm3")


_v1 = _cut_targets['loft_feature_strip'].volume
_v2 = _cut_targets['loft_feature_strip2'].volume
print(f"Before cut — strip: {_v1:.0f}  strip2: {_v2:.0f}")

_cut_targets['loft_feature_strip2'] = _cut_targets['loft_feature_strip2'].cut(cyl_part.part)
_cut_targets['loft_feature_strip'] = _cut_targets['loft_feature_strip'].cut(cyl_part.part)

_v1b = _cut_targets['loft_feature_strip'].volume
_v2b = _cut_targets['loft_feature_strip2'].volume
print(f"After cut  — strip: {_v1b:.0f} (Δ={_v1-_v1b:.0f})  strip2: {_v2b:.0f} (Δ={_v2-_v2b:.0f})")
# ── Boss cylinder — SUBTRACT (boolean cut) ────────────────────────────
with BuildPart() as cyl_boss:
    with Locations(Location(cyl_main["pos"], cyl_main["rot"])):
        Cylinder(
            radius=cyl_main["id"] / 2,   # 205.3mm
            height=450,
            mode=Mode.ADD                 # pehle solid banao
        )

# Ab isse saare parts cut karo
_parts_to_cut = [
    'snap_fit', 'loft_part', 'swept_combined', 'profile_part',
    'profile_part2', 'loft_part2', 'loft_feature1', 'loft_feature_3prof',
    'loft_feature_strip', 'loft_feature_strip2', 'loft_feature_strip3',
    'loft_feature_strip4', 'loft_smooth', 'loft_tail',
]
for _pname in _parts_to_cut:
    _v1 = _cut_targets[_pname].volume
    _cut_targets[_pname] = _cut_targets[_pname].cut(cyl_boss.part)
    _delta = _v1 - _cut_targets[_pname].volume
    if _delta > 1:
        print(f"CUT HIT: {_pname}  delta={_delta:.0f} mm3")
    else:
        print(f"no cut : {_pname}")

_v1 = enclosure_final.volume
enclosure_final = enclosure_final.cut(cyl_boss.part)
_delta = _v1 - enclosure_final.volume
print(f"{'CUT HIT' if _delta > 1 else 'no cut'}: enclosure_final  delta={_delta:.0f}")

_v1 = loft_connector.volume
loft_connector = loft_connector.cut(cyl_boss.part)
_delta = _v1 - loft_connector.volume
print(f"{'CUT HIT' if _delta > 1 else 'no cut'}: loft_connector  delta={_delta:.0f}")

_v1 = loft_new2.volume
loft_new2 = loft_new2.cut(cyl_boss.part)
_delta = _v1 - loft_new2.volume
print(f"{'CUT HIT' if _delta > 1 else 'no cut'}: loft_new2  delta={_delta:.0f}")

print("cyl_boss boolean cut done")
# ── New boundary cut (XZ plane, 100mm Y extrude) ──────────────────────
_cut3_pts = [
    (-1040.6867, 0.0, -44.132),
    (-1040.6867, 0.0, -202.8018),
    (-981.6076,  0.0, -202.8018),
    (-841.5055,  0.0, -96.4592),
    (-971.4797,  0.0,  29.295),
]

_cut3_face = _make_planar_cap(_cut3_pts)
_cut3_tool_pos = Solid(BRepPrimAPI_MakePrism(_cut3_face.wrapped, gp_Vec(0, 100, 0)).Shape())
_cut3_tool_neg = Solid(BRepPrimAPI_MakePrism(_cut3_face.wrapped, gp_Vec(0, -100, 0)).Shape())

for _pname in [
    'snap_fit', 'loft_part', 'swept_combined', 'profile_part',
    'profile_part2', 'loft_part2', 'loft_feature1', 'loft_feature_3prof',
    'loft_feature_strip', 'loft_feature_strip2', 'loft_feature_strip3',
    'loft_feature_strip4', 'loft_smooth', 'loft_tail',
]:
    _v1 = _cut_targets[_pname].volume
    _cut_targets[_pname] = _cut_targets[_pname].cut(_cut3_tool_pos).cut(_cut3_tool_neg)
    _delta = _v1 - _cut_targets[_pname].volume
    if _delta > 1:
        print(f"CUT HIT: {_pname}  delta={_delta:.0f} mm3")
    else:
        print(f"no cut : {_pname}")

enclosure_final = enclosure_final.cut(_cut3_tool_pos).cut(_cut3_tool_neg)
loft_connector = loft_connector.cut(_cut3_tool_pos).cut(_cut3_tool_neg)
loft_new2 = loft_new2.cut(_cut3_tool_pos).cut(_cut3_tool_neg)
print("cut3 XZ plane done")
# ── New boss (XZ plane, 80.662mm negative Y only) ─────────────────────
_boss_pts = [
    (-939.9901, 0.0, -118.3304),
    (-927.833,  0.0, -121.1359),
    (-916.611,  0.0, -126.2793),
    (-908.6621, 0.0, -130.9551),
    (-901.1808, 0.0, -138.4365),
    (-896.0374, 0.0, -146.8529),
    (-892.2967, 0.0, -155.737),
    (-889.9588, 0.0, -164.6211),
    (-889.9588, 0.0, -172.57),
    (-891.8291, 0.0, -181.4541),
    (-894.167,  0.0, -189.403),
    (-899.778,  0.0, -199.6898),
    (-909.5973, 0.0, -209.9766),
    (-921.7544, 0.0, -216.0552),
    (-931.9539, 0.0, -219.068),
    (-944.359,  0.0, -220.6516),
    (-955.4444, 0.0, -223.0271),
    (-962.8347, 0.0, -225.1386),
    (-971.8085, 0.0, -230.4173),
    (-977.3513, 0.0, -234.6403),
    (-983.9497, 0.0, -240.9749),
    (-988.1727, 0.0, -247.0454),
    (-991.34,   0.0, -254.6996),
    (-992.9236, 0.0, -263.4096),
    (-991.6028, 0.0, -203.1372),
    (-1031.7909,0.0, -203.1372),
    (-1027.0849,0.0, -135.5906),
    (-1003.2106,0.0, -104.3819),
]

_boss_face = _make_planar_cap(_boss_pts)
_boss_solid = Solid(BRepPrimAPI_MakePrism(
    _boss_face.wrapped,
    gp_Vec(0, -80.662, 0)
).Shape())

print(f"Boss volume: {_boss_solid.volume:.2f} mm3")
# ── Circle cut (Ø58mm, XZ plane, -Y direction 90mm) ───────────────────
import math

_circle_center = (-939.2051, 0.0, -170.7584)
_circle_radius = 29.0   # 58mm diameter / 2
_n_pts = 64

_circle_pts = [
    (
        _circle_center[0] + _circle_radius * math.cos(2 * math.pi * i / _n_pts),
        0.0,
        _circle_center[2] + _circle_radius * math.sin(2 * math.pi * i / _n_pts),
    )
    for i in range(_n_pts)
]

_circle_face = _make_planar_cap(_circle_pts)
_circle_cut_tool = Solid(BRepPrimAPI_MakePrism(
    _circle_face.wrapped,
    gp_Vec(0, -90, 0)
).Shape())

print(f"Circle cut tool volume: {_circle_cut_tool.volume:.2f} mm3")

# Cut on boss
_v1 = _boss_solid.volume
_boss_solid = _boss_solid.cut(_circle_cut_tool)
print(f"Boss cut: {_v1:.0f} -> {_boss_solid.volume:.0f}  delta={_v1 - _boss_solid.volume:.0f}")

# ── Helper function — define once ─────────────────────────────────────
def _apply_cut(obj, tool, name):
    try:
        v1 = obj.volume
        result = obj.cut(tool)
        out = result if hasattr(result, 'volume') else Compound(list(result))
        delta = v1 - out.volume
        if delta > 1:
            print(f"CUT HIT: {name}  delta={delta:.0f} mm3")
        else:
            print(f"no cut : {name}")
        return out
    except Exception as e:
        print(f"skip: {name} ({e})")
        return obj
# ── XZ plane cut5 (-Y 90mm) ───────────────────────────────────────────
from OCP.BRepBuilderAPI import BRepBuilderAPI_MakeFace, BRepBuilderAPI_MakePolygon

_cut5_pts = [
    (-947.0513,  0.0, -253.9977),
    (-961.3773,  0.0, -257.8533),
    (-973.7116,  0.0, -264.2672),
    (-983.579,   0.0, -272.1611),
    (-992.9531,  0.0, -282.5219),
    (-999.8603,  0.0, -291.896),
    (-1006.7675, 0.0, -301.7634),
    (-1012.1945, 0.0, -313.1109),
    (-1016.1415, 0.0, -326.9253),
    (-1019.1017, 0.0, -338.7662),
    (-1022.062,  0.0, -352.5806),
    (-1022.062,  0.0, -367.8752),
    (-1024.5288, 0.0, -384.6498),
    (-1030.9427, 0.0, -331.3657),
    (-1030.9427, 0.0, -274.8534),
    (-1043.5794, 0.0, -79.8203),
    (-862.9931,  0.0, -99.5206),
    (-878.0967,  0.0, -241.3629),
]

_poly5 = BRepBuilderAPI_MakePolygon()
for _p in _cut5_pts:
    _poly5.Add(gp_Pnt(_p[0], _p[1], _p[2]))
_poly5.Close()
_cut5_face_occ = BRepBuilderAPI_MakeFace(_poly5.Wire(), True).Face()

_cut5_tool = Solid(BRepPrimAPI_MakePrism(
    _cut5_face_occ,
    gp_Vec(0, -90, 0)
).Shape())
print(f"Cut5 tool volume: {_cut5_tool.volume:.2f} mm3")

for _pname in [
    'snap_fit', 'loft_part', 'swept_combined', 'profile_part',
    'profile_part2', 'loft_part2', 'loft_feature1', 'loft_feature_3prof',
    'loft_feature_strip', 'loft_feature_strip2', 'loft_feature_strip3',
    'loft_feature_strip4', 'loft_smooth', 'loft_tail',
]:
    _v1 = _cut_targets[_pname].volume
    _result = _cut_targets[_pname].cut(_cut5_tool)
    _cut_targets[_pname] = _result if hasattr(_result, 'volume') else Compound(list(_result))
    _delta = _v1 - _cut_targets[_pname].volume
    if _delta > 1:
        print(f"CUT HIT: {_pname}  delta={_delta:.0f} mm3")
    else:
        print(f"no cut : {_pname}")

enclosure_final = enclosure_final.cut(_cut5_tool)
loft_connector = loft_connector.cut(_cut5_tool)
loft_new2 = loft_new2.cut(_cut5_tool)
print("cut5 done")
# ── XZ plane cut6 (-Y 120mm) ──────────────────────────────────────────
from OCP.BRepBuilderAPI import BRepBuilderAPI_MakeFace, BRepBuilderAPI_MakePolygon

_cut6_pts = [
    (-1028.7479, 0.0, -78.6026),
    (-873.9604,  0.0, -130.1984),
    (-886.4146,  0.0, -227.163),
    (-926.4458,  0.0, -237.838),
    (-960.25,    0.0, -247.6234),
    (-995.8333,  0.0, -264.5255),
    (-1062.552,  0.0, -303.6671),
    (-1062.552,  0.0, -216.488),
    (-1047.4291, 0.0, -115.0755),
]

_poly6 = BRepBuilderAPI_MakePolygon()
for _p in _cut6_pts:
    _poly6.Add(gp_Pnt(_p[0], _p[1], _p[2]))
_poly6.Close()
_cut6_face_occ = BRepBuilderAPI_MakeFace(_poly6.Wire(), True).Face()
_cut6_tool = Solid(BRepPrimAPI_MakePrism(_cut6_face_occ, gp_Vec(0, -120, 0)).Shape())
print(f"Cut6 tool volume: {_cut6_tool.volume:.2f} mm3")

for pname in list(_cut_targets.keys()):
    _cut_targets[pname] = _apply_cut(_cut_targets[pname], _cut6_tool, pname)

enclosure_final = _apply_cut(enclosure_final, _cut6_tool, 'enclosure_final')
loft_connector  = _apply_cut(loft_connector,  _cut6_tool, 'loft_connector')
loft_new2       = _apply_cut(loft_new2,       _cut6_tool, 'loft_new2')
_boss_solid     = _apply_cut(_boss_solid,      _cut6_tool, 'boss_solid')

try:
    poly_extrude = _apply_cut(poly_extrude, _cut6_tool, 'poly_extrude')
except Exception:
    pass

_ep_solids = [
    _apply_cut(ep, _cut6_tool, f'ep_solid[{i}]')
    for i, ep in enumerate(_ep_solids)
]

print("cut6 done")
# ── XZ plane cut7 (-Y 100mm) ──────────────────────────────────────────
_cut7_pts = [
    (-993.0437,  0.0, -229.0185),
    (-1006.6058, 0.0, -267.1381),
    (-1020.5801, 0.0, -314.8436),
    (-1028.2901, 0.0, -357.2485),
    (-1034.0726, 0.0, -407.8453),
    (-1062.3262, 0.0, -299.144),
]

_poly7 = BRepBuilderAPI_MakePolygon()
for _p in _cut7_pts:
    _poly7.Add(gp_Pnt(_p[0], _p[1], _p[2]))
_poly7.Close()
_cut7_face_occ = BRepBuilderAPI_MakeFace(_poly7.Wire(), True).Face()
_cut7_tool = Solid(BRepPrimAPI_MakePrism(_cut7_face_occ, gp_Vec(0, -100, 0)).Shape())
print(f"Cut7 tool volume: {_cut7_tool.volume:.2f} mm3")

for pname in list(_cut_targets.keys()):
    _cut_targets[pname] = _apply_cut(_cut_targets[pname], _cut7_tool, pname)

enclosure_final = _apply_cut(enclosure_final, _cut7_tool, 'enclosure_final')
loft_connector  = _apply_cut(loft_connector,  _cut7_tool, 'loft_connector')
loft_new2       = _apply_cut(loft_new2,       _cut7_tool, 'loft_new2')
_boss_solid     = _apply_cut(_boss_solid,      _cut7_tool, 'boss_solid')

try:
    poly_extrude = _apply_cut(poly_extrude, _cut7_tool, 'poly_extrude')
except Exception:
    pass

_ep_solids = [
    _apply_cut(ep, _cut7_tool, f'ep_solid[{i}]')
    for i, ep in enumerate(_ep_solids)
]

print("cut7 done")



if True and len(_ep_solids) == 7:
    combined = Compound([
        _cut_targets['snap_fit'],
        _cut_targets['loft_part'],
        _cut_targets['swept_combined'],
        _cut_targets['profile_part'],
        _cut_targets['profile_part2'],
        _cut_targets['loft_part2'],
        _cut_targets['loft_feature1'],
        _cut_targets['loft_feature_3prof'],
        _cut_targets['loft_feature_strip'],
        _cut_targets['loft_feature_strip2'],
        _cut_targets['loft_feature_strip3'],
        _cut_targets['loft_feature_strip4'],
        enclosure_final,
        _cut_targets['loft_smooth'],
        _cut_targets['loft_tail'],
        cyl_part.part,
        loft_connector,
        loft_new2,
        _boss_solid,
        *([poly_extrude] if poly_extrude is not None else []),
        *_ep_solids,
    ])
_v_before = _cut_targets['loft_feature_strip2'].volume
_cut_targets['loft_feature_strip2'] = _cut_targets['loft_feature_strip2'].cut(cyl_part.part)
_v_after = _cut_targets['loft_feature_strip2'].volume
print(f"strip2 cut: {_v_before:.0f} → {_v_after:.0f}  Δ={_v_before - _v_after:.0f}")

_v_before2 = _cut_targets['loft_feature_strip'].volume
_cut_targets['loft_feature_strip'] = _cut_targets['loft_feature_strip'].cut(cyl_part.part)
_v_after2 = _cut_targets['loft_feature_strip'].volume
print(f"strip cut:  {_v_before2:.0f} → {_v_after2:.0f}  Δ={_v_before2 - _v_after2:.0f}")
# ── Inner diameter boss (visible solid) ───────────────────────────────
with BuildPart() as cyl_part:
    with Locations(Location(cyl_main["pos"], cyl_main["rot"])):
        Cylinder(radius=cyl_main["od"] / 2, height=cyl_main["height"])
        Cylinder(radius=cyl_main["id"] / 2, height=cyl_main["height"], mode=Mode.SUBTRACT)
    with Locations(Location(cyl_sub["pos"], cyl_sub["rot"])):
        Cylinder(radius=cyl_sub["od"] / 2, height=cyl_sub["height"], mode=Mode.SUBTRACT)




total_volume = sum(
        p.volume for p in [
            _cut_targets['snap_fit'],
            _cut_targets['loft_part'],
            _cut_targets['swept_combined'],
            _cut_targets['profile_part'],
            _cut_targets['profile_part2'],
            _cut_targets['loft_part2'],
            _cut_targets['loft_feature1'],
            _cut_targets['loft_feature_3prof'],
            _cut_targets['loft_feature_strip'],
            _cut_targets['loft_feature_strip2'],
            _cut_targets['loft_feature_strip3'],
            _cut_targets['loft_feature_strip4'],
            loft_6, loft_7, loft_8,
            loft_smooth, loft_tail,
            new_feature,
            cyl_part.part,
            loft_connector,
            loft_new2,
            cyl_boss.part,
            _boss_solid,
            *_ep_solids,
            poly_extrude if poly_extrude else None,
        ] if p is not None
    )
    

print(f"\nTOTAL VOLUME : {total_volume * 1000:.4e} mm3")
print(f"TOTAL VOLUME : {total_volume * 1000 / 1000:.4f} cm3")

export_step(combined, "/Users/softage/Desktop/3x5+3_low_clipped.step")
export_stl(combined,  "/Users/softage/Desktop/3x5+3_low_clipped.stl")
print("Combined export DONE")

try:
    from ocp_vscode import show, Camera
    show(combined, reset_camera=Camera.RESET, transparent=True)
    print("Preview opened")
except ImportError:
    print("ocp_vscode not found")